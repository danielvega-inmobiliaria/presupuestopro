"""
Exportar usuarios a contactar (retención) — agregado 15/07/2026, pedido de
Daniel. Extendido varias veces el mismo mes:
  - 15/07/2026: Segmento C ("validado sin actividad") + hoja "Todos los usuarios".
  - 20/07/2026: Segmento D ("validado, solo usó Costo/m², nunca presupuestó")
    + columna "Segmento" en la hoja "Todos los usuarios". Antes, un usuario
    validado que solo había consultado Costo/m² (0 presupuestos, 0
    borradores, 1+ costo/m²) no caía en ningún segmento — ni B (pedía
    presup+borr==1) ni C (pedía costo/m²==0) — y quedaba invisible para la
    campaña de retención. Daniel lo pidió explícitamente: la lista tiene que
    cubrir a TODOS los usuarios, no dejar a nadie afuera de algún segmento.

Antes esto se armaba a mano: entrar a Admin > Usuarios, sacar capturas de
pantalla de la tabla y transcribir fila por fila a un Excel (lento y con
riesgo de error, sobre todo en teléfonos y contadores que se leían muy
chicos en la captura). Ahora se arma directo desde la base, en un click, con
la MISMA lógica que ya usa la tabla de Admin > Usuarios.

Segmentación (cubre el 100% de los usuarios is_admin=0, sin huecos):

  Segmento A — "Sin validar (email)": u.email_verificado = 0.
  (Se usa email_verificado y no el badge visual de la tabla porque el badge
  no se muestra cuando u.metodo_verificacion está vacío — cuentas viejas
  "grandfatheradas", ver utils/verificacion.py::get_verificacion_status.
  email_verificado sigue siendo la señal real de si activó la cuenta por
  mail o no, esté vacío el método o no.)

  Entre los validados (email_verificado = 1):
    Segmento B — hizo exactamente 1 presupuesto o borrador en total
      (n_presupuestos + n_borradores) == 1.
    Segmento C — CERO actividad de cualquier tipo: 0 presupuestos,
      0 borradores, 0 consultas de costo/m².
    Segmento D — nunca presupuestó (0 presupuestos, 0 borradores) pero SÍ
      usó la calculadora de Costo/m² al menos una vez. Es el hueco que
      había antes de 20/07/2026.
    Sin segmento de retención — 2 o más presupuestos/borradores en total:
      ya es un usuario activo de la app, no es el público de esta campaña,
      pero igual aparece listado en "Todos los usuarios" (con esa etiqueta)
      para que Daniel tenga el panorama completo.
"""
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.email_tracking import ETIQUETAS_EVENTO

APP_URL = 'https://web-production-0c9c1.up.railway.app/login'

# 21/07/2026: los emails de retención salen de noreply@presupuestopro.com.ar.
# Para que el usuario tenga cómo contestar, cada mensaje de email suma esta
# línea con el link de WhatsApp al final (además, si responde por WhatsApp
# se abre la ventana de 24h para mensajes libres desde el 2009).
# Actualizado 24/07/2026: desde ahora el mail SÍ se puede responder --
# `reply_to` en admin.py::seguimiento_email lo manda a
# contacto@presupuestopro.com.ar (Cloudflare Email Routing, ya armado y
# activo, reenvía a presupuestopro.app@gmail.com) -- se saca el aviso de
# "no respondas, no lo lee nadie" que ya no es cierto, y se deja como una
# segunda opción además de WhatsApp.
WA_LINK = 'https://wa.me/5493417542009'
WA_CTA = (f"\n\nPodés responder este mail directamente, o si preferís, "
          f"escribinos por WhatsApp: {WA_LINK}")

COLS_RETENCION = ["Email retención: último envío (ART)", "WhatsApp retención: último envío (ART)"]

HEADERS_SEGMENTO = ["Nombre", "Email", "Teléfono", "Ciudad", "Provincia", "País",
                     "Presup.", "Borr.", "Costo/m²", "Estado activación", "Creado", "Vence",
                     "Mensaje WhatsApp sugerido", "Mensaje email sugerido", "Comentarios",
                     "Mail: estado", "Mail: fecha"] + COLS_RETENCION

HEADERS_TODOS = ["Nombre", "Email", "Teléfono", "Ciudad", "Provincia", "País",
                  "Presup.", "Borr.", "Costo/m²", "Estado activación", "Segmento",
                  "Creado", "Vence", "Comentarios", "Mail: estado", "Mail: fecha"] + COLS_RETENCION

# Agregado 03/08/2026, pedido de Daniel: columna "Comentarios" para anotar,
# llamada por llamada, la causa del uso escaso/nulo -- en TODAS las hojas
# (no solo Vencidos/Abonados, que fue la 1ra versión: Daniel aclaró que
# también contacta gente de "Todos los usuarios" y de los segmentos A/B/C/D,
# y quiere que el comentario se mantenga igual sea cual sea la hoja/segmento
# desde donde contactó). Se guarda siempre en el mismo campo de la base
# (users.comentario_seguimiento) -- no hay una columna por hoja, es UN
# comentario por usuario que se muestra en cualquier hoja donde ese usuario
# aparezca. Ver generar_excel_usuarios_a_contactar() para el criterio exacto
# de Vencidos/Abonados (mismo que ya usa el resto de la app --
# subscription_expires/es_trial/active -- en routes/pagos.py y el dashboard
# de Admin) y routes/admin.py::_importar_comentarios_xlsx() para cómo
# vuelven los comentarios anotados a la base para la próxima exportación
# (lee CUALQUIER hoja que tenga columnas Email + Comentarios, no una lista
# fija de nombres de hoja).
HEADERS_VENCIDOS = ["Nombre", "Email", "Teléfono", "Ciudad", "Provincia", "País",
                     "Presup.", "Borr.", "Costo/m²", "Tipo", "Venció el", "Comentarios",
                     "Mail: estado", "Mail: fecha"] + COLS_RETENCION

# 06/08/2026, pedido de Daniel: 2 columnas nuevas en Abonados para ver de un
# vistazo cuánto ingresa realmente por cada suscriptor. "Plan / $mes" muestra
# el plan actual (Mensual/Trimestral/Semestral/Anual) y su precio de lista
# por mes. "$ cobrados" es ESE precio mensual ya descontado el 14% que se
# queda Mercado Pago -- normalizado a base mensual para que las 4 duraciones
# sean comparables entre sí (no el monto real del último pago, que varía
# según cuántos meses pagó de una vez). Ver _escribir_hoja_abonados() para el
# cálculo y el total que va arriba de la columna.
HEADERS_ABONADOS = ["Nombre", "Email", "Teléfono", "Ciudad", "Provincia", "País",
                     "Presup.", "Borr.", "Costo/m²", "Plan / $mes", "$ cobrados (mensual, neto MP 14%)",
                     "Abonado desde", "Vence", "Comentarios", "Mail: estado", "Mail: fecha"] + COLS_RETENCION

MP_COMISION_PCT = 14  # % que se queda Mercado Pago (Checkout Pro) -- confirmado por Daniel 06/08/2026

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SEG_A = "A - Sin validar email"
SEG_B = "B - 1 presup./borrador en total"
SEG_C = "C - Validado, cero actividad"
SEG_D = "D - Validado, solo usó Costo/m2"
SEG_ACTIVO = "Activo (2+ presup./borr.) - no prioritario"


def _art(fecha_utc):
    """UTC (como se guarda created_at en SQLite) -> ART (UTC-3, sin horario
    de verano), formateado DD/MM/YYYY HH:MM. Mismo offset fijo que ya usa el
    filtro Jinja local_dt (app.py) -- acá se repite en Python plano porque
    esto arma un .xlsx, no un template."""
    if not fecha_utc:
        return ''
    try:
        dt = datetime.fromisoformat(str(fecha_utc).replace(' ', 'T'))
    except ValueError:
        return ''
    return (dt - timedelta(hours=3)).strftime('%d/%m/%Y %H:%M')


def _retencion_cols(u):
    """[último email de retención (ART), último WhatsApp de retención (ART)]
    -- agregado 07/08/2026, pedido de Daniel. A diferencia de _mail_cols()
    (que sale de email_eventos, TODOS los mails sin distinguir tipo --
    bienvenida, recordatorio de inactividad, retención, etc.), esto sale de
    retencion_contactos filtrado por canal (ver subqueries
    ultimo_email_retencion/ultimo_whatsapp_retencion en
    routes/admin.py::_usuarios_para_exportar) -- specific al mensaje de
    retención de Admin > Seguimiento (manual o automático), y solo cuenta
    envíos que salieron OK (resultado='ok'), no los que fallaron."""
    ue = u['ultimo_email_retencion'] if 'ultimo_email_retencion' in u.keys() else None
    uw = u['ultimo_whatsapp_retencion'] if 'ultimo_whatsapp_retencion' in u.keys() else None
    return [_art(ue), _art(uw)]


def _mail_cols(u):
    """[estado del último mail mandado, fecha] listo para pegar al final de
    una fila -- agregado 04/08/2026 (cont. 20, pedido de Daniel: quiere ver
    desde la App si los mails llegan/se abren, sin depender del dashboard de
    Resend). Sale de users.mail_estado/mail_estado_fecha (subquery
    SQL_MAIL_ESTADO en routes/admin.py -- ver utils/email_tracking.py). Si
    todavía no llegó ningún webhook de Resend para ese email (no está
    configurado en Resend, o recién se mandó y no hay respuesta todavía),
    queda vacío -- no es un error, es "sin datos todavía"."""
    estado = u['mail_estado'] if 'mail_estado' in u.keys() else None
    fecha = u['mail_estado_fecha'] if 'mail_estado_fecha' in u.keys() else None
    return [ETIQUETAS_EVENTO.get(estado, estado) if estado else '', (fecha or '')[:10]]


def _segmento(u):
    if not u['email_verificado']:
        return SEG_A
    total = u['n_presupuestos'] + u['n_borradores']
    if total == 1:
        return SEG_B
    if total == 0 and u['n_costo_m2'] == 0:
        return SEG_C
    if total == 0 and u['n_costo_m2'] > 0:
        return SEG_D
    return SEG_ACTIVO


def _mensaje_activacion(nombre):
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Somos de PresupuestoPRO. Vimos que te registraste pero "
          f"todavía no activaste tu cuenta por mail. Si te trabó algo del proceso "
          f"contanos, te ayudamos. Y si preferís, te reenviamos el mail de "
          f"activación. Ingresá en: {APP_URL}")
    # Actualizado 05/08/2026: mismo enfoque "bondades + link" que la plantilla
    # de WhatsApp retencion_activar_cuenta_promo, en vez del enfoque viejo de
    # "¿tuviste una dificultad?".
    email = (f"Hola {nombre}, con PresupuestoPRO armás el presupuesto completo de tu "
             f"obra —materiales, mano de obra y costo por m²— en minutos, con precios "
             f"siempre actualizados contra corralones reales. Para empezar a usarla, "
             f"solo te falta activar tu cuenta. Entrá acá: "
             f"https://presupuestopro.com.ar/login" + WA_CTA)
    return wa, email


def _mensaje_seguimiento(nombre):
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Vimos que hiciste tu primer presupuesto en PresupuestoPRO. "
          f"¿Qué te pareció? ¿Tuviste alguna dificultad usando la app? Nos ayuda "
          f"mucho tu opinión, y si necesitás una mano con el próximo presupuesto "
          f"contanos.")
    # Actualizado 05/08/2026: mismo enfoque que retencion_primer_presupuesto
    # (ahora empuja a probar Costo/m², no solo pide feedback).
    email = (f"Hola {nombre}, ya armaste un presupuesto en PresupuestoPRO. ¿Sabías que "
             f"también podés calcular el costo por m² de tu obra en segundos, con "
             f"precios siempre actualizados contra corralones reales? Probalo acá: "
             f"https://presupuestopro.com.ar/costo-m2/" + WA_CTA)
    return wa, email


def _mensaje_sin_uso(nombre):
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Vimos que activaste tu cuenta en PresupuestoPRO pero "
          f"todavía no hiciste tu primer presupuesto. ¿Te trabaste en algún paso o "
          f"tuviste alguna duda? Contanos y te ayudamos a armarlo — no lleva más de "
          f"unos minutos. Ingresá en: {APP_URL}")
    # Actualizado 05/08/2026: mismo enfoque que retencion_sin_uso reformulada.
    email = (f"Hola {nombre}, activaste tu cuenta en PresupuestoPRO pero todavía no la "
             f"probaste. Con la app armás el presupuesto completo de tu obra "
             f"—materiales, mano de obra y costo por m²— en minutos, con precios "
             f"siempre actualizados contra corralones reales. Empezá ahora: "
             f"https://presupuestopro.com.ar/login" + WA_CTA)
    return wa, email


def _mensaje_solo_costo_m2(nombre):
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Vimos que probaste la calculadora de Costo/m² en "
          f"PresupuestoPRO pero todavía no armaste un presupuesto completo. Es el "
          f"paso siguiente natural y no lleva mucho más tiempo — ¿te ayudamos a "
          f"armar el primero?")
    # Actualizado 05/08/2026: mismo enfoque que retencion_solo_costo_m2 reformulada.
    email = (f"Hola {nombre}, ya probaste la calculadora de Costo/m² en PresupuestoPRO. "
             f"El paso siguiente es armar el presupuesto completo de tu obra "
             f"—materiales y mano de obra al detalle, con precios siempre actualizados "
             f"contra corralones reales— en minutos. Hacé tu primer presupuesto: "
             f"https://presupuestopro.com.ar/presupuesto/nuevo" + WA_CTA)
    return wa, email


def _mensaje_prueba_por_vencer(nombre):
    """Agregado 20/07/2026, pedido de Daniel — trigger por ciclo de vida
    (prueba gratis), no por uso, así que no participa de _segmento()."""
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Tu prueba gratis de PresupuestoPRO está por terminar. Si "
          f"te sirvió, podés suscribirte desde la app para seguir usándola sin "
          f"cortes. Cualquier duda sobre el pago, contanos.")
    # Actualizado 05/08/2026: mismo enfoque que retencion_prueba_por_vencer_promo.
    email = (f"Hola {nombre}, tu prueba gratis de PresupuestoPRO está por terminar. "
             f"Seguí armando presupuestos completos de tu obra —materiales, mano de "
             f"obra y costo por m²— con precios siempre actualizados contra corralones "
             f"reales, sin cortes. Suscribite acá: "
             f"https://presupuestopro.com.ar/pagos/planes" + WA_CTA)
    return wa, email


def _mensaje_suscripcion_vencida(nombre):
    """Agregado 20/07/2026, pedido de Daniel — trigger por ciclo de vida
    (suscripción paga que no se renovó), no por uso.
    Actualizado 05/08/2026: texto espejado de la plantilla `retencion_suscripcion_vencida`
    reformulada y aprobada por Meta el 04/08/2026 (enfoque de valor + link directo a planes,
    en vez del enfoque de soporte/"¿tuviste una dificultad?" original). El mail ahora
    usa el mismo enfoque (antes tenía el texto viejo de soporte)."""
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Con PresupuestoPRO armás el presupuesto completo de tu obra "
          f"—materiales, mano de obra y costo por m²— en minutos, con precios siempre "
          f"actualizados contra corralones reales, así nunca te quedás corto. Reactivá tu "
          f"suscripción y seguí ahorrando tiempo en cada obra: "
          f"https://presupuestopro.com.ar/pagos/planes")
    email = (f"Hola {nombre}, con PresupuestoPRO armás el presupuesto completo de tu "
             f"obra —materiales, mano de obra y costo por m²— en minutos, con precios "
             f"siempre actualizados contra corralones reales, así nunca te quedás "
             f"corto. Reactivá tu suscripción y seguí ahorrando tiempo en cada obra: "
             f"https://presupuestopro.com.ar/pagos/planes" + WA_CTA)
    return wa, email


def _mensaje_checkin_activo(nombre):
    """Agregado 06/08/2026, pedido de Daniel: mensaje para la categoría
    "ESTUVO USANDO" del nuevo Seguimiento (routes/admin.py) cuando el usuario
    ya tiene 2+ presupuestos/borradores -- antes este grupo (SEG_ACTIVO) no
    tenía ningún mensaje de retención armado porque no participaba de
    ninguna campaña. Es un check-in genérico, no de reactivación (ya está
    usando la app)."""
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Somos de PresupuestoPRO. Vimos que ya armaste varios "
          f"presupuestos con la app -- ¿cómo te está yendo? Si tenés alguna consulta "
          f"o idea para mejorarla, nos encantaría escucharte.")
    email = (f"Hola {nombre}, ya armaste varios presupuestos con PresupuestoPRO. "
             f"¿Cómo te está yendo? Si tenés alguna consulta, sugerencia o encontraste "
             f"algo que se pueda mejorar, contanos -- nos ayuda un montón." + WA_CTA)
    return wa, email


def _mensaje_checkin_abonado(nombre):
    """Agregado 07/08/2026: Daniel preguntó qué se manda al usar "Mail/
    WhatsApp a todos" en la categoría "Abonados" (nueva, ver
    routes/admin.py::_categoria) y se encontró un problema real -- estaba
    reusando _mensaje_checkin_activo(), que dice "vimos que ya armaste
    varios presupuestos". Eso es cierto para SEG_ACTIVO (2+ presupuestos,
    para quien se escribió ese texto originalmente el 06/08), pero "Abonados"
    agrupa a CUALQUIER suscriptor pago esté vencido o no -- alguien recién
    suscripto (ej. Claudio/Rodrigo, abonados desde el 04-05/08) puede tener 0
    o 1 presupuesto todavía, y ese texto le mentiría. Este mensaje es
    genérico a propósito, no asume ningún nivel de uso."""
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Somos de PresupuestoPRO. Gracias por confiar en la app "
          f"para tus obras -- ¿cómo te está yendo? Si tenés alguna consulta, sugerencia "
          f"o algo que se pueda mejorar, contanos, nos ayuda un montón.")
    email = (f"Hola {nombre}, gracias por confiar en PresupuestoPRO para tus obras. "
             f"¿Cómo te está yendo? Si tenés alguna consulta, sugerencia o encontraste "
             f"algo que se pueda mejorar, contanos -- nos ayuda un montón." + WA_CTA)
    return wa, email


def _mensaje_conversion_d(nombre, link=None, vence_str=None):
    """Campaña de conversión, vencidos que eran Segmento D -- 07/08/2026,
    oferta 50%/48hs (reemplaza el 3x1 placeholder original). `link`/
    `vence_str` son opcionales: si Daniel ya generó el link de pago con
    descuento (botón "Generar link" en el perfil, ver routes/admin.py), se
    suman al texto del mail y del WhatsApp "manual" -- la plantilla API
    aprobada en Meta NO lleva el link (Meta no deja agregar texto libre a una
    plantilla ya aprobada), por eso el envío por API sigue mandando el texto
    fijo sin importar estos parámetros; esto es solo para el mail y para
    "Abrir en WhatsApp (manual)"."""
    nombre = nombre or ''
    extra = f" Pagá acá con el descuento ya aplicado: {link}" if link else ""
    vence_txt = f" (válido hasta el {vence_str} hora Argentina)" if vence_str and link else ""
    base = (f"Hola {nombre}! Vimos que probaste la calculadora de Costo/m² en PresupuestoPRO, "
            f"pero eso es solo una parte de lo que resuelve la app. Con el presupuesto completo "
            f"tenés materiales y mano de obra calculados al detalle, con precios siempre "
            f"actualizados. Te ahorrás las vueltas de cotizar a mano llamando al corralón. "
            f"Por 48hs: reactivá tu cuenta con 50% de descuento en el plan que elijas.")
    wa = base + extra + vence_txt if link else base + " Respondé este mensaje y te ayudamos a reactivar con el descuento."
    email = base + extra + vence_txt + WA_CTA if link else base + " Respondé este mail y te ayudamos a reactivar con el descuento." + WA_CTA
    return wa, email


def _mensaje_conversion_b(nombre, link=None, vence_str=None):
    """Campaña de conversión, vencidos que eran Segmento B -- mismo criterio
    y misma nota que _mensaje_conversion_d() de arriba."""
    nombre = nombre or ''
    extra = f" Pagá acá con el descuento ya aplicado: {link}" if link else ""
    vence_txt = f" (válido hasta el {vence_str} hora Argentina)" if vence_str and link else ""
    base = (f"Hola {nombre}! Empezaste un presupuesto en PresupuestoPRO y no lo terminaste. "
            f"La app te calcula materiales y mano de obra al detalle, con precios siempre "
            f"actualizados — te ahorrás las vueltas de cotizar a mano llamando al corralón. "
            f"Por 48hs: reactivá tu cuenta con 50% de descuento en el plan que elijas.")
    wa = base + extra + vence_txt if link else base + " Respondé este mensaje y seguimos donde quedaste, con el descuento."
    email = base + extra + vence_txt + WA_CTA if link else base + " Respondé este mail y seguimos donde quedaste, con el descuento." + WA_CTA
    return wa, email


def _mensaje_testimonio(nombre):
    """Agregado 08/08/2026, pedido de Daniel (vía META_ADS): pedido de
    testimonio real para el creative de la campaña de retargeting a
    Purchase -- se descartaron testimonios inventados por riesgo legal
    (Ley 22.802/24.240) y de política de Meta, así que hace falta juntar
    reales. Mensaje corto, sin presión, aclara que es opcional y que se
    pediría permiso para usarlo (dato de compliance: aunque no haga falta
    firma para un comentario breve, es más prolijo curarse en salud antes
    de publicarlo). Genérico para toda la base -- para Claudio/Rodrigo (los
    2 abonados reales) Daniel manda un texto más personalizado a mano desde
    su perfil (Seguimiento > Ver > editar mensaje), no este genérico."""
    nombre = nombre or ''
    wa = (f"Hola {nombre}! Somos de PresupuestoPRO. Te escribimos por otra cosa: "
          f"¿nos contás en 2-3 líneas qué te pareció usar la app? Nos ayudaría muchísimo "
          f"para mostrarle a otros profesionales cómo les sirvió a otros como vos. Totalmente "
          f"opcional, y si nos das el OK te pedimos permiso antes de usarlo en algún lado.")
    email = (f"Hola {nombre}, te escribimos por algo distinto a lo de siempre: "
             f"¿nos contás en 2-3 líneas qué te pareció usar PresupuestoPRO? Nos ayudaría "
             f"muchísimo para mostrarle a otros profesionales de la construcción cómo les "
             f"sirvió a otros como vos. Es totalmente opcional, y si nos das el OK te "
             f"pedimos permiso antes de usarlo en algún lado." + WA_CTA)
    return wa, email


def _escribir_hoja_segmento(ws, usuarios, mensaje_fn):
    for c, h in enumerate(HEADERS_SEGMENTO, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for r, u in enumerate(usuarios, start=2):
        wa_msg, email_msg = mensaje_fn(u['nombre'])
        estado = 'Validado' if u['email_verificado'] else 'Sin validar (email)'
        row = [
            u['nombre'] or '—', u['email'], u['telefono'] or '', u['ciudad'] or '',
            u['provincia'] or '', u['pais'] or '', u['n_presupuestos'], u['n_borradores'],
            u['n_costo_m2'], estado, (u['created_at'] or '')[:10],
            u['subscription_expires'] or '∞', wa_msg, email_msg,
            u['comentario_seguimiento'] or '',
        ] + _mail_cols(u) + _retencion_cols(u)
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [18, 30, 18, 20, 20, 8, 8, 8, 9, 18, 12, 12, 55, 55, 55, 16, 12, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _escribir_hoja_todos(ws, usuarios):
    for c, h in enumerate(HEADERS_TODOS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for r, u in enumerate(usuarios, start=2):
        estado = 'Validado' if u['email_verificado'] else 'Sin validar (email)'
        row = [
            u['nombre'] or '—', u['email'], u['telefono'] or '', u['ciudad'] or '',
            u['provincia'] or '', u['pais'] or '', u['n_presupuestos'], u['n_borradores'],
            u['n_costo_m2'], estado, _segmento(u), (u['created_at'] or '')[:10],
            u['subscription_expires'] or '∞', u['comentario_seguimiento'] or '',
        ] + _mail_cols(u) + _retencion_cols(u)
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [18, 30, 18, 20, 20, 8, 8, 8, 9, 18, 32, 12, 12, 55, 16, 12, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _escribir_hoja_vencidos(ws, usuarios):
    """Cuentas con subscription_expires < hoy -- tanto prueba gratis vencida
    sin convertir como suscripción paga que no se renovó (columna "Tipo"
    distingue las dos, porque el motivo de la llamada es distinto en cada
    caso)."""
    for c, h in enumerate(HEADERS_VENCIDOS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for r, u in enumerate(usuarios, start=2):
        tipo = 'Prueba gratis (no convirtió)' if u['es_trial'] else 'Suscripción paga (no renovó)'
        row = [
            u['nombre'] or '—', u['email'], u['telefono'] or '', u['ciudad'] or '',
            u['provincia'] or '', u['pais'] or '', u['n_presupuestos'], u['n_borradores'],
            u['n_costo_m2'], tipo, u['subscription_expires'] or '',
            u['comentario_seguimiento'] or '',
        ] + _mail_cols(u) + _retencion_cols(u)
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [18, 30, 18, 20, 20, 8, 8, 8, 9, 26, 12, 55, 16, 12, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _fmt_ars(n):
    return f"$ {n:,.0f}".replace(',', '.')


def _plan_y_cobrado(u, mp_planes):
    """Devuelve (etiqueta_plan, monto_neto_mensual) para la fila de un
    abonado. `u['plan_nombre_actual']` sale de la suscripción 'authorized'
    más reciente de ese usuario (ver subquery en
    routes/admin.py::_usuarios_para_exportar); si por algún motivo viene
    vacía (no debería, plan_nombre tiene default 'mensual' en la tabla desde
    siempre) se asume 'mensual' como plan más común. El monto ya sale
    normalizado a base mensual y neto del {MP_COMISION_PCT}% de Mercado Pago
    -- ver comentario en HEADERS_ABONADOS."""
    plan_key = (u['plan_nombre_actual'] if 'plan_nombre_actual' in u.keys() else None) or 'mensual'
    plan_cfg = mp_planes.get(plan_key) or mp_planes.get('mensual', {})
    precio_mes = plan_cfg.get('precio_mes', 14499)
    nombre_plan = plan_cfg.get('nombre', 'Plan Mensual').replace('Plan ', '')
    neto_mensual = round(precio_mes * (100 - MP_COMISION_PCT) / 100)
    etiqueta = f"{nombre_plan} / {_fmt_ars(precio_mes)}/mes"
    return etiqueta, neto_mensual


def _escribir_hoja_abonados(ws, usuarios, mp_planes):
    """Cuentas pagas y activas ahora mismo (mismo criterio 'sub_activa' que
    ya usan routes/pagos.py::planes()/dashboard() -- es_trial=0, active=1,
    subscription_expires >= hoy). Son los que SÍ están pagando -- el
    segmento que le interesa llamar a Daniel para entender el uso bajo.

    06/08/2026: fila 2 (justo debajo del encabezado, "arriba" de los datos
    como pidió Daniel) trae el TOTAL de la columna "$ cobrados" sumado entre
    todos los abonados de esta hoja -- un vistazo rápido al ingreso mensual
    neto real, sin tener que sumar la columna a mano. Los datos arrancan en
    la fila 3."""
    for c, h in enumerate(HEADERS_ABONADOS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    filas = []
    total_neto = 0
    for u in usuarios:
        etiqueta_plan, neto_mensual = _plan_y_cobrado(u, mp_planes)
        total_neto += neto_mensual
        filas.append([
            u['nombre'] or '—', u['email'], u['telefono'] or '', u['ciudad'] or '',
            u['provincia'] or '', u['pais'] or '', u['n_presupuestos'], u['n_borradores'],
            u['n_costo_m2'], etiqueta_plan, _fmt_ars(neto_mensual),
            (u['abonado_desde'] or '')[:10], u['subscription_expires'] or '',
            u['comentario_seguimiento'] or '',
        ] + _mail_cols(u) + _retencion_cols(u))

    TOTAL_FONT = Font(name="Arial", bold=True, size=10)
    TOTAL_FILL = PatternFill("solid", fgColor="E5E7EB")
    fila_total = [''] * len(HEADERS_ABONADOS)
    fila_total[0] = f"TOTAL ({len(usuarios)} abonados)"
    fila_total[10] = _fmt_ars(total_neto)  # columna "$ cobrados" (índice 10, 0-based)
    for c, val in enumerate(fila_total, start=1):
        cell = ws.cell(row=2, column=c, value=val)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")

    for r, row in enumerate(filas, start=3):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [18, 30, 18, 20, 20, 8, 8, 8, 9, 22, 24, 12, 12, 55, 16, 12, 20, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def generar_excel_usuarios_a_contactar(usuarios, mp_planes=None):
    """usuarios: lista de sqlite3.Row con al menos email_verificado,
    n_presupuestos, n_borradores, n_costo_m2, nombre, email, telefono,
    ciudad, provincia, pais, created_at, subscription_expires, plan_nombre_actual
    (ver query en routes/admin.py::_usuarios_para_exportar).

    mp_planes: dict de config.py::Config.MP_PLANES (current_app.config['MP_PLANES']),
    usado solo por la hoja "Abonados" para mostrar plan/$mes y el cobrado
    neto -- ver _escribir_hoja_abonados(). Si no se pasa, usa un fallback
    mínimo con el plan mensual actual.

    Devuelve (BytesIO, download_name) listo para send_file.
    """
    mp_planes = mp_planes or {'mensual': {'nombre': 'Plan Mensual', 'precio_mes': 14499}}
    segmento_a = [u for u in usuarios if _segmento(u) == SEG_A]
    segmento_b = [u for u in usuarios if _segmento(u) == SEG_B]
    segmento_c = [u for u in usuarios if _segmento(u) == SEG_C]
    segmento_d = [u for u in usuarios if _segmento(u) == SEG_D]

    # 03/08/2026: mismo criterio que ya usa el resto de la app (dashboard de
    # Admin y routes/pagos.py::planes()/dashboard()) -- no se inventa una
    # regla nueva acá.
    # Fix 07/08/2026 (cont. 24): date.today() usa la hora del servidor (UTC en
    # Railway), no ART -- podía clasificar mal a un usuario en la hoja
    # Vencidos/Abonados hasta 3hs antes de tiempo. Mismo ajuste -3hs que en
    # utils/auth.py y routes/admin.py.
    hoy_str = (datetime.utcnow() - timedelta(hours=3)).date().isoformat()
    vencidos = [u for u in usuarios
                if u['subscription_expires'] and u['subscription_expires'] < hoy_str]
    abonados = [u for u in usuarios
                if not u['es_trial'] and u['active']
                and u['subscription_expires'] and u['subscription_expires'] >= hoy_str]

    wb = Workbook()
    ws_todos = wb.active
    ws_todos.title = "Todos los usuarios"
    _escribir_hoja_todos(ws_todos, usuarios)

    ws_vencidos = wb.create_sheet("Vencidos")
    _escribir_hoja_vencidos(ws_vencidos, vencidos)

    ws_abonados = wb.create_sheet("Abonados")
    _escribir_hoja_abonados(ws_abonados, abonados, mp_planes)

    ws_a = wb.create_sheet("A - Sin validar email")
    _escribir_hoja_segmento(ws_a, segmento_a, _mensaje_activacion)

    ws_b = wb.create_sheet("B - 1 presup o borrador")
    _escribir_hoja_segmento(ws_b, segmento_b, _mensaje_seguimiento)

    ws_c = wb.create_sheet("C - Validado sin actividad")
    _escribir_hoja_segmento(ws_c, segmento_c, _mensaje_sin_uso)

    ws_d = wb.create_sheet("D - Solo uso Costo-m2")
    _escribir_hoja_segmento(ws_d, segmento_d, _mensaje_solo_costo_m2)

    notas = wb.create_sheet("Leer primero")
    notas_font = Font(name="Arial", size=11)
    titulo_font = Font(name="Arial", size=13, bold=True)
    lineas = [
        ("Exportado automáticamente desde Admin > Usuarios", titulo_font),
        ("", notas_font),
        (f"Generado el {date.today().isoformat()}. Total de usuarios registrados: "
         f"{len(usuarios)}.", notas_font),
        ("", notas_font),
        ("Hoja 'Todos los usuarios': listado completo, con columna 'Segmento' para ver "
         "de un vistazo en qué grupo cae cada uno (ninguno queda afuera).", notas_font),
        ("Hoja 'Vencidos': prueba gratis o suscripción paga que venció y no se renovó "
         "(columna 'Tipo' distingue cuál de las dos).", notas_font),
        ("Hoja 'Abonados': cuentas pagas y activas ahora mismo -- las que sí están "
         "pagando la suscripción. Columna 'Plan / $mes' muestra el plan actual y su "
         f"precio de lista mensual; '$ cobrados' es ese precio ya neto del "
         f"{MP_COMISION_PCT}% que se queda Mercado Pago, normalizado a mes (para poder "
         "comparar Mensual/Trimestral/Semestral/Anual en la misma base). La fila justo "
         "debajo del encabezado trae el TOTAL de esa columna sumando todos los abonados.",
         notas_font),
("TODAS las hojas (incluidas 'Todos los usuarios' y los segmentos A/B/C/D) "
         "traen una columna 'Comentarios' para anotar, llamada por llamada, la causa "
         "del uso escaso o nulo -- da igual desde qué hoja contactaste a cada uno, es "
         "el mismo comentario en cualquier lado donde aparezca. Lo que escribas ahí NO "
         "se guarda solo -- al tocar 'Exportar' de nuevo, elegí este mismo Excel cuando "
         "te lo pregunte y los comentarios se suman a la base antes de bajar el "
         "siguiente, que ya sale con todo precargado.", notas_font),
        ("Columnas 'Mail: estado'/'Mail: fecha' (todas las hojas): estado del ÚLTIMO mail "
         "que le mandó la app (Entregado/Abierto/Rebotado/etc.), según los webhooks de "
         "Resend. Quedan vacías para todos hasta que se configure el webhook en Resend "
         "(resend.com/webhooks → apuntar a /webhooks/resend) -- ver routes/webhooks_resend.py.",
         notas_font),
        ("Hoja 'A - Sin validar email': cuentas con email_verificado=0.", notas_font),
        ("Hoja 'B - 1 presup o borrador': cuentas validadas con exactamente 1 "
         "presupuesto o borrador en total.", notas_font),
        ("Hoja 'C - Validado sin actividad': cuentas validadas que nunca hicieron nada "
         "(0 presupuestos, 0 borradores, 0 consultas de costo/m²).", notas_font),
        ("Hoja 'D - Solo uso Costo/m2': cuentas validadas que nunca presupuestaron pero "
         "sí usaron la calculadora de costo/m² — antes de hoy quedaban afuera de "
         "cualquier segmento.", notas_font),
        ("Los que tienen 2+ presupuestos/borradores no entran en ningún segmento de "
         "retención (ya son usuarios activos) pero figuran en 'Todos los usuarios' "
         "igual, etiquetados como tal.", notas_font),
        ("", notas_font),
        ("Cada fila de A/B/C/D ya trae un mensaje de WhatsApp y de email sugeridos, "
         "personalizados con el nombre — listos para copiar y pegar.", notas_font),
    ]
    for i, (texto, font) in enumerate(lineas, start=1):
        c = notas.cell(row=i, column=1, value=texto)
        c.font = font
    notas.column_dimensions["A"].width = 95

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    download_name = f"PresupuestoPRO_usuarios_a_contactar_{date.today().isoformat()}.xlsx"
    return buf, download_name
