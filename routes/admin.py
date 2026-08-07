import json
import math
import os
import urllib.request
from datetime import date, datetime, timedelta
from flask import Blueprint, render_template, render_template_string, request, redirect, url_for, flash, g, send_file, current_app
from werkzeug.security import generate_password_hash
from openpyxl import load_workbook
from utils.auth import admin_required
from utils.calculations import PAISES
from utils.normalizacion import PROVINCIAS_AR, telefono_normalizado, telefono_valido
from utils.exportar_contactos import (
    generar_excel_usuarios_a_contactar, _segmento, SEG_A, SEG_B, SEG_C, SEG_D, SEG_ACTIVO,
    _mensaje_activacion, _mensaje_seguimiento, _mensaje_sin_uso, _mensaje_solo_costo_m2,
    _mensaje_prueba_por_vencer, _mensaje_suscripcion_vencida, _mensaje_checkin_activo,
    _mensaje_checkin_abonado, _mensaje_conversion_d, _mensaje_conversion_b,
)
from utils.email_tracking import registrar_envio, ETIQUETAS_EVENTO, SQL_MAIL_ESTADO, SQL_MAIL_ESTADO_FECHA
from database import get_db, recalcular_precio_mo_ars

# Admin > Seguimiento (20/07/2026): nombres de plantilla EXACTOS que hay que
# dar de alta en Meta Business Manager para que esto funcione — ver
# conversación del proyecto para el texto completo de cada una.
TEMPLATES_WHATSAPP = {
    # 05/08/2026: 'A' y 'trial' apuntan a plantillas NUEVAS en categoría
    # Marketing (retencion_activar_cuenta_promo / retencion_prueba_por_vencer_promo)
    # -- las viejas Utility (retencion_activar_cuenta / retencion_prueba_por_vencer)
    # quedaron sin usar acá porque Meta no deja editar una plantilla activa
    # cambiándole la categoría (error real: "Category update is only allowed
    # when updating from old Categories to new ones"). NO DEPLOYAR este cambio
    # hasta que las 2 plantillas nuevas estén APROBADAS en Meta -- si se manda
    # con el nombre nuevo antes de que exista/apruebe, el envío falla con
    # el mismo error 132001 que ya se vio el 21/07.
    'A': 'retencion_activar_cuenta_promo',
    'B': 'retencion_primer_presupuesto',
    'C': 'retencion_sin_uso',
    'D': 'retencion_solo_costo_m2',
    'trial': 'retencion_prueba_por_vencer_promo',
    'vencido': 'retencion_suscripcion_vencida',
    # 06/08/2026: plantilla NUEVA para la categoría "ESTUVO USANDO" (usuarios
    # con 2+ presupuestos/borradores) -- FALTA CREARLA Y QUE LA APRUEBEN EN
    # META antes de que el botón de WhatsApp funcione para este grupo (mismo
    # requisito que el resto, ver aviso en la pantalla de Seguimiento). Hasta
    # entonces el envío por WhatsApp para 'activo' va a dar error visible
    # (no falla en silencio) -- el email sí funciona ya (usa Resend, no Meta).
    'activo': 'retencion_checkin_usuario_activo',
    # 07/08/2026: plantilla NUEVA y distinta de 'activo' -- ver
    # _mensaje_checkin_abonado() en utils/exportar_contactos.py para el
    # motivo (el texto de 'activo' asume 2+ presupuestos, falso para un
    # abonado recién suscripto). FALTA CREARLA Y QUE LA APRUEBEN EN META,
    # mismo criterio que el resto: hasta entonces el WhatsApp da error
    # visible, el email ya funciona (Resend).
    'abonado': 'retencion_checkin_abonado',
    # 07/08/2026: campaña de conversión (oferta 50%/48hs) para vencidos que
    # ERAN Segmento D o B antes de vencer -- ver _tipo_mensaje() más abajo,
    # que ahora distingue esto dentro de la categoría 'vencidos' en vez de
    # mandarles a todos el mismo mensaje de soporte genérico ('vencido').
    'conversion_d': 'retencion_conversion_d',
    'conversion_b': 'retencion_conversion_b',
}

MENSAJES_EMAIL = {
    'A': _mensaje_activacion,
    'B': _mensaje_seguimiento,
    'C': _mensaje_sin_uso,
    'D': _mensaje_solo_costo_m2,
    'trial': _mensaje_prueba_por_vencer,
    'vencido': _mensaje_suscripcion_vencida,
    'activo': _mensaje_checkin_activo,
    'abonado': _mensaje_checkin_abonado,
    'conversion_d': _mensaje_conversion_d,
    'conversion_b': _mensaje_conversion_b,
}

TIPO_LABEL = {
    'A': 'Activar cuenta',
    'B': '1 presup./borrador',
    'C': 'Sin actividad',
    'D': 'Solo Costo/m²',
    'trial': 'Prueba por vencer',
    'vencido': 'Suscripción vencida',
    'activo': 'Check-in (usuario activo)',
    'abonado': 'Check-in (abonado)',
    'conversion_d': 'Conversión 50%/48hs (ex-D)',
    'conversion_b': 'Conversión 50%/48hs (ex-B)',
}

# ── Seguimiento > Retención de usuarios (rediseño 06/08/2026, pedido de
# Daniel) ────────────────────────────────────────────────────────────────
# 5 categorías SIN superposición -- cada usuario cae en UNA sola. Reemplaza
# la lista plana de antes (todos los "tipos" aplicables mezclados). Orden de
# prioridad (lo más urgente primero): un usuario con suscripción vencida cae
# en VENCIDOS aunque también tenga uso previo; uno con prueba por vencer cae
# en POR_VENCER aunque nunca haya validado el email, etc. Confirmado con
# Daniel: VENCIDOS junta prueba gratis vencida + suscripción paga vencida
# (mismo criterio que ya usa el Excel/dashboard); POR_VENCER es SOLO prueba
# gratis por terminar (hoy no hay noción de "pago que vence pronto", ver
# nota en _categoria()).
CATEGORIAS_SEGUIMIENTO = ['vencidos', 'por_vencer', 'abonados', 'sin_validar', 'nunca_probo', 'estuvo_usando']

# Orden por importancia (07/08/2026, pedido de Daniel al agregar "Abonados"):
# 1-2. Vencidos/Por vencer -- urgente, plata que se está perdiendo o a punto
#      de perderse.
# 3. Abonados -- ya están pagando, es la plata que HOY entra: no es urgente
#    (no tienen nada por vencer), pero es la más importante de cuidar/no
#    perder, así que va inmediatamente después de los 2 urgentes y antes de
#    los segmentos que todavía no convirtieron.
# 4-6. Sin validar / Nunca probó / Estuvo usando -- prospectos, en el mismo
#      orden que ya tenían (sin tocar ese criterio, era el de Daniel del
#      06/08/2026).
CATEGORIA_LABEL = {
    'vencidos':      'Vencidos',
    'por_vencer':    'Por vencer',
    'abonados':      'Abonados',
    'sin_validar':   'Sin validar',
    'nunca_probo':   'Nunca probó',
    'estuvo_usando': 'Estuvo usando',
}

CATEGORIA_DESC = {
    'vencidos':      'Prueba gratis vencida sin convertir, o suscripción paga que no se renovó.',
    'por_vencer':    'Prueba gratis por terminar (3 días o menos, o último presupuesto disponible).',
    'abonados':      'Suscripción paga y al día -- no están por vencer ni vencidos, es la base que hoy factura.',
    'sin_validar':   'No activaron la cuenta por mail todavía.',
    'nunca_probo':   'Validaron la cuenta pero nunca hicieron un presupuesto ni usaron Costo/m².',
    'estuvo_usando': 'Ya usaron la app (1+ presupuesto, borrador o Costo/m²) y no están por vencer ni vencidos.',
}


def _categoria(fila):
    """Devuelve UNA de las 6 categorías de CATEGORIAS_SEGUIMIENTO para esta
    fila (ya con 'segmento'/'trial_por_vencer'/'suscripcion_vencida'
    calculados por _usuarios_seguimiento()). 'abonados' (07/08/2026) se
    chequea antes que segmento (sin_validar/nunca_probo/estuvo_usando)
    porque alguien que ya paga puede tener cualquier segmento de uso -- acá
    interesa verlo agrupado como abonado, no mezclado en esos 3."""
    if fila['suscripcion_vencida']:
        return 'vencidos'
    if fila['trial_por_vencer']:
        return 'por_vencer'
    if not fila['es_trial']:
        return 'abonados'
    if fila['segmento'] == SEG_A:
        return 'sin_validar'
    if fila['segmento'] == SEG_C:
        return 'nunca_probo'
    return 'estuvo_usando'  # SEG_B, SEG_D o SEG_ACTIVO


def _tipo_mensaje(fila, categoria):
    """Tipo de mensaje (clave de TEMPLATES_WHATSAPP/MENSAJES_EMAIL/TIPO_LABEL)
    a usar para ESTA fila dentro de su categoría. Para 'estuvo_usando' varía
    según el segmento real (B/D/activo) porque cada uno tiene un mensaje
    distinto ya armado -- las demás categorías son 1 a 1 con un tipo fijo.
    'abonados' reusa el mismo mensaje de check-in que ya existía para
    usuarios activos ('activo') -- mismo texto, ahora accesible también
    desde acá para cualquier abonado sea cual sea su uso."""
    if categoria == 'sin_validar':
        return 'A'
    if categoria == 'nunca_probo':
        return 'C'
    if categoria == 'por_vencer':
        return 'trial'
    if categoria == 'vencidos':
        # 07/08/2026: campaña de conversión 50%/48hs solo para quienes eran
        # D o B antes de vencer (los más prometedores, ver PROYECTO.md) -- el
        # resto de vencidos (A/C) sigue con el mensaje de soporte genérico.
        # fila['segmento'] sale de _segmento(), que mira uso real (presup./
        # borradores/costo_m2) sin importar si está vencido o no -- no se
        # pierde este dato al vencerse la prueba/suscripción.
        if fila['segmento'] == SEG_D:
            return 'conversion_d'
        if fila['segmento'] == SEG_B:
            return 'conversion_b'
        return 'vencido'
    if categoria == 'abonados':
        return 'abonado'
    if fila['segmento'] == SEG_B:
        return 'B'
    if fila['segmento'] == SEG_D:
        return 'D'
    return 'activo'

SEG_A_CODE = {SEG_A: 'A', SEG_B: 'B', SEG_C: 'C', SEG_D: 'D'}


def _tipos_aplicables(fila):
    """Devuelve la lista de tipos de mensaje que aplican a esta fila (0, 1 o
    2: el segmento de uso A/B/C/D es mutuamente excluyente, pero los
    triggers de ciclo de vida -- prueba por vencer / suscripción vencida --
    son independientes y pueden sumarse al mismo usuario)."""
    tipos = []
    code = SEG_A_CODE.get(fila['segmento'])
    if code:
        tipos.append(code)
    if fila.get('trial_por_vencer'):
        tipos.append('trial')
    if fila.get('suscripcion_vencida'):
        # 07/08/2026: mismo criterio que _tipo_mensaje() para la categoría
        # 'vencidos' -- los que eran D/B antes de vencer entran a la
        # campaña de conversión 50%/48hs en vez del mensaje de soporte
        # genérico.
        if fila['segmento'] == SEG_D:
            tipos.append('conversion_d')
        elif fila['segmento'] == SEG_B:
            tipos.append('conversion_b')
        else:
            tipos.append('vencido')
    return tipos

bp = Blueprint('admin', __name__, url_prefix='/admin')

@bp.route('/')
@admin_required
def dashboard():
    db = get_db()
    # Fix 07/08/2026, pedido de Daniel: "Usuarios totales" y "Activos" daban
    # siempre el mismo número (active=1 es el flag de cuenta habilitada, no
    # de plan pago -- prácticamente todos lo tienen en 1) y no decía nada
    # útil a primera vista. Se reemplaza por 3 cuadros SIN superposición que
    # sí importan para el negocio: Abonados (pagando, al día -- mismo
    # criterio que la hoja "Abonados" del export: es_trial=0/active=1/no
    # vencido), En prueba (trial vigente, no vencido) y Vencidos (se deja
    # igual que antes: prueba o pago vencido, sin convertir). 'activos' se
    # mantiene en el dict (lo sigue usando la barra de "conectados ahora"
    # más abajo) pero ya no es uno de los 4 KPI de arriba.
    stats = {
        'total_users':   db.execute("SELECT COUNT(*) as c FROM users WHERE is_admin=0").fetchone()['c'],
        'activos':       db.execute("SELECT COUNT(*) as c FROM users WHERE active=1 AND is_admin=0").fetchone()['c'],
        # Fix 06/08/2026 (cont. 23): date('now') es UTC -- ver mismo fix en
        # utils/auth.py. Contaba usuarios como vencidos hasta 3hs antes de
        # tiempo en la franja 21:00-23:59 ART.
        'vencidos':      db.execute("SELECT COUNT(*) as c FROM users WHERE subscription_expires < date('now', '-3 hours') AND is_admin=0").fetchone()['c'],
        'abonados':      db.execute(
            "SELECT COUNT(*) as c FROM users WHERE is_admin=0 AND es_trial=0 AND active=1 "
            "AND (subscription_expires IS NULL OR subscription_expires >= date('now', '-3 hours'))"
        ).fetchone()['c'],
        'en_prueba':     db.execute(
            "SELECT COUNT(*) as c FROM users WHERE is_admin=0 AND es_trial=1 "
            "AND (subscription_expires IS NULL OR subscription_expires >= date('now', '-3 hours'))"
        ).fetchone()['c'],
        'presupuestos':  db.execute("SELECT COUNT(*) as c FROM presupuestos").fetchone()['c'],
        'mensajes_nuevos': db.execute("SELECT COUNT(*) as c FROM contactos WHERE leido=0").fetchone()['c'],
        'sugerencias_nuevas': db.execute("SELECT COUNT(*) as c FROM sugerencias WHERE leido=0").fetchone()['c'],
        # Fix 08/07/2026: badge de inscriptos nuevos — el botón "Inscriptos"
        # ya existía en el dashboard pero sin ningún contador, así que un
        # inscripto nuevo (ej. Ricardo Jordan) pasaba desapercibido salvo que
        # Daniel entrara a revisar la lista sin motivo aparente.
        'leads_nuevos': db.execute("SELECT COUNT(*) as c FROM leads WHERE estado='nuevo'").fetchone()['c'],
        # Fix 20/07/2026, pedido de Daniel: badge de consultas de WhatsApp sin
        # responder (el bot no supo contestar y quedaron para revisión manual
        # -- ver admin.whatsapp_inbox).
        'whatsapp_pendientes': db.execute(
            "SELECT COUNT(*) as c FROM whatsapp_consultas_sin_responder WHERE respondida=0"
        ).fetchone()['c'],
        # Fix 25/07/2026, pedido de Daniel (fase 1 del CRM unificado): mismo
        # badge que WhatsApp, para Messenger + Instagram (routes/social_bot.py).
        'redes_pendientes': db.execute(
            "SELECT COUNT(*) as c FROM redes_consultas_sin_responder WHERE respondida=0"
        ).fetchone()['c'],
        # Fix 25/07/2026, pedido de Daniel (CRM unificado, paso 1 -- webhook
        # de mail entrante): mismo badge que WhatsApp/Redes, para el mail
        # entrante a contacto@presupuestopro.com.ar (ver routes/email_bot.py).
        'email_pendientes': db.execute(
            "SELECT COUNT(*) as c FROM email_consultas_entrantes WHERE respondida=0"
        ).fetchone()['c'],
    }
    # Fix 20/07/2026, pedido de Daniel: badge de usuarios con algo pendiente
    # en Admin > Seguimiento (segmento A/B/C/D, prueba por vencer o
    # suscripción vencida). Se calcula en Python (no en SQL) porque el
    # segmento sale de utils.exportar_contactos._segmento — misma lógica
    # que usa esa pantalla, sin duplicar la regla acá.
    stats['seguimiento_pendientes'] = len([f for f in _usuarios_seguimiento() if f['tipos']])
    # "Próximos vencimientos" sacado 06/08/2026, pedido de Daniel -- esa info
    # ahora vive en Seguimiento (POR VENCER/VENCIDOS), con acciones en vez de
    # ser solo una lista para mirar.
    # Fix 06/08/2026, pedido de Daniel: cuadro "Actividad de usuarios" --
    # conectados ahora (ventana de 5 min, ver utils/auth.py::get_current_user)
    # + gráfico de usuarios conectados por día (actividad_diaria, últimos 30
    # días con datos). El gráfico solo tiene datos desde que se agregó este
    # tracking (06/08/2026) -- no hay forma de reconstruir días anteriores.
    stats['conectados_ahora'] = db.execute(
        "SELECT COUNT(*) as c FROM users WHERE is_admin=0 AND ultima_actividad >= datetime('now','-5 minutes')"
    ).fetchone()['c']
    dias_rows = db.execute(
        "SELECT fecha, COUNT(*) as c FROM actividad_diaria GROUP BY fecha ORDER BY fecha DESC LIMIT 30"
    ).fetchall()
    chart_dias = [{'fecha': r['fecha'], 'cantidad': r['c']} for r in reversed(dias_rows)]
    db.close()
    return render_template('admin/dashboard.html', stats=stats,
                            chart_dias=chart_dias, user=g.user)


@bp.route('/api/conectados-ahora')
@admin_required
def conectados_ahora_json():
    """Endpoint liviano para refrescar solo el número/barra de "conectados
    ahora" del dashboard cada 30s (JS, ver dashboard.html) sin recargar toda
    la página ni volver a pegarle al gráfico de días."""
    db = get_db()
    n = db.execute(
        "SELECT COUNT(*) as c FROM users WHERE is_admin=0 AND ultima_actividad >= datetime('now','-5 minutes')"
    ).fetchone()['c']
    db.close()
    return {'conectados_ahora': n}

# USUARIOS
@bp.route('/usuarios')
@admin_required
def usuarios():
    """Fix 05/07/2026: agregado conteo de presupuestos completos/borradores y
    consultas de Costo/m2 por usuario, + filtros por localidad/provincia/pais
    (pedido de Daniel para tener a mano el uso real de cada usuario)."""
    f_ciudad    = (request.args.get('f_ciudad') or '').strip()
    f_provincia = (request.args.get('f_provincia') or '').strip()
    f_pais      = (request.args.get('f_pais') or '').strip()

    where = ["is_admin=0"]
    params = []
    if f_ciudad:
        where.append("ciudad LIKE ?")
        params.append(f"%{f_ciudad}%")
    if f_provincia:
        # Fix 10/07/2026: provincia ya es lista cerrada (select) — match exacto,
        # no LIKE. Con LIKE, "Buenos Aires" también traía a los de "Ciudad
        # Autónoma de Buenos Aires" por ser substring.
        where.append("provincia = ?")
        params.append(f_provincia)
    if f_pais:
        where.append("pais = ?")
        params.append(f_pais)

    db = get_db()
    users = db.execute(
        f"""SELECT u.*,
                   (SELECT COUNT(*) FROM presupuestos p WHERE p.user_id=u.id AND p.status='completo'
                      AND (p.es_demo IS NULL OR p.es_demo=0))                                          AS n_presupuestos,
                   (SELECT COUNT(*) FROM presupuestos p WHERE p.user_id=u.id AND p.status='borrador')  AS n_borradores,
                   (SELECT COUNT(*) FROM costo_m2_consultas c WHERE c.user_id=u.id)                    AS n_costo_m2,
                   {SQL_MAIL_ESTADO}       AS mail_estado,
                   {SQL_MAIL_ESTADO_FECHA} AS mail_estado_fecha
            FROM users u
            WHERE {' AND '.join(where)}
            ORDER BY u.created_at DESC""",
        params
    ).fetchall()

    # Fix 10/07/2026 (cont. 8, pedido de Daniel): el badge "Usuarios" del
    # encabezado tiene que marcar SIEMPRE el total de la base (sin filtrar),
    # no la cantidad de filas que quedaron después de aplicar los filtros
    # (eso ya lo muestra `users|length` en la tabla misma).
    total_usuarios = db.execute("SELECT COUNT(*) as c FROM users WHERE is_admin=0").fetchone()['c']

    # Fix 10/07/2026 (2da vuelta, pedido de Daniel): en vez de un cartel de
    # "también en" aparte, mostrar la cantidad de cada nivel (ciudad,
    # provincia, país) pegada al lado de su propio label. Cuando un nivel no
    # está elegido explícitamente, se infiere del nivel más específico que sí
    # esté activo (ciudad -> su provincia; provincia -> su país), tomando la
    # provincia/país más frecuente entre los usuarios que matchean, por si
    # hay datos mezclados. Si no hay nada de qué inferir, ese contador queda
    # en None y el template no muestra badge.
    contadores = {'ciudad': None, 'provincia': None, 'pais': None}

    def _mas_frecuente(campo, where_extra, params_extra):
        filas = db.execute(
            f"SELECT {campo} as v, COUNT(*) as c FROM users WHERE is_admin=0 AND {where_extra} AND {campo} != '' "
            f"GROUP BY {campo} ORDER BY c DESC LIMIT 1",
            params_extra
        ).fetchone()
        return filas['v'] if filas else None

    if f_ciudad:
        contadores['ciudad'] = db.execute(
            "SELECT COUNT(*) as c FROM users WHERE is_admin=0 AND ciudad LIKE ?", (f"%{f_ciudad}%",)
        ).fetchone()['c']

    provincia_efectiva = f_provincia or (
        _mas_frecuente('provincia', "ciudad LIKE ?", (f"%{f_ciudad}%",)) if f_ciudad else None
    )
    if provincia_efectiva:
        contadores['provincia'] = {
            'nombre': provincia_efectiva,
            'cantidad': db.execute(
                "SELECT COUNT(*) as c FROM users WHERE is_admin=0 AND provincia=?", (provincia_efectiva,)
            ).fetchone()['c'],
        }

    pais_efectivo = f_pais or (
        _mas_frecuente('pais', "provincia=?", (provincia_efectiva,)) if provincia_efectiva else
        (_mas_frecuente('pais', "ciudad LIKE ?", (f"%{f_ciudad}%",)) if f_ciudad else None)
    )
    if pais_efectivo:
        contadores['pais'] = {
            'nombre': PAISES.get(pais_efectivo, {}).get('nombre', pais_efectivo),
            'cantidad': db.execute(
                "SELECT COUNT(*) as c FROM users WHERE is_admin=0 AND pais=?", (pais_efectivo,)
            ).fetchone()['c'],
        }

    # Fix 10/07/2026: antes esto era un DISTINCT de lo cargado en users.provincia
    # (texto libre, con duplicados). Provincia ahora es lista cerrada — se
    # usa PROVINCIAS_AR para que el filtro sea un <select> real, igual que País.
    # Localidad sigue siendo abierta, pero ahora hay tabla `localidades`
    # autoalimentada (fix de hoy) para ofrecer autocompletado real, no el
    # autocompletado del navegador que Daniel vio antes.
    # Fix 12/07/2026: antes ordenaba por uso (veces_usada DESC) -> las
    # sugerencias del filtro de Localidad aparecían desordenadas. Pedido de
    # Daniel: alfabético, para elegir más fácil desde el celular.
    # Fix 05/08/2026 (bug reportado por Daniel): esta lista salía de la tabla
    # `localidades`, que solo se mantiene sincronizada cuando la ciudad se
    # carga por registro (`_guardar_localidad`) o se toca desde Admin >
    # Localidades (renombrar/fusionar). `usuario_editar` (editar un usuario a
    # mano desde Admin > Usuarios) actualiza `users.ciudad` directo, sin tocar
    # `localidades` — así quedó el caso de Claudio: se corrigió su ciudad de
    # "T" a "Tostado", pero el filtro seguía ofreciendo "T" porque esa fila
    # vieja nunca se actualizó ni se borró de `localidades`. Se cambia la
    # fuente del filtro a `users.ciudad` en vivo (DISTINCT) — así siempre
    # coincide con lo que realmente tiene cargado cada usuario, sin importar
    # por dónde se haya editado, y no depende de mantener 2 tablas en sync.
    localidades_lista = [r['ciudad'] for r in db.execute(
        "SELECT DISTINCT ciudad FROM users WHERE is_admin=0 AND ciudad != '' ORDER BY ciudad COLLATE NOCASE ASC"
    ).fetchall()]
    db.close()

    return render_template('admin/usuarios.html', users=users, user=g.user,
                            provincias=PROVINCIAS_AR, localidades_lista=localidades_lista, paises=PAISES,
                            f_ciudad=f_ciudad, f_provincia=f_provincia, f_pais=f_pais,
                            contadores=contadores, total_usuarios=total_usuarios,
                            etiquetas_mail=ETIQUETAS_EVENTO)


def _usuarios_para_exportar(db):
    """Query compartida por usuarios_exportar_contactar() y usuarios_exportar()
    -- se separó acá el 03/08/2026 (2da vuelta) para no duplicarla entre las
    2 rutas."""
    return db.execute(
        f"""SELECT u.*,
                  (SELECT COUNT(*) FROM presupuestos p WHERE p.user_id=u.id AND p.status='completo'
                     AND (p.es_demo IS NULL OR p.es_demo=0))                                          AS n_presupuestos,
                  (SELECT COUNT(*) FROM presupuestos p WHERE p.user_id=u.id AND p.status='borrador')  AS n_borradores,
                  (SELECT COUNT(*) FROM costo_m2_consultas c WHERE c.user_id=u.id)                    AS n_costo_m2,
                  (SELECT MIN(fecha_inicio) FROM suscripciones s
                     WHERE s.user_id=u.id AND s.estado='authorized')                                  AS abonado_desde,
                  (SELECT plan_nombre FROM suscripciones s
                     WHERE s.user_id=u.id AND s.estado='authorized'
                     ORDER BY s.fecha_inicio DESC, s.id DESC LIMIT 1)                                  AS plan_nombre_actual,
                  {SQL_MAIL_ESTADO}       AS mail_estado,
                  {SQL_MAIL_ESTADO_FECHA} AS mail_estado_fecha,
                  (SELECT MAX(rc.created_at) FROM retencion_contactos rc
                     WHERE rc.user_id=u.id AND rc.canal='email'
                       AND rc.resultado='ok')                                                          AS ultimo_email_retencion,
                  (SELECT MAX(rc.created_at) FROM retencion_contactos rc
                     WHERE rc.user_id=u.id AND rc.canal='whatsapp'
                       AND rc.resultado='ok')                                                          AS ultimo_whatsapp_retencion
           FROM users u
           WHERE u.is_admin=0
           ORDER BY u.created_at DESC"""
    ).fetchall()


def _importar_comentarios_xlsx(db, archivo):
    """Lee TODAS las hojas de `archivo` (un .xlsx ya anotado por Daniel) que
    tengan columnas "Email" y "Comentarios" -- no una lista fija de nombres
    de hoja, así funciona sea cual sea la hoja desde donde Daniel contactó
    (Vencidos, Abonados, Todos los usuarios, A/B/C/D -- todas las tienen
    desde 03/08/2026 2da vuelta, pedido de Daniel: quiere que el comentario
    se mantenga sin importar desde qué hoja/segmento contactó a cada uno).
    Detecta las columnas por nombre de encabezado, no por posición fija, para
    no romper si alguna hoja las reordena. Guarda en users.comentario_seguimiento,
    matcheando por email (case-insensitive). Solo pisa un comentario si el
    texto cambió. La hoja "Leer primero" se salta sola (no tiene esos
    encabezados). Usada tanto por usuarios_importar_comentarios (standalone)
    como por usuarios_exportar() (combinado en 1 solo click). Devuelve (ok,
    error_o_None, cantidad_actualizados) -- NO cierra `db`, eso lo maneja
    quien la llama."""
    if not archivo.filename.lower().endswith('.xlsx'):
        return False, 'Tiene que ser un .xlsx -- el mismo que bajaste con "Exportar".', 0
    try:
        wb_in = load_workbook(archivo, data_only=True)
    except Exception:
        return False, 'No se pudo leer ese archivo -- ¿es el Excel exportado desde acá?', 0

    hoy_str = date.today().isoformat()
    actualizados = 0
    for nombre_hoja in wb_in.sheetnames:
        ws = wb_in[nombre_hoja]
        primera_fila = next(ws.iter_rows(min_row=1, max_row=1), None)
        if primera_fila is None:
            continue
        headers = [c.value for c in primera_fila]
        if 'Email' not in headers or 'Comentarios' not in headers:
            continue
        col_email = headers.index('Email')
        col_comentario = headers.index('Comentarios')

        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_email >= len(row):
                continue
            email = (row[col_email] or '').strip().lower()
            comentario = (row[col_comentario] or '').strip() if col_comentario < len(row) and row[col_comentario] else ''
            if not email or not comentario:
                continue
            actual = db.execute(
                "SELECT comentario_seguimiento FROM users WHERE lower(email)=?", (email,)
            ).fetchone()
            if actual is None:
                continue
            if (actual['comentario_seguimiento'] or '') != comentario:
                db.execute(
                    "UPDATE users SET comentario_seguimiento=?, comentario_actualizado=? WHERE lower(email)=?",
                    (comentario, hoy_str, email)
                )
                actualizados += 1
    db.commit()
    return True, None, actualizados


@bp.route('/usuarios/exportar-contactar')
@admin_required
def usuarios_exportar_contactar():
    """Pedido de Daniel 15/07/2026: antes esta lista se armaba a mano
    (capturas de pantalla de esta misma tabla + transcripción manual a un
    Excel — lento y con riesgo de error en teléfonos/contadores). Este botón
    arma el mismo Excel en un click, leyendo directo de la base. Ver
    utils/exportar_contactos.py para la lógica de segmentación exacta
    (incluye Segmento C — validado sin ninguna actividad — agregado el mismo
    día a pedido de Daniel, por eso también se trae n_costo_m2 acá).

    Export "directo" (sin merge de comentarios) -- lo usa el botón "Exportar"
    cuando Daniel cancela el picker de archivo (ver usuarios_exportar() más
    abajo, que es el combinado)."""
    db = get_db()
    usuarios = _usuarios_para_exportar(db)
    db.close()

    buf, download_name = generar_excel_usuarios_a_contactar(
        usuarios, mp_planes=current_app.config.get('MP_PLANES', {}))
    return send_file(buf,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                      as_attachment=True, download_name=download_name)


@bp.route('/usuarios/exportar', methods=['POST'])
@admin_required
def usuarios_exportar():
    """Pedido de Daniel 03/08/2026 (2da vuelta): en vez de 2 pasos separados
    ("Importar comentarios" y después "Exportar"), un solo click. El botón
    "Exportar" de la tabla ahora pregunta (confirm de JS, ver usuarios.html)
    si Daniel tiene un Excel anterior con comentarios para sumar:
      - Si elige un archivo: esta ruta primero llama a
        _importar_comentarios_xlsx() (mismos comentarios guardados en la
        base) y RECIÉN DESPUÉS genera y devuelve el Excel nuevo -- ya con
        esos comentarios adentro -- todo en una sola respuesta/descarga.
      - Si cancela el picker: el JS del template ni siquiera pega acá, va
        directo a usuarios_exportar_contactar() (GET, sin archivo).
    No hay forma de que el servidor "busque solo" el Excel anterior en la
    compu de Daniel -- ningún navegador permite leer archivos locales sin
    que el usuario los elija (restricción de seguridad, no una limitación
    de esta implementación). Este flujo es lo más cercano a "un solo click"
    que se puede lograr respetando esa restricción."""
    archivo = request.files.get('excel')
    db = get_db()
    if archivo and archivo.filename:
        ok, error, actualizados = _importar_comentarios_xlsx(db, archivo)
        if not ok:
            db.close()
            flash(error, 'error')
            return redirect(url_for('admin.usuarios'))

    usuarios = _usuarios_para_exportar(db)
    db.close()

    buf, download_name = generar_excel_usuarios_a_contactar(
        usuarios, mp_planes=current_app.config.get('MP_PLANES', {}))
    return send_file(buf,
                      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                      as_attachment=True, download_name=download_name)


@bp.route('/usuarios/importar-comentarios', methods=['POST'])
@admin_required
def usuarios_importar_comentarios():
    """Guarda comentarios de un Excel anterior SIN exportar uno nuevo al
    toque (por si Daniel solo quiere ir bancando el progreso de las llamadas
    sin necesitar un Excel fresco en ese momento). El flujo de 1 solo click
    (importar + exportar juntos) es usuarios_exportar() arriba -- ver ese
    docstring para el detalle completo de por qué existe la subida manual."""
    archivo = request.files.get('excel')
    if not archivo or not archivo.filename:
        flash('Elegí primero el Excel que ya tenés con los comentarios anotados.', 'error')
        return redirect(url_for('admin.usuarios'))

    db = get_db()
    ok, error, actualizados = _importar_comentarios_xlsx(db, archivo)
    db.close()
    if not ok:
        flash(error, 'error')
        return redirect(url_for('admin.usuarios'))

    if actualizados:
        flash(f'{actualizados} comentario(s) guardado(s). El próximo Excel que exportes ya '
              f'va a salir con estos comentarios cargados.', 'success')
    else:
        flash('No se encontró ningún comentario nuevo para trasladar en ese archivo.', 'success')
    return redirect(url_for('admin.usuarios'))


def _usuarios_seguimiento():
    """Arma la lista de usuarios con su segmento de uso (A/B/C/D, misma
    lógica que el Excel de utils/exportar_contactos.py) + los dos triggers
    de ciclo de vida que pidió Daniel 20/07/2026: prueba gratis por vencer
    (usa el mismo criterio que utils/trial.py, calculado acá para no abrir
    una conexión a DB por usuario) y suscripción vencida (mismo criterio que
    el contador 'vencidos' del dashboard). Devuelve la lista ya con
    'segmento', 'trial_por_vencer', 'dias_restantes', 'presup_restantes',
    'suscripcion_vencida', 'ultimo_contacto' y 'ultimo_resultado' agregados
    a cada fila. 'ultimo_resultado' (21/07/2026, pedido de Daniel) es para
    poder mostrar si el último envío salió bien o mal directo en la lista,
    sin depender de que el flash message se vea (ver nota en seguimiento())."""
    db = get_db()
    usuarios = db.execute(
        """SELECT u.*,
                  (SELECT COUNT(*) FROM presupuestos p WHERE p.user_id=u.id AND p.status='completo'
                     AND (p.es_demo IS NULL OR p.es_demo=0))                                          AS n_presupuestos,
                  (SELECT COUNT(*) FROM presupuestos p WHERE p.user_id=u.id AND p.status='borrador')  AS n_borradores,
                  (SELECT COUNT(*) FROM costo_m2_consultas c WHERE c.user_id=u.id)                    AS n_costo_m2,
                  (SELECT MAX(created_at) FROM retencion_contactos rc WHERE rc.user_id=u.id)          AS ultimo_contacto,
                  (SELECT resultado FROM retencion_contactos rc WHERE rc.user_id=u.id
                     ORDER BY rc.created_at DESC LIMIT 1)                                              AS ultimo_resultado
           FROM users u
           WHERE u.is_admin=0 AND (u.retencion_opt_out IS NULL OR u.retencion_opt_out=0)
           ORDER BY u.created_at DESC"""
    ).fetchall()
    db.close()

    # Fix 07/08/2026 (cont. 24): date.today() usa la hora del servidor (UTC en
    # Railway), no ART -- en la franja 21:00-23:59 ART del día de vencimiento,
    # marcaba "suscripción vencida" hasta 3hs antes de tiempo (afecta el
    # segmento VENCIDOS de Admin > Seguimiento y sus envíos masivos). Mismo
    # ajuste -3hs que en utils/auth.py.
    hoy_str = (datetime.utcnow() - timedelta(hours=3)).date().isoformat()
    ahora = datetime.utcnow()
    filas = []
    for u in usuarios:
        fila = dict(u)
        fila['segmento'] = _segmento(u)

        trial_por_vencer = False
        dias_restantes = presup_restantes = None
        if u['es_trial']:
            try:
                creado = datetime.fromisoformat((u['created_at'] or '').replace(' ', 'T'))
                dias_pasados = (ahora - creado).days
            except (ValueError, TypeError):
                dias_pasados = 0
            dias_restantes = max(0, 14 - dias_pasados)
            presup_restantes = max(0, 3 - u['n_presupuestos'])
            vencido_trial = u['n_presupuestos'] >= 3 or dias_pasados >= 14
            trial_por_vencer = (not vencido_trial) and (dias_restantes <= 3 or presup_restantes <= 1)
        fila['trial_por_vencer'] = trial_por_vencer
        fila['dias_restantes'] = dias_restantes
        fila['presup_restantes'] = presup_restantes

        fila['suscripcion_vencida'] = bool(u['subscription_expires']) and u['subscription_expires'] < hoy_str

        fila['tipos'] = _tipos_aplicables(fila)
        filas.append(fila)
    return filas


_FLASH_BLOCK = """
{% with messages = get_flashed_messages(with_categories=true) %}
  {% if messages %}
  <div class="mb-3">
    {% for category, msg in messages %}
    <div class="alert alert-{{ 'danger' if category=='error' else category }} py-2 mb-1">{{ msg }}</div>
    {% endfor %}
  </div>
  {% endif %}
{% endwith %}
"""


@bp.route('/seguimiento')
@admin_required
def seguimiento():
    """Rediseño 06/08/2026, pedido de Daniel: reemplaza la lista plana de
    antes (todos los usuarios accionables mezclados, con ?todos=1 para ver
    el resto) por un landing con las 5 categorías de retención SIN
    superposición (CATEGORIAS_SEGUIMIENTO/_categoria(), ver arriba) y su
    cantidad -- entrando a cada una se ve la lista real con botones
    "Mail a todos"/"WhatsApp a todos" (ver seguimiento_categoria)."""
    filas = _usuarios_seguimiento()
    conteos = {c: 0 for c in CATEGORIAS_SEGUIMIENTO}
    for f in filas:
        conteos[_categoria(f)] += 1

    return render_template_string("""
<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Seguimiento - Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css">
</head><body class="bg-light">
<div class="container-fluid py-4" style="max-width:900px">
  <a href="/admin/" class="btn btn-outline-secondary btn-sm mb-2">Volver</a>
  <h4 class="fw-bold mb-1">Seguimiento — Retención de usuarios</h4>
  <p class="text-muted small mb-3">{{ total }} usuario{{ 's' if total != 1 else '' }} en total. Elegí una categoría para ver la lista y contactarlos.</p>
  """ + _FLASH_BLOCK + """
  <div class="row g-3">
    {% for cat in categorias %}
    <div class="col-6 col-md-4">
      <a href="{{ url_for('admin.seguimiento_categoria', categoria=cat) }}" class="text-decoration-none">
        <div class="card shadow-sm h-100 {{ 'border-danger' if cat=='vencidos' else ('border-warning' if cat=='por_vencer' else ('border-success' if cat=='abonados' else '')) }}">
          <div class="card-body text-center py-4">
            <div class="display-6 fw-bold {{ 'text-danger' if cat=='vencidos' else ('text-warning' if cat=='por_vencer' else ('text-success' if cat=='abonados' else 'text-dark')) }}">{{ conteos[cat] }}</div>
            <div class="fw-semibold text-uppercase small mt-2">{{ cat_label[cat] }}</div>
            <div class="text-muted small mt-1">{{ cat_desc[cat] }}</div>
          </div>
        </div>
      </a>
    </div>
    {% endfor %}
  </div>
</div></body></html>
""", categorias=CATEGORIAS_SEGUIMIENTO, conteos=conteos, cat_label=CATEGORIA_LABEL,
    cat_desc=CATEGORIA_DESC, total=len(filas), user=g.user)


@bp.route('/seguimiento/categoria/<categoria>')
@admin_required
def seguimiento_categoria(categoria):
    """Lista de una sola categoría (ver _categoria()), con los botones
    "Mail a todos"/"WhatsApp a todos" (1 click, sin revisar mensaje por
    mensaje -- confirmado por Daniel 06/08/2026) arriba de todo, y debajo la
    lista con los datos de cada uno (igual que Admin > Usuarios) para poder
    llamarlos, más las acciones individuales por si hace falta reenviar a
    uno solo o editar el texto antes de mandarlo (Ver → seguimiento_detalle)."""
    if categoria not in CATEGORIAS_SEGUIMIENTO:
        flash('Categoría no reconocida.', 'error')
        return redirect(url_for('admin.seguimiento'))

    filas = [f for f in _usuarios_seguimiento() if _categoria(f) == categoria]
    for f in filas:
        f['tipo_msg'] = _tipo_mensaje(f, categoria)

    return render_template_string("""
<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{{ cat_label[categoria] }} - Seguimiento - Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css">
<style>.badge-seg { font-size: .72rem; }</style>
</head><body class="bg-light">
<div class="container-fluid py-4" style="max-width:1100px">
  <a href="{{ url_for('admin.seguimiento') }}" class="btn btn-outline-secondary btn-sm mb-2">← Volver a categorías</a>
  <h4 class="fw-bold mb-1">{{ cat_label[categoria] }} <span class="text-muted fs-6">({{ filas|length }})</span></h4>
  <p class="text-muted small mb-3">{{ cat_desc[categoria] }}</p>
  """ + _FLASH_BLOCK + """
  <div class="alert alert-warning small">
    <i class="bi bi-exclamation-triangle"></i> El WhatsApp (individual o "a todos") solo funciona una
    vez que la plantilla correspondiente esté <strong>aprobada en Meta Business Manager</strong>. Hasta
    entonces va a dar error acá arriba (se ve el detalle). Mientras tanto usá el email, o "Ver" para
    mandar el WhatsApp a mano.
  </div>

  <div class="d-flex gap-2 mb-3">
    <form method="POST" action="{{ url_for('admin.seguimiento_categoria_email_todos', categoria=categoria) }}"
          onsubmit="return confirm('¿Mandar el email a los {{ filas|length }} usuarios de esta lista?');">
      <button type="submit" class="btn btn-outline-primary btn-sm" {{ 'disabled' if not filas }}>
        <i class="bi bi-envelope"></i> Mail a todos ({{ filas|length }})
      </button>
    </form>
    <form method="POST" action="{{ url_for('admin.seguimiento_categoria_whatsapp_todos', categoria=categoria) }}"
          onsubmit="return confirm('¿Mandar el WhatsApp a los {{ filas|length }} usuarios de esta lista?');">
      <button type="submit" class="btn btn-success btn-sm" {{ 'disabled' if not filas }}>
        <i class="bi bi-whatsapp"></i> WhatsApp a todos ({{ filas|length }})
      </button>
    </form>
  </div>

  {% for f in filas %}
  <div class="card mb-2 shadow-sm">
    <div class="card-body py-2">
      <div class="row align-items-center g-2">
        <div class="col-md-3">
          <div class="fw-semibold">{{ f.nombre or '—' }}</div>
          <div class="small text-muted">{{ f.email }}</div>
          {% if f.telefono %}<div class="small text-success"><i class="bi bi-whatsapp"></i> {{ f.telefono }}</div>{% endif %}
          <div class="small text-muted">{{ f.ciudad or '' }}{{ ', ' if f.ciudad and f.provincia }}{{ f.provincia or '' }}</div>
        </div>
        <div class="col-md-3 small text-muted">
          Registrado: {{ f.created_at|local_dt('%d/%m/%Y') }}<br>Vence: {{ f.subscription_expires or '∞' }}
          {% if f.trial_por_vencer %}<br><span class="badge bg-warning text-dark badge-seg">{{ f.dias_restantes }}d / {{ f.presup_restantes }} presup. restantes</span>{% endif %}
          {% if f.suscripcion_vencida %}<br><span class="badge bg-danger badge-seg">Vencida ({{ f.subscription_expires }})</span>{% endif %}
        </div>
        <div class="col-md-2 small text-muted">
          {% if f.ultimo_contacto %}
            Último contacto:<br>{{ f.ultimo_contacto|local_dt }}
            {% if f.ultimo_resultado == 'ok' %}<span class="badge bg-success">enviado</span>
            {% elif f.ultimo_resultado == 'error' %}<span class="badge bg-danger">error</span>{% endif %}
          {% else %}Sin contactar todavía{% endif %}
        </div>
        <div class="col-md-3">
          <div class="d-flex gap-1 align-items-center">
            <span class="small text-muted" style="min-width:110px">{{ tipo_label[f.tipo_msg] }}:</span>
            <form method="POST" action="{{ url_for('admin.seguimiento_whatsapp', uid=f.id) }}">
              <input type="hidden" name="tipo" value="{{ f.tipo_msg }}">
              <button type="submit" class="btn btn-sm btn-success" {{ 'disabled title=Sin teléfono' if not f.telefono }}>
                <i class="bi bi-whatsapp"></i>
              </button>
            </form>
            <form method="POST" action="{{ url_for('admin.seguimiento_email', uid=f.id) }}">
              <input type="hidden" name="tipo" value="{{ f.tipo_msg }}">
              <button type="submit" class="btn btn-sm btn-outline-primary">
                <i class="bi bi-envelope"></i>
              </button>
            </form>
          </div>
        </div>
        <div class="col-md-1 text-end">
          <a href="{{ url_for('admin.seguimiento_detalle', uid=f.id) }}" class="btn btn-sm btn-outline-secondary">Ver</a>
        </div>
      </div>
    </div>
  </div>
  {% else %}
  <p class="text-muted text-center py-4">No hay usuarios en esta categoría.</p>
  {% endfor %}
</div></body></html>
""", filas=filas, categoria=categoria, cat_label=CATEGORIA_LABEL, cat_desc=CATEGORIA_DESC,
    tipo_label=TIPO_LABEL, user=g.user)


@bp.route('/seguimiento/categoria/<categoria>/email-a-todos', methods=['POST'])
@admin_required
def seguimiento_categoria_email_todos(categoria):
    """"Mail a todos" -- 1 click, manda el mail de retención correspondiente
    a CADA usuario de la categoría de una sola vez (confirmado por Daniel
    06/08/2026: sin revisar mensaje por mensaje). Dentro de "ESTUVO USANDO"
    cada usuario recibe el mensaje que le corresponde según su uso real (1
    presup./borrador, solo Costo/m², o check-in si ya tiene 2+) -- no es un
    único texto genérico para toda la categoría."""
    if categoria not in CATEGORIAS_SEGUIMIENTO:
        flash('Categoría no reconocida.', 'error')
        return redirect(url_for('admin.seguimiento'))

    db = get_db()
    filas = [f for f in _usuarios_seguimiento() if _categoria(f) == categoria]
    enviados = errores = salteados = 0
    for f in filas:
        tipo = _tipo_mensaje(f, categoria)
        # 07/08/2026: no repetir a quien ya le llegó ESTE mensaje puntual --
        # ver _ya_contactado(), pensado para no duplicar contra las tandas
        # horarias automáticas.
        if _ya_contactado(db, f['id'], 'email', tipo):
            salteados += 1
            continue
        ok, _ = _enviar_email_seguimiento(db, f, tipo)
        enviados += 1 if ok else 0
        errores += 0 if ok else 1
    db.commit()
    db.close()
    flash(f'Mail a todos ({CATEGORIA_LABEL[categoria]}): {enviados} enviados, {errores} con error, '
          f'{salteados} salteados (ya lo habían recibido, ej. por la tanda automática).',
          'success' if errores == 0 else 'error')
    return redirect(url_for('admin.seguimiento_categoria', categoria=categoria))


@bp.route('/seguimiento/categoria/<categoria>/whatsapp-a-todos', methods=['POST'])
@admin_required
def seguimiento_categoria_whatsapp_todos(categoria):
    """"WhatsApp a todos" -- mismo criterio que el mail (1 click, sin
    revisar uno por uno). Los que no tengan teléfono cargado o cuya
    plantilla no esté aprobada en Meta cuentan como error (se ve el detalle
    en el Historial de cada usuario, Seguimiento > Ver) -- no cortan el
    envío al resto de la lista."""
    if categoria not in CATEGORIAS_SEGUIMIENTO:
        flash('Categoría no reconocida.', 'error')
        return redirect(url_for('admin.seguimiento'))

    db = get_db()
    filas = [f for f in _usuarios_seguimiento() if _categoria(f) == categoria]
    enviados = errores = salteados = 0
    for f in filas:
        tipo = _tipo_mensaje(f, categoria)
        # Mismo chequeo que "Mail a todos" -- hoy WhatsApp no tiene tanda
        # automática todavía, pero deja esto listo para cuando la tenga y
        # evita duplicar si Daniel lo clickea 2 veces seguidas.
        if _ya_contactado(db, f['id'], 'whatsapp', tipo):
            salteados += 1
            continue
        ok, detalle, plantilla = _enviar_whatsapp_seguimiento(db, f, tipo)
        enviados += 1 if ok else 0
        errores += 0 if ok else 1
    db.commit()
    db.close()
    flash(f'WhatsApp a todos ({CATEGORIA_LABEL[categoria]}): {enviados} enviados, {errores} con error '
          f'(sin teléfono o plantilla no aprobada en Meta), {salteados} salteados (ya contactados).',
          'success' if errores == 0 else 'error')
    return redirect(url_for('admin.seguimiento_categoria', categoria=categoria))


@bp.route('/seguimiento/<int:uid>/generar-promo', methods=['POST'])
@admin_required
def seguimiento_generar_promo(uid):
    """Genera (o reusa el vigente) el link de pago con 50% off por 48hs para
    este usuario -- botón en su perfil, solo tiene sentido para vencidos
    D/B (conversion_d/conversion_b), pero no se restringe duro por si hace
    falta para otro caso puntual. Reusa uno existente si ya hay uno sin usar
    y sin vencer (ver _crear_promo), así clickearlo de nuevo no genera un
    2do link distinto -- pero si ya venció, uno nuevo le da otras 48hs
    frescas desde ahora."""
    from routes.pagos import _crear_promo, _art_str
    db = get_db()
    u = db.execute("SELECT nombre, email FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        db.close()
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('admin.seguimiento'))
    promo = _crear_promo(db, uid, horas=48, descuento_pct=50)
    db.close()
    app_url = os.environ.get('APP_BASE_URL', 'https://web-production-0c9c1.up.railway.app')
    link = f"{app_url}/pagos/promo/{promo['token']}"
    vence_str = _art_str(promo['vence_at'], '%d/%m a las %H:%M')
    flash(f'Link generado (válido hasta el {vence_str} ART): {link} — ya se suma solo al mail de este usuario; '
          f'para WhatsApp, copialo y pegalo en tu respuesta.', 'success')
    return redirect(url_for('admin.seguimiento_detalle', uid=uid))


@bp.route('/seguimiento/<int:uid>/opt-out', methods=['POST'])
@admin_required
def seguimiento_opt_out(uid):
    """Marcar/desmarcar "Baja de retención" -- agregado 07/08/2026, pedido
    de Daniel al ver varias respuestas de "Baja" en Admin > WhatsApp desde
    que los mails son automáticos. Mientras esté marcado, este usuario
    desaparece de _usuarios_seguimiento() (Seguimiento, tanda automática de
    mails y los botones "a todos" -- las 3 comparten esa función), así no
    se lo vuelve a contactar por error. NO borra el historial de
    retencion_contactos ni impide mandarle un mail/whatsapp puntual a mano
    desde esta misma pantalla si hiciera falta (eso queda como excepción
    explícita, no automática)."""
    valor = 1 if request.form.get('accion') == 'baja' else 0
    db = get_db()
    db.execute("UPDATE users SET retencion_opt_out=? WHERE id=?", (valor, uid))
    db.commit()
    db.close()
    flash('Usuario dado de BAJA de retención -- no se lo va a volver a contactar automático.' if valor
          else 'Baja de retención revertida -- vuelve a aparecer en Seguimiento.', 'success')
    return redirect(url_for('admin.seguimiento_detalle', uid=uid))


@bp.route('/seguimiento/<int:uid>')
@admin_required
def seguimiento_detalle(uid):
    """Agregado 21/07/2026, pedido de Daniel: poder ver toda la actividad de
    un usuario (perfil, presupuestos/borradores/costo_m2, historial de
    contactos previos) y el texto sugerido de cada mensaje EN UN TEXTAREA
    EDITABLE antes de mandarlo -- para agregar algo puntual o corregir, en
    vez de que se mande el texto fijo de una. El email manda lo que quede
    escrito en el textarea al momento de tocar "Enviar" (ver
    seguimiento_email). El WhatsApp por plantilla NO puede llevar texto
    editado -- Meta solo permite rellenar las variables ({{1}}=nombre) de
    una plantilla ya aprobada, no cambiar el cuerpo -- así que para el
    textarea de WhatsApp se ofrece en cambio "Abrir en WhatsApp (manual)":
    abre wa.me con el texto editado, listo para que Daniel lo mande él
    mismo desde su teléfono, sin depender de que la plantilla esté
    aprobada. Esa opción funciona HOY."""
    db = get_db()
    u = db.execute(
        """SELECT u.*,
                  (SELECT COUNT(*) FROM presupuestos p WHERE p.user_id=u.id AND p.status='completo'
                     AND (p.es_demo IS NULL OR p.es_demo=0))                                          AS n_presupuestos,
                  (SELECT COUNT(*) FROM presupuestos p WHERE p.user_id=u.id AND p.status='borrador')  AS n_borradores,
                  (SELECT COUNT(*) FROM costo_m2_consultas c WHERE c.user_id=u.id)                    AS n_costo_m2
           FROM users u WHERE u.id=?""",
        (uid,)
    ).fetchone()
    if not u:
        db.close()
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('admin.seguimiento'))

    historial = db.execute(
        "SELECT * FROM retencion_contactos WHERE user_id=? ORDER BY created_at DESC", (uid,)
    ).fetchall()

    # Fix 24/07/2026, pedido de Daniel: probó el flujo real con un usuario
    # (Rafael) que respondió por WhatsApp a un mensaje de retención, y esa
    # respuesta no aparecía en ninguna parte de ESTA pantalla — solo estaba
    # en Admin > WhatsApp, una página distinta. Acá se agrega también,
    # cruzando por teléfono (mismo criterio que whatsapp_inbox).
    eventos = [dict(h, tipo_evento='enviado') for h in historial]
    if u['telefono']:
        tel_norm = telefono_normalizado(u['telefono'])
        for c in db.execute(
            "SELECT * FROM whatsapp_consultas_sin_responder ORDER BY created_at DESC"
        ).fetchall():
            if telefono_normalizado(c['telefono']) == tel_norm:
                eventos.append(dict(c, tipo_evento='recibido'))
    eventos.sort(key=lambda e: e['created_at'] or '', reverse=True)
    db.close()

    fila = dict(u)
    fila['segmento'] = _segmento(u)
    # Fix 07/08/2026 (cont. 24): date.today() usa la hora del servidor (UTC en
    # Railway), no ART -- en la franja 21:00-23:59 ART del día de vencimiento,
    # marcaba "suscripción vencida" hasta 3hs antes de tiempo (afecta el
    # segmento VENCIDOS de Admin > Seguimiento y sus envíos masivos). Mismo
    # ajuste -3hs que en utils/auth.py.
    hoy_str = (datetime.utcnow() - timedelta(hours=3)).date().isoformat()
    ahora = datetime.utcnow()
    trial_por_vencer = False
    dias_restantes = presup_restantes = None
    if u['es_trial']:
        try:
            creado = datetime.fromisoformat((u['created_at'] or '').replace(' ', 'T'))
            dias_pasados = (ahora - creado).days
        except (ValueError, TypeError):
            dias_pasados = 0
        dias_restantes = max(0, 14 - dias_pasados)
        presup_restantes = max(0, 3 - u['n_presupuestos'])
        vencido_trial = u['n_presupuestos'] >= 3 or dias_pasados >= 14
        trial_por_vencer = (not vencido_trial) and (dias_restantes <= 3 or presup_restantes <= 1)
    fila['trial_por_vencer'] = trial_por_vencer
    fila['dias_restantes'] = dias_restantes
    fila['presup_restantes'] = presup_restantes
    fila['suscripcion_vencida'] = bool(u['subscription_expires']) and u['subscription_expires'] < hoy_str
    tipos = _tipos_aplicables(fila)

    # 07/08/2026: si hay un link de pago con descuento vigente para este
    # usuario, se muestra acá y se suma a los mensajes de conversion_d/b
    # (ver _link_promo_vigente arriba).
    db2 = get_db()
    link_promo, vence_promo = _link_promo_vigente(db2, uid)
    db2.close()

    mensajes = {}
    for tipo in tipos:
        if tipo in ('conversion_d', 'conversion_b'):
            wa_msg, email_msg = MENSAJES_EMAIL[tipo](u['nombre'], link_promo, vence_promo)
        else:
            wa_msg, email_msg = MENSAJES_EMAIL[tipo](u['nombre'])
        mensajes[tipo] = {'wa': wa_msg, 'email': email_msg}

    return render_template_string("""
<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{{ f.nombre or f.email }} - Seguimiento</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css">
</head><body class="bg-light">
<div class="container py-4" style="max-width:760px">
  <a href="{{ url_for('admin.seguimiento', todos=1) }}" class="btn btn-outline-secondary btn-sm mb-3">Volver a Seguimiento</a>
  """ + _FLASH_BLOCK + """
  <div class="card mb-3">
    <div class="card-body">
      <h5 class="fw-bold mb-1">{{ f.nombre or '—' }}</h5>
      <div class="text-muted">{{ f.email }}</div>
      {% if f.telefono %}<div class="text-success"><i class="bi bi-whatsapp"></i> {{ f.telefono }}</div>{% endif %}
      <div class="small text-muted mt-1">
        {{ f.ciudad or '' }}{% if f.ciudad and f.provincia %}, {% endif %}{{ f.provincia or '' }}
      </div>
      <hr>
      <div class="row small">
        <div class="col-4"><strong>Registrado:</strong><br>{{ f.created_at|local_dt('%d/%m/%Y') }}</div>
        <div class="col-4"><strong>Vence:</strong><br>{{ f.subscription_expires or '∞' }}</div>
        <div class="col-4"><strong>Segmento:</strong><br><span class="badge bg-secondary">{{ f.segmento }}</span></div>
      </div>
      <div class="row small mt-2">
        <div class="col-4"><strong>Presupuestos:</strong> {{ f.n_presupuestos }}</div>
        <div class="col-4"><strong>Borradores:</strong> {{ f.n_borradores }}</div>
        <div class="col-4"><strong>Costo/m² usados:</strong> {{ f.n_costo_m2 }}</div>
      </div>
      {% if f.trial_por_vencer %}
      <div class="alert alert-warning small mt-2 mb-0">Prueba gratis por vencer: {{ f.dias_restantes }} días
        o {{ f.presup_restantes }} presupuesto(s) restantes.</div>
      {% endif %}
      {% if f.suscripcion_vencida %}
      <div class="alert alert-danger small mt-2 mb-0">Suscripción vencida el {{ f.subscription_expires }}.</div>
      {% endif %}
      {% if f.retencion_opt_out %}
      <div class="alert alert-secondary small mt-2 mb-0 d-flex justify-content-between align-items-center">
        <span><i class="bi bi-slash-circle"></i> Usuario dado de <strong>BAJA</strong> de retención -- no se lo contacta automático.</span>
        <form method="POST" action="{{ url_for('admin.seguimiento_opt_out', uid=f.id) }}" class="ms-2">
          <input type="hidden" name="accion" value="alta">
          <button type="submit" class="btn btn-sm btn-outline-secondary">Revertir</button>
        </form>
      </div>
      {% else %}
      <form method="POST" action="{{ url_for('admin.seguimiento_opt_out', uid=f.id) }}" class="mt-2"
            onsubmit="return confirm('¿Dar de baja a {{ f.nombre or f.email }} de retención? No se lo va a volver a contactar automático (podés revertirlo después).');">
        <input type="hidden" name="accion" value="baja">
        <button type="submit" class="btn btn-sm btn-outline-danger"><i class="bi bi-slash-circle"></i> Dar de baja de retención</button>
      </form>
      {% endif %}
    </div>
  </div>

  {% if not tipos %}
  <p class="text-muted">Este usuario no tiene ningún mensaje de retención sugerido en este momento
    (ya es un usuario activo, o no encaja en ningún segmento).</p>
  {% endif %}

  {% for tipo in tipos %}
  <div class="card mb-3">
    <div class="card-header fw-bold">{{ tipo_label[tipo] }}</div>
    <div class="card-body">
      {% if tipo in ('conversion_d', 'conversion_b') %}
        {% if link_promo %}
        <div class="alert alert-success small">
          <i class="bi bi-tag"></i> Link con 50% off vigente hasta el <strong>{{ vence_promo }}</strong> (ART) --
          ya está sumado al mensaje de abajo. <a href="{{ link_promo }}" target="_blank">{{ link_promo }}</a>
        </div>
        {% else %}
        <div class="alert alert-warning small">
          <i class="bi bi-exclamation-triangle"></i> Todavía no generaste el link de pago con descuento para este usuario.
        </div>
        {% endif %}
        <form method="POST" action="{{ url_for('admin.seguimiento_generar_promo', uid=f.id) }}" class="mb-3">
          <button type="submit" class="btn btn-sm btn-outline-success">
            <i class="bi bi-tag"></i> {{ 'Generar otro link (48hs frescas)' if link_promo else 'Generar link 50% off (48hs)' }}
          </button>
        </form>
      {% endif %}
      <form method="POST" action="{{ url_for('admin.seguimiento_email', uid=f.id) }}" class="mb-3">
        <input type="hidden" name="tipo" value="{{ tipo }}">
        <label class="form-label small text-muted">Mensaje por email (editable):</label>
        <textarea name="mensaje" class="form-control mb-2" rows="4">{{ mensajes[tipo].email }}</textarea>
        <button type="submit" class="btn btn-sm btn-primary"><i class="bi bi-envelope"></i> Enviar email</button>
      </form>
      <div>
        <label class="form-label small text-muted">Mensaje por WhatsApp:</label>
        <textarea id="wa-{{ tipo }}" class="form-control mb-2" rows="3">{{ mensajes[tipo].wa }}</textarea>
        <button type="button" class="btn btn-sm btn-success"
                onclick="abrirWhatsapp('{{ f.telefono|e }}', document.getElementById('wa-{{ tipo }}').value)"
                {{ 'disabled title=Sin teléfono' if not f.telefono }}>
          <i class="bi bi-whatsapp"></i> Abrir en WhatsApp (manual, funciona ya)
        </button>
        <form method="POST" action="{{ url_for('admin.seguimiento_whatsapp', uid=f.id) }}" class="d-inline">
          <input type="hidden" name="tipo" value="{{ tipo }}">
          <button type="submit" class="btn btn-sm btn-outline-success" {{ 'disabled title=Sin teléfono' if not f.telefono }}>
            Enviar por plantilla API (necesita aprobación de Meta)
          </button>
        </form>
      </div>
    </div>
  </div>
  {% endfor %}

  <div class="card">
    <div class="card-header fw-bold">Historial de contactos</div>
    <ul class="list-group list-group-flush">
      {% for e in eventos %}
      {% if e.tipo_evento == 'enviado' %}
      <li class="list-group-item small">
        <strong>{{ e.created_at|local_dt }}</strong> — {{ e.canal }} ({{ tipo_label.get(e.segmento, e.segmento) }})
        <span class="badge {{ 'bg-success' if e.resultado == 'ok' else 'bg-danger' }}">{{ e.resultado }}</span>
        <div class="text-muted">{{ e.mensaje }}</div>
      </li>
      {% else %}
      <li class="list-group-item small bg-light">
        <strong>{{ e.created_at|local_dt }}</strong> — <i class="bi bi-reply"></i> respondió por WhatsApp
        {% if e.respondida %}<span class="badge bg-success">respondida</span>
        {% else %}<span class="badge bg-warning text-dark">sin responder</span>{% endif %}
        <div>{{ e.mensaje }}</div>
        {% if e.respondida %}<div class="text-muted"><strong>Tu respuesta:</strong> {{ e.respuesta_admin }}</div>
        {% else %}<a href="{{ url_for('admin.whatsapp_inbox') }}" class="small">Responder en Admin &gt; WhatsApp →</a>{% endif %}
      </li>
      {% endif %}
      {% else %}
      <li class="list-group-item text-muted small">Sin contactos registrados todavía.</li>
      {% endfor %}
    </ul>
  </div>
</div>
<script>
function abrirWhatsapp(tel, mensaje) {
  let num = (tel || '').replace(/[^0-9+]/g, '');
  if (!num.startsWith('+') && !num.startsWith('54')) num = '549' + num;
  else if (num.startsWith('54') && !num.startsWith('549')) num = '549' + num.slice(2);
  num = num.replace(/^\\+/, '');
  window.open('https://wa.me/' + num + '?text=' + encodeURIComponent(mensaje), '_blank');
}
</script>
</body></html>
""", f=fila, tipos=tipos, mensajes=mensajes, tipo_label=TIPO_LABEL, eventos=eventos, user=g.user,
    link_promo=link_promo, vence_promo=vence_promo)


def _link_promo_vigente(db, user_id):
    """(link, vence_str_ART) del último link de pago con descuento vigente
    para este usuario, o (None, None) si nunca se generó uno o ya venció/se
    usó. Import perezoso de routes.pagos (mismo criterio que
    routes.whatsapp_bot en otras funciones de este archivo) para evitar
    import circular a nivel de módulo."""
    from routes.pagos import _art_str
    promo = db.execute(
        "SELECT * FROM retencion_promos WHERE user_id=? AND usado=0 AND vence_at > datetime('now') "
        "ORDER BY creado_at DESC LIMIT 1",
        (user_id,)
    ).fetchone()
    if not promo:
        return None, None
    app_url = os.environ.get('APP_BASE_URL', 'https://web-production-0c9c1.up.railway.app')
    link = f"{app_url}/pagos/promo/{promo['token']}"
    vence_str = _art_str(promo['vence_at'], '%d/%m a las %H:%M')
    return link, vence_str


def _ya_contactado(db, user_id, canal, segmento):
    """True si ya existe un envío OK registrado en retencion_contactos para
    ese usuario, canal y tipo puntual -- agregado 07/08/2026, pedido de
    Daniel: desde que utils/recordatorios.py::enviar_backlog_email_segmentos
    manda los mails de A/C/trial/estuvo_usando solo por tandas horarias, los
    botones "a todos" de acá (que mandan a TODA la categoría sin fijarse
    quién ya recibió el automático) duplicaban el envío. Con este chequeo,
    "a todos" pasa a ser "a todos los que todavía no tienen ESTE mensaje
    puntual" -- sigue sirviendo para mandar ya (sin esperar la próxima
    tanda) a los que falten, sin repetirle a nadie."""
    return db.execute(
        "SELECT 1 FROM retencion_contactos WHERE user_id=? AND canal=? AND segmento=? AND resultado='ok' LIMIT 1",
        (user_id, canal, segmento)
    ).fetchone() is not None


def _enviar_whatsapp_seguimiento(db, u, tipo):
    """Manda 1 WhatsApp de retención (plantilla de Meta) a un usuario y
    guarda el resultado en retencion_contactos. Devuelve (ok, detalle,
    plantilla). NO hace commit ni close -- lo maneja quien llama, para poder
    loopear varios usuarios en una sola transacción (ver
    seguimiento_categoria_whatsapp_todos). Extraído 06/08/2026 del handler
    de 1 solo usuario (seguimiento_whatsapp) para reusar en el envío masivo
    "WhatsApp a todos" del rediseño de Seguimiento."""
    plantilla = TEMPLATES_WHATSAPP.get(tipo)
    if not plantilla:
        return False, 'Tipo de mensaje no reconocido.', None
    if not u['telefono']:
        return False, 'Sin teléfono cargado.', plantilla

    from routes.whatsapp_bot import enviar_plantilla_whatsapp
    ok, detalle = enviar_plantilla_whatsapp(u['telefono'], plantilla, parametros={'nombre': u['nombre'] or ''})
    # Fix 24/07/2026: se guarda el texto real (no solo el nombre de la
    # plantilla) para que el Historial (Seguimiento > Ver) muestre lo mismo
    # que se mandó, igual que el email. Nota 07/08/2026: para conversion_d/b
    # esto queda como el texto SIN link (no se busca la promo acá) -- lo que
    # de verdad se manda por la API de Meta es siempre el texto fijo
    # aprobado, este valor es solo para el registro/historial.
    generador = MENSAJES_EMAIL.get(tipo)
    wa_texto, _ = generador(u['nombre']) if generador else ('', '')
    mensaje_guardado = wa_texto if ok else f"{plantilla} — ERROR: {detalle}"
    db.execute(
        "INSERT INTO retencion_contactos (user_id, canal, segmento, mensaje, resultado) VALUES (?,?,?,?,?)",
        (u['id'], 'whatsapp', tipo, mensaje_guardado, 'ok' if ok else 'error')
    )
    return ok, detalle, plantilla


@bp.route('/seguimiento/<int:uid>/whatsapp', methods=['POST'])
@admin_required
def seguimiento_whatsapp(uid):
    tipo = request.form.get('tipo', '')
    volver = request.form.get('volver') or url_for('admin.seguimiento')

    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        db.close()
        flash('Usuario no encontrado.', 'error')
        return redirect(volver)

    ok, detalle, plantilla = _enviar_whatsapp_seguimiento(db, u, tipo)
    db.commit()
    db.close()
    if ok:
        flash(f'WhatsApp ({plantilla}) enviado a {u["nombre"] or u["email"]}.', 'success')
    elif not plantilla:
        flash(detalle, 'error')
    else:
        flash(f'No se pudo enviar "{plantilla}": {detalle}. Mientras tanto usá "Abrir en WhatsApp '
              f'(manual)" desde el detalle del usuario.', 'error')
    return redirect(request.referrer or url_for('admin.seguimiento'))


def _enviar_email_seguimiento(db, u, tipo, mensaje_override=None):
    """Manda 1 email de retención (Resend) a un usuario y guarda el
    resultado en retencion_contactos. Devuelve (ok, cuerpo_email). NO hace
    commit ni close -- lo maneja quien llama. Extraído 06/08/2026 del handler
    de 1 solo usuario (seguimiento_email) para reusar en el envío masivo
    "Mail a todos" del rediseño de Seguimiento (que nunca manda
    mensaje_override -- eso es solo para el textarea editable de
    seguimiento_detalle)."""
    generador = MENSAJES_EMAIL.get(tipo)
    if not generador:
        return False, None

    if mensaje_override:
        cuerpo_email = mensaje_override
    elif tipo in ('conversion_d', 'conversion_b'):
        # 07/08/2026: si ya existe un link de pago con descuento generado
        # para este usuario (botón "Generar link 50% (48hs)" en su perfil),
        # se lo suma al mail -- si todavía no lo generaste, manda el texto
        # sin link ("respondé este mail y te ayudamos").
        link, vence_str = _link_promo_vigente(db, u['id'])
        _, cuerpo_email = generador(u['nombre'], link, vence_str)
    else:
        _, cuerpo_email = generador(u['nombre'])

    ok = False
    api_key = os.environ.get('RESEND_API_KEY')
    if api_key:
        try:
            import resend
            resend.api_key = api_key
            resp = resend.Emails.send({
                "from": "PresupuestoPRO <noreply@presupuestopro.com.ar>",
                "to": [u['email']],
                "subject": "PresupuestoPRO",
                "text": cuerpo_email,
                "reply_to": ["contacto@presupuestopro.com.ar"],
            })
            registrar_envio(resp.get('id', ''), u['email'], f'seguimiento_{tipo}')
            ok = True
        except Exception as e:
            print(f"[seguimiento_email] error enviando a {u['email']}: {e}")

    db.execute(
        "INSERT INTO retencion_contactos (user_id, canal, segmento, mensaje, resultado) VALUES (?,?,?,?,?)",
        (u['id'], 'email', tipo, cuerpo_email[:500], 'ok' if ok else 'error')
    )
    return ok, cuerpo_email


@bp.route('/seguimiento/<int:uid>/email', methods=['POST'])
@admin_required
def seguimiento_email(uid):
    """Fix 21/07/2026, pedido de Daniel: antes esto SIEMPRE regeneraba el
    texto fijo de la plantilla interna, ignorando cualquier edición. Ahora,
    si el form manda 'mensaje' (viene del textarea editable de
    seguimiento_detalle), se usa ESE texto tal cual; si no viene (por
    ejemplo el botón rápido de la lista, que no tiene textarea), se genera
    el texto por default como antes."""
    tipo = request.form.get('tipo', '')
    if tipo not in MENSAJES_EMAIL:
        flash('Tipo de mensaje no reconocido.', 'error')
        return redirect(request.referrer or url_for('admin.seguimiento'))

    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        db.close()
        flash('Usuario no encontrado.', 'error')
        return redirect(request.referrer or url_for('admin.seguimiento'))

    mensaje_editado = (request.form.get('mensaje') or '').strip()
    ok, _ = _enviar_email_seguimiento(db, u, tipo, mensaje_override=mensaje_editado or None)
    db.commit()
    db.close()
    if ok:
        flash(f'Email enviado a {u["email"]}.', 'success')
    else:
        flash('No se pudo enviar el email (revisar RESEND_API_KEY en Railway).', 'error')
    return redirect(request.referrer or url_for('admin.seguimiento'))


@bp.route('/localidades')
@admin_required
def localidades():
    db = get_db()
    filas = db.execute("SELECT * FROM localidades ORDER BY merged_en != '', veces_usada DESC").fetchall()
    # Para poder mostrar "fusionada en: <nombre>" en vez de solo la clave.
    por_clave = {f['clave_normalizada']: f for f in filas}
    db.close()
    return render_template('admin/localidades.html', filas=filas, por_clave=por_clave, user=g.user)


@bp.route('/localidades/<int:lid>/renombrar', methods=['POST'])
@admin_required
def localidad_renombrar(lid):
    nuevo_nombre = (request.form.get('nombre_display') or '').strip()
    if not nuevo_nombre:
        flash('El nombre no puede quedar vacío.', 'error')
        return redirect(url_for('admin.localidades'))
    db = get_db()
    fila = db.execute("SELECT * FROM localidades WHERE id=?", (lid,)).fetchone()
    if not fila:
        db.close()
        flash('No se encontró esa localidad.', 'error')
        return redirect(url_for('admin.localidades'))
    # Los usuarios que ya tenían la grafía vieja se actualizan también, para
    # que no quede desincronizado lo que ve Daniel acá vs. lo que dice cada
    # usuario en Admin > Usuarios.
    db.execute("UPDATE users SET ciudad=? WHERE ciudad=?", (nuevo_nombre, fila['nombre_display']))
    db.execute("UPDATE localidades SET nombre_display=? WHERE id=?", (nuevo_nombre, lid))
    db.commit()
    db.close()
    flash(f'Renombrada a "{nuevo_nombre}".', 'success')
    return redirect(url_for('admin.localidades'))


@bp.route('/localidades/fusionar', methods=['POST'])
@admin_required
def localidad_fusionar():
    """Fusiona `origen_id` hacia `destino_id`: todos los usuarios que tenían
    la grafía de origen pasan a la de destino, y de acá en más cualquiera que
    se registre escribiendo la clave de origen también cae en destino (ver
    routes/landing.py::_guardar_localidad, sigue la cadena `merged_en`)."""
    try:
        origen_id = int(request.form.get('origen_id'))
        destino_id = int(request.form.get('destino_id'))
    except (TypeError, ValueError):
        flash('Elegí las dos localidades a fusionar.', 'error')
        return redirect(url_for('admin.localidades'))
    if origen_id == destino_id:
        flash('Elegí dos localidades distintas.', 'error')
        return redirect(url_for('admin.localidades'))

    db = get_db()
    origen = db.execute("SELECT * FROM localidades WHERE id=?", (origen_id,)).fetchone()
    destino = db.execute("SELECT * FROM localidades WHERE id=?", (destino_id,)).fetchone()
    if not origen or not destino:
        db.close()
        flash('No se encontró alguna de las dos localidades.', 'error')
        return redirect(url_for('admin.localidades'))

    db.execute("UPDATE users SET ciudad=? WHERE ciudad=?", (destino['nombre_display'], origen['nombre_display']))
    db.execute(
        "UPDATE localidades SET merged_en=?, veces_usada=0 WHERE id=?",
        (destino['clave_normalizada'], origen_id)
    )
    db.execute(
        "UPDATE localidades SET veces_usada = veces_usada + ? WHERE id=?",
        (origen['veces_usada'], destino_id)
    )
    db.commit()
    db.close()
    flash(f'"{origen["nombre_display"]}" fusionada en "{destino["nombre_display"]}".', 'success')
    return redirect(url_for('admin.localidades'))

@bp.route('/usuarios/nuevo', methods=['GET', 'POST'])
@admin_required
def usuario_nuevo():
    if request.method == 'POST':
        email     = request.form.get('email', '').strip().lower()
        nombre    = request.form.get('nombre', '')
        telefono  = request.form.get('telefono', '').strip()
        ciudad    = request.form.get('ciudad', '').strip()
        provincia = request.form.get('provincia', '').strip()
        password  = request.form.get('password', '')
        pais      = request.form.get('pais', 'AR')
        vence     = request.form.get('subscription_expires', '')
        # Fix 05/08/2026: Daniel detectó un usuario con el email pegado en el
        # campo Teléfono/WhatsApp — no había ninguna validación de formato acá.
        if telefono and not telefono_valido(telefono):
            flash('El teléfono no es válido (solo números, sin letras ni email).', 'error')
            return redirect(url_for('admin.usuario_nuevo'))
        db = get_db()
        try:
            db.execute(
                "INSERT INTO users (email, password_hash, nombre, telefono, ciudad, provincia, pais, active, subscription_expires) VALUES (?,?,?,?,?,?,?,1,?)",
                (email, generate_password_hash(password), nombre, telefono, ciudad, provincia, pais, vence or None)
            )
            db.commit()
            flash(f'Usuario {email} creado.', 'success')
        except Exception as e:
            flash(f'Error: {e}', 'error')
        finally:
            db.close()
        return redirect(url_for('admin.usuarios'))
    return render_template('admin/usuario_form.html', u=None, user=g.user,
                           now_date=date.today(), timedelta=timedelta)

@bp.route('/usuarios/<int:uid>/editar', methods=['GET', 'POST'])
@admin_required
def usuario_editar(uid):
    db = get_db()
    u = db.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        db.close(); return redirect(url_for('admin.usuarios'))

    if request.method == 'POST':
        nombre    = request.form.get('nombre', '')
        telefono  = request.form.get('telefono', '').strip()
        ciudad    = request.form.get('ciudad', '').strip()
        provincia = request.form.get('provincia', '').strip()
        pais      = request.form.get('pais', 'AR')
        active    = 1 if request.form.get('active') else 0
        vence     = request.form.get('subscription_expires', '')
        new_pw    = request.form.get('password', '').strip()
        # Fix 05/08/2026: mismo chequeo que en usuario_nuevo (ver docstring de
        # telefono_valido) — evita guardar un email u otro texto en el campo.
        if telefono and not telefono_valido(telefono):
            db.close()
            flash('El teléfono no es válido (solo números, sin letras ni email).', 'error')
            return redirect(url_for('admin.usuario_editar', uid=uid))
        if new_pw:
            db.execute(
                "UPDATE users SET nombre=?, telefono=?, ciudad=?, provincia=?, pais=?, active=?, subscription_expires=?, password_hash=? WHERE id=?",
                (nombre, telefono, ciudad, provincia, pais, active, vence or None, generate_password_hash(new_pw), uid)
            )
        else:
            db.execute(
                "UPDATE users SET nombre=?, telefono=?, ciudad=?, provincia=?, pais=?, active=?, subscription_expires=? WHERE id=?",
                (nombre, telefono, ciudad, provincia, pais, active, vence or None, uid)
            )
        if not active:
            db.execute("UPDATE users SET session_token=NULL WHERE id=?", (uid,))
        db.commit()
        flash('Usuario actualizado.', 'success')
        db.close()
        return redirect(url_for('admin.usuarios'))

    db.close()
    return render_template('admin/usuario_form.html', u=u, user=g.user)

@bp.route('/usuarios/<int:uid>/eliminar', methods=['POST'])
@admin_required
def usuario_eliminar(uid):
    """Borra un usuario y todo lo asociado. Agregado 10/07/2026 — pedido de
    Daniel para poder limpiar las cuentas ficticias/de prueba que va cargando
    mientras testea el flujo de validación de cuenta, sin tener que pedirle
    a un dev que lo haga a mano en la base. Irreversible: borra en cascada
    presupuestos, suscripciones, consultas de costo/m2, sugerencias, perfil
    de empresa, tokens de reset de password y códigos de verificación del
    usuario, antes de borrar la fila de users."""
    db = get_db()
    u = db.execute("SELECT email, nombre, is_admin FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        db.close()
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('admin.usuarios'))
    if u['is_admin']:
        db.close()
        flash('No se puede eliminar una cuenta de administrador desde acá.', 'error')
        return redirect(url_for('admin.usuarios'))

    for tabla in ('presupuestos', 'suscripciones', 'costo_m2_consultas', 'sugerencias',
                  'empresa_perfil', 'password_reset_tokens', 'verificacion_codigos',
                  'retencion_contactos'):
        db.execute(f"DELETE FROM {tabla} WHERE user_id=?", (uid,))
    db.execute("DELETE FROM users WHERE id=?", (uid,))
    db.commit()
    db.close()
    flash(f"Usuario {u['email']} eliminado.", 'success')
    return redirect(url_for('admin.usuarios'))


@bp.route('/usuarios/<int:uid>/enviar-activacion', methods=['POST'])
@admin_required
def usuario_enviar_activacion(uid):
    db = get_db()
    u = db.execute("SELECT email, nombre, subscription_expires FROM users WHERE id=?", (uid,)).fetchone()
    if not u:
        db.close()
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('admin.usuarios'))

    # Si no tiene fecha de vencimiento, asignar hoy + 30 días y activar
    exp_str = u['subscription_expires']
    if not exp_str:
        exp_str = (date.today() + timedelta(days=30)).isoformat()
        db.execute(
            "UPDATE users SET active=1, subscription_expires=? WHERE id=?",
            (exp_str, uid)
        )
        db.commit()

    db.close()

    from routes.pagos import _enviar_email_activacion
    from datetime import datetime
    exp_display = exp_str
    try:
        exp_display = datetime.strptime(exp_str, '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        pass

    ok = _enviar_email_activacion(
        user_email=u['email'],
        user_nombre=u['nombre'],
        fecha_vencimiento=exp_display,
    )
    if ok:
        flash(f'Email de activacion enviado a {u["email"]}.', 'success')
    else:
        flash('No se pudo enviar el email (revisar RESEND_API_KEY).', 'error')
    return redirect(url_for('admin.usuarios'))


# CONTACTOS
@bp.route('/contactos')
@admin_required
def contactos():
    db = get_db()
    msgs = db.execute("SELECT * FROM contactos ORDER BY created_at DESC").fetchall()
    db.execute("UPDATE contactos SET leido=1 WHERE leido=0")
    db.commit()
    db.close()
    return render_template_string("""
<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Mensajes - Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
  .card-nuevo  { border-left: 4px solid #0d6efd; }
  .card-leido  { border-left: 4px solid #dee2e6; }
  .card-contest{ border-left: 4px solid #198754; }
</style>
</head><body class="bg-light">
<div class="container py-4" style="max-width:760px">
  <a href="/admin/" class="btn btn-outline-secondary btn-sm mb-3">Volver</a>
  <h4 class="fw-bold mb-1">Mensajes de contacto</h4>
  <p class="text-muted small mb-3">
    <span class="badge bg-primary">{{ msgs|length }} total</span>
    <span class="badge bg-success ms-1">{{ msgs|selectattr('contestado','equalto',1)|list|length }} contestados</span>
    <span class="badge bg-warning text-dark ms-1">{{ msgs|selectattr('leido','equalto',0)|list|length }} nuevos</span>
  </p>
  {% if not msgs %}<p class="text-muted">No hay mensajes aun.</p>{% endif %}
  {% for m in msgs %}
  {% set card_class = 'card-contest' if m.contestado else ('card-nuevo' if not m.leido else 'card-leido') %}
  <div class="card mb-3 shadow-sm {{ card_class }}">
    <div class="card-body">
      <div class="d-flex justify-content-between align-items-start mb-1">
        <div>
          <strong>{{ m.nombre }} {{ m.apellido or '' }}</strong>
          {% if not m.leido %}<span class="badge bg-primary ms-2">NUEVO</span>{% endif %}
          {% if m.contestado %}<span class="badge bg-success ms-2">Contestado</span>{% endif %}
        </div>
        <small class="text-muted text-nowrap ms-2">{{ m.created_at|local_dt }}</small>
      </div>
      <div class="text-muted small mb-2">
        {% if m.email %}<a href="mailto:{{ m.email }}">{{ m.email }}</a>&nbsp;{% endif %}
        {% if m.telefono %}{{ m.telefono }}&nbsp;{% endif %}
        {% if m.ciudad or m.provincia %}{{ m.ciudad or '' }}{% if m.ciudad and m.provincia %}, {% endif %}{{ m.provincia or '' }}{% endif %}
      </div>
      <p class="mb-3 border rounded p-2 bg-white">{{ m.mensaje }}</p>
      <div class="d-flex gap-2 flex-wrap">
        {% if m.email %}
        <a href="mailto:{{ m.email }}" class="btn btn-sm btn-outline-primary">Responder</a>
        {% endif %}
        <form method="POST" action="/admin/contactos/{{ m.id }}/contestado" style="display:inline">
          <button type="submit" class="btn btn-sm {{ 'btn-success' if m.contestado else 'btn-outline-success' }}">
            {% if m.contestado %}Contestado{% else %}Marcar contestado{% endif %}
          </button>
        </form>
      </div>
    </div>
  </div>
  {% endfor %}
</div></body></html>
""", msgs=msgs, user=g.user)

@bp.route('/contactos/<int:mid>/contestado', methods=['POST'])
@admin_required
def contacto_contestado(mid):
    db = get_db()
    row = db.execute("SELECT contestado FROM contactos WHERE id=?", (mid,)).fetchone()
    if row:
        nuevo = 0 if row['contestado'] else 1
        db.execute("UPDATE contactos SET contestado=? WHERE id=?", (nuevo, mid))
        db.commit()
    db.close()
    return redirect(url_for('admin.contactos'))

# SUGERENCIAS (05/07/2026)
@bp.route('/sugerencias')
@admin_required
def sugerencias():
    db = get_db()
    msgs = db.execute("""
        SELECT s.*, u.nombre as user_nombre, u.email as user_email
        FROM sugerencias s LEFT JOIN users u ON u.id = s.user_id
        ORDER BY s.created_at DESC
    """).fetchall()
    db.execute("UPDATE sugerencias SET leido=1 WHERE leido=0")
    db.commit()
    db.close()
    return render_template_string("""
<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Sugerencias - Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<style>
  .card-nuevo  { border-left: 4px solid #0d6efd; }
  .card-leido  { border-left: 4px solid #dee2e6; }
  .card-resp   { border-left: 4px solid #198754; }
</style>
</head><body class="bg-light">
<div class="container py-4" style="max-width:760px">
  <a href="/admin/" class="btn btn-outline-secondary btn-sm mb-3">Volver</a>
  <h4 class="fw-bold mb-1">Sugerencias de usuarios</h4>
  <p class="text-muted small mb-3">
    <span class="badge bg-primary">{{ msgs|length }} total</span>
    <span class="badge bg-success ms-1">{{ msgs|selectattr('respondida','equalto',1)|list|length }} respondidas</span>
    <span class="badge bg-warning text-dark ms-1">{{ msgs|selectattr('leido','equalto',0)|list|length }} nuevas</span>
  </p>
  {% if not msgs %}<p class="text-muted">No hay sugerencias aun.</p>{% endif %}
  {% for m in msgs %}
  {% set card_class = 'card-resp' if m.respondida else ('card-nuevo' if not m.leido else 'card-leido') %}
  <div class="card mb-3 shadow-sm {{ card_class }}">
    <div class="card-body">
      <div class="d-flex justify-content-between align-items-start mb-1">
        <div>
          <strong>{{ m.user_nombre or m.user_email or 'Usuario #' ~ m.user_id }}</strong>
          {% if not m.leido %}<span class="badge bg-primary ms-2">NUEVA</span>{% endif %}
          {% if m.respondida %}<span class="badge bg-success ms-2">Respondida</span>{% endif %}
        </div>
        <small class="text-muted text-nowrap ms-2">{{ m.created_at|local_dt }}</small>
      </div>
      <div class="text-muted small mb-2">
        {% if m.user_email %}<a href="mailto:{{ m.user_email }}">{{ m.user_email }}</a>{% endif %}
      </div>
      <p class="mb-3 border rounded p-2 bg-white">{{ m.mensaje }}</p>
      <div class="d-flex gap-2 flex-wrap">
        {% if m.user_email %}
        <a href="mailto:{{ m.user_email }}" class="btn btn-sm btn-outline-primary">Responder</a>
        {% endif %}
        <form method="POST" action="/admin/sugerencias/{{ m.id }}/respondida" style="display:inline">
          <button type="submit" class="btn btn-sm {{ 'btn-success' if m.respondida else 'btn-outline-success' }}">
            {% if m.respondida %}Respondida{% else %}Marcar respondida{% endif %}
          </button>
        </form>
      </div>
    </div>
  </div>
  {% endfor %}
</div></body></html>
""", msgs=msgs, user=g.user)

@bp.route('/sugerencias/<int:mid>/respondida', methods=['POST'])
@admin_required
def sugerencia_respondida(mid):
    db = get_db()
    row = db.execute("SELECT respondida FROM sugerencias WHERE id=?", (mid,)).fetchone()
    if row:
        nuevo = 0 if row['respondida'] else 1
        db.execute("UPDATE sugerencias SET respondida=? WHERE id=?", (nuevo, mid))
        db.commit()
    db.close()
    return redirect(url_for('admin.sugerencias'))


# WHATSAPP (bandeja de respuesta manual, 20/07/2026)
@bp.route('/whatsapp')
@admin_required
def whatsapp_inbox():
    """Bandeja para responder a mano las consultas que el bot de WhatsApp
    (routes/whatsapp_bot.py) no supo contestar solo. Pedido de Daniel
    20/07/2026.

    Importante — regla de Meta que no depende de nuestro código: se puede
    mandar texto libre por acá SOLO dentro de las 24hs desde que la persona
    escribió (columna `dentro_ventana` de abajo); pasadas esas 24hs, Meta
    rechaza el envío de texto libre y exige una plantilla (template)
    pre-aprobada — igual que ya pasa con el código de verificación en
    utils/verificacion.py::enviar_codigo_whatsapp.

    Actualizado 21/07/2026: esta bandeja SÍ sirve para la campaña de
    retención — antes solo servía para conversaciones que la otra persona
    arrancaba de cero, pero una vez que un usuario de retención RESPONDE al
    mensaje que le mandamos (plantilla aprobada por Cloud API), esa
    respuesta entra por el mismo webhook y cae acá igual que cualquier otro
    mensaje entrante, abriendo la ventana de 24hs para contestarle texto
    libre de verdad. Para dar contexto de quién escribe, se cruza el
    teléfono contra `users` (incluso si vino con formato distinto:
    telefono_normalizado se queda con los últimos 10 dígitos) y contra el
    último envío de `retencion_contactos` para ese usuario."""
    db = get_db()
    consultas = db.execute(
        """SELECT c.*, v.ultima_interaccion
           FROM whatsapp_consultas_sin_responder c
           LEFT JOIN whatsapp_conversaciones v ON v.telefono = c.telefono
           ORDER BY c.respondida ASC, c.created_at DESC"""
    ).fetchall()

    usuarios_por_tel = {}
    for u in db.execute("SELECT id, nombre, email, telefono FROM users WHERE telefono IS NOT NULL AND telefono != ''").fetchall():
        usuarios_por_tel[telefono_normalizado(u['telefono'])] = dict(u)

    ultimo_contacto_por_user = {}
    for r in db.execute(
        """SELECT rc.user_id, rc.segmento, rc.mensaje, rc.canal, rc.created_at
           FROM retencion_contactos rc
           ORDER BY rc.created_at DESC"""
    ).fetchall():
        ultimo_contacto_por_user.setdefault(r['user_id'], dict(r))
    db.close()

    ahora = datetime.utcnow()
    filas = []
    for c in consultas:
        dentro_ventana = None
        if c['ultima_interaccion']:
            try:
                ultima = datetime.fromisoformat(str(c['ultima_interaccion']))
                dentro_ventana = (ahora - ultima) <= timedelta(hours=24)
            except ValueError:
                dentro_ventana = None
        fila = dict(c)
        fila['dentro_ventana'] = dentro_ventana
        usuario = usuarios_por_tel.get(telefono_normalizado(c['telefono']))
        fila['usuario'] = usuario
        fila['retencion'] = ultimo_contacto_por_user.get(usuario['id']) if usuario else None
        filas.append(fila)

    return render_template_string("""
<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>WhatsApp - Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css">
<style>
  .card-pendiente { border-left: 4px solid #ffc107; }
  .card-respondida { border-left: 4px solid #198754; }
</style>
</head><body class="bg-light">
<div class="container py-4" style="max-width:760px">
  <a href="/admin/" class="btn btn-outline-secondary btn-sm mb-3">Volver</a>
  <h4 class="fw-bold mb-1">WhatsApp — consultas sin responder por el bot</h4>
  """ + _FLASH_BLOCK + """
  <p class="text-muted small mb-3">
    <span class="badge bg-warning text-dark">{{ consultas|selectattr('respondida','equalto',0)|list|length }} pendientes</span>
    <span class="badge bg-success ms-1">{{ consultas|selectattr('respondida','equalto',1)|list|length }} respondidas</span>
  </p>
  <div class="alert alert-info small">
    <i class="bi bi-info-circle"></i> Solo se puede responder texto libre si la persona escribió
    hace menos de 24hs (columna "ventana"). Pasado ese plazo, Meta exige una plantilla aprobada —
    no es algo que se pueda evitar desde acá.
  </div>
  {% if not consultas %}<p class="text-muted">No hay consultas todavía.</p>{% endif %}
  {% for c in consultas %}
  <div class="card mb-3 shadow-sm {{ 'card-respondida' if c.respondida else 'card-pendiente' }}">
    <div class="card-body">
      <div class="d-flex justify-content-between align-items-start mb-1">
        <div>
          <strong>{{ c.usuario.nombre if c.usuario and c.usuario.nombre else c.telefono }}</strong>
          {% if c.usuario %}<small class="text-muted">({{ c.telefono }})</small>{% endif %}
          {% if not c.respondida %}<span class="badge bg-warning text-dark ms-2">PENDIENTE</span>{% endif %}
          {% if c.respondida %}<span class="badge bg-success ms-2">Respondida</span>{% endif %}
          {% if c.dentro_ventana %}
          <span class="badge bg-success ms-1"><i class="bi bi-clock-history"></i> dentro de ventana (24hs)</span>
          {% elif c.dentro_ventana is not none %}
          <span class="badge bg-danger ms-1"><i class="bi bi-clock-history"></i> fuera de ventana</span>
          {% endif %}
        </div>
        <small class="text-muted text-nowrap ms-2">{{ c.created_at|local_dt }}</small>
      </div>
      {% if c.retencion %}
      <p class="mb-1 small text-primary">
        <i class="bi bi-reply"></i> Responde a un mensaje de retención — Segmento {{ c.retencion.segmento }},
        enviado por {{ c.retencion.canal }} el {{ c.retencion.created_at|local_dt('%d/%m/%Y') }}
      </p>
      {% endif %}
      {% if c.usuario %}
      <p class="mb-1 small"><a href="{{ url_for('admin.seguimiento_detalle', uid=c.usuario.id) }}">Ver perfil completo en Seguimiento →</a></p>
      {% endif %}
      <p class="mb-2 border rounded p-2 bg-white">{{ c.mensaje }}</p>
      {% if c.respondida %}
      <p class="mb-0 small text-muted"><strong>Tu respuesta:</strong> {{ c.respuesta_admin }}</p>
      {% else %}
      <form method="POST" action="{{ url_for('admin.whatsapp_responder', cid=c.id) }}">
        <div class="input-group">
          <textarea name="respuesta" class="form-control" rows="2" placeholder="Escribí la respuesta..." required></textarea>
          <button type="submit" class="btn btn-success">Enviar</button>
        </div>
      </form>
      {% endif %}
    </div>
  </div>
  {% endfor %}
</div></body></html>
""", consultas=filas, user=g.user)


@bp.route('/whatsapp/<int:cid>/responder', methods=['POST'])
@admin_required
def whatsapp_responder(cid):
    texto = (request.form.get('respuesta') or '').strip()
    if not texto:
        flash('Escribí un mensaje antes de enviar.', 'error')
        return redirect(url_for('admin.whatsapp_inbox'))

    db = get_db()
    consulta = db.execute("SELECT * FROM whatsapp_consultas_sin_responder WHERE id=?", (cid,)).fetchone()
    if not consulta:
        db.close()
        flash('Consulta no encontrada.', 'error')
        return redirect(url_for('admin.whatsapp_inbox'))

    from routes.whatsapp_bot import enviar_mensaje_whatsapp
    ok, detalle = enviar_mensaje_whatsapp(consulta['telefono'], texto)
    if ok:
        db.execute(
            "UPDATE whatsapp_consultas_sin_responder SET respondida=1, respuesta_admin=? WHERE id=?",
            (texto, cid)
        )
        db.commit()
        flash('Respuesta enviada.', 'success')
    else:
        flash(f'No se pudo enviar: {detalle}. Si ya pasaron las 24hs desde que esa persona '
              'escribió, Meta exige una plantilla aprobada para mandarle texto libre (no es un '
              'error nuestro).', 'error')
    db.close()
    return redirect(url_for('admin.whatsapp_inbox'))


@bp.route('/redes')
@admin_required
def social_inbox():
    """Agregado 25/07/2026, fase 1 del CRM unificado (WhatsApp + Email +
    Messenger + Instagram en un solo lugar, pedido de Daniel). Bandeja para
    Facebook Messenger e Instagram DM -- mismo patrón que whatsapp_inbox,
    pero SIN cruce contra `users` todavía: Messenger/Instagram solo dan un
    ID de plataforma (PSID/IGSID), no hay teléfono ni email para matchear.
    Si Daniel confirma que quiere vincular estos contactos a un usuario real
    de la app, se agrega en una fase 2 (pidiéndole el dato en el chat)."""
    db = get_db()
    consultas = db.execute(
        "SELECT * FROM redes_consultas_sin_responder ORDER BY respondida ASC, created_at DESC"
    ).fetchall()
    db.close()

    return render_template_string("""
<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Redes sociales - Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css">
<style>
  .card-pendiente { border-left: 4px solid #ffc107; }
  .card-respondida { border-left: 4px solid #198754; }
</style>
</head><body class="bg-light">
<div class="container py-4" style="max-width:760px">
  <a href="/admin/" class="btn btn-outline-secondary btn-sm mb-3">Volver</a>
  <h4 class="fw-bold mb-1">Messenger / Instagram — consultas sin responder</h4>
  """ + _FLASH_BLOCK + """
  <p class="text-muted small mb-3">
    <span class="badge bg-warning text-dark">{{ consultas|selectattr('respondida','equalto',0)|list|length }} pendientes</span>
    <span class="badge bg-success ms-1">{{ consultas|selectattr('respondida','equalto',1)|list|length }} respondidas</span>
  </p>
  <div class="alert alert-info small">
    <i class="bi bi-info-circle"></i> Estos contactos NO están vinculados a ningún usuario de la app
    todavía — Messenger e Instagram no comparten teléfono ni email, solo un ID de la plataforma.
    Si la persona cuenta quién es en el chat, buscala a mano en Admin &gt; Usuarios.
  </div>
  {% if not consultas %}<p class="text-muted">No hay consultas todavía.</p>{% endif %}
  {% for c in consultas %}
  <div class="card mb-3 shadow-sm {{ 'card-respondida' if c.respondida else 'card-pendiente' }}">
    <div class="card-body">
      <div class="d-flex justify-content-between align-items-start mb-1">
        <div>
          <span class="badge {{ 'bg-primary' if c.canal == 'messenger' else 'bg-danger' }}">
            <i class="bi {{ 'bi-messenger' if c.canal == 'messenger' else 'bi-instagram' }}"></i>
            {{ 'Messenger' if c.canal == 'messenger' else 'Instagram' }}
          </span>
          <small class="text-muted ms-1">{{ c.remitente_id }}</small>
          {% if not c.respondida %}<span class="badge bg-warning text-dark ms-2">PENDIENTE</span>{% endif %}
          {% if c.respondida %}<span class="badge bg-success ms-2">Respondida</span>{% endif %}
        </div>
        <small class="text-muted text-nowrap ms-2">{{ c.created_at|local_dt }}</small>
      </div>
      <p class="mb-2 border rounded p-2 bg-white">{{ c.mensaje }}</p>
      {% if c.respondida %}
      <p class="mb-0 small text-muted"><strong>Tu respuesta:</strong> {{ c.respuesta_admin }}</p>
      {% else %}
      <form method="POST" action="{{ url_for('admin.social_responder', cid=c.id) }}">
        <div class="input-group">
          <textarea name="respuesta" class="form-control" rows="2" placeholder="Escribí la respuesta..." required></textarea>
          <button type="submit" class="btn btn-success">Enviar</button>
        </div>
      </form>
      {% endif %}
    </div>
  </div>
  {% endfor %}
</div></body></html>
""", consultas=consultas, user=g.user)


@bp.route('/redes/<int:cid>/responder', methods=['POST'])
@admin_required
def social_responder(cid):
    texto = (request.form.get('respuesta') or '').strip()
    if not texto:
        flash('Escribí un mensaje antes de enviar.', 'error')
        return redirect(url_for('admin.social_inbox'))

    db = get_db()
    consulta = db.execute("SELECT * FROM redes_consultas_sin_responder WHERE id=?", (cid,)).fetchone()
    if not consulta:
        db.close()
        flash('Consulta no encontrada.', 'error')
        return redirect(url_for('admin.social_inbox'))

    from routes.social_bot import enviar_mensaje_social
    ok, detalle = enviar_mensaje_social(consulta['remitente_id'], texto)
    if ok:
        db.execute(
            "UPDATE redes_consultas_sin_responder SET respondida=1, respuesta_admin=? WHERE id=?",
            (texto, cid)
        )
        db.commit()
        flash('Respuesta enviada.', 'success')
    else:
        flash(f'No se pudo enviar: {detalle}. Si ya pasaron las 24hs desde que esa persona '
              'escribió, Meta exige un mensaje "fuera de ventana" especial (no es un error '
              'nuestro).', 'error')
    db.close()
    return redirect(url_for('admin.social_inbox'))


# EMAIL (bandeja de respuesta manual, 25/07/2026 -- paso 1 del CRM unificado)
@bp.route('/email')
@admin_required
def email_inbox():
    """Bandeja para el mail entrante a contacto@presupuestopro.com.ar --
    mismo patrón que whatsapp_inbox, pero cruzando por email en vez de
    teléfono (ver routes/email_bot.py). El mail sigue llegando también a
    Gmail (el Worker de Cloudflare lo reenvía además de mandarlo acá), así
    que esta pantalla es un complemento, no un reemplazo."""
    db = get_db()
    consultas = db.execute(
        "SELECT * FROM email_consultas_entrantes ORDER BY respondida ASC, created_at DESC"
    ).fetchall()

    usuarios_por_email = {}
    for u in db.execute("SELECT id, nombre, email FROM users WHERE email IS NOT NULL AND email != ''").fetchall():
        usuarios_por_email[u['email'].strip().lower()] = dict(u)

    ultimo_contacto_por_user = {}
    for r in db.execute(
        """SELECT rc.user_id, rc.segmento, rc.mensaje, rc.canal, rc.created_at
           FROM retencion_contactos rc
           ORDER BY rc.created_at DESC"""
    ).fetchall():
        ultimo_contacto_por_user.setdefault(r['user_id'], dict(r))
    db.close()

    filas = []
    for c in consultas:
        fila = dict(c)
        usuario = usuarios_por_email.get(c['email_remitente'])
        fila['usuario'] = usuario
        fila['retencion'] = ultimo_contacto_por_user.get(usuario['id']) if usuario else None
        filas.append(fila)

    return render_template_string("""
<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Email - Admin</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.0/font/bootstrap-icons.css">
<style>
  .card-pendiente { border-left: 4px solid #ffc107; }
  .card-respondida { border-left: 4px solid #198754; }
</style>
</head><body class="bg-light">
<div class="container py-4" style="max-width:760px">
  <a href="/admin/" class="btn btn-outline-secondary btn-sm mb-3">Volver</a>
  <h4 class="fw-bold mb-1">Email — mail entrante a contacto@presupuestopro.com.ar</h4>
  """ + _FLASH_BLOCK + """
  <p class="text-muted small mb-3">
    <span class="badge bg-warning text-dark">{{ consultas|selectattr('respondida','equalto',0)|list|length }} pendientes</span>
    <span class="badge bg-success ms-1">{{ consultas|selectattr('respondida','equalto',1)|list|length }} respondidas</span>
  </p>
  <div class="alert alert-info small">
    <i class="bi bi-info-circle"></i> Este mail también sigue llegando a Gmail como antes --
    acá es solo para verlo cruzado con el usuario y responder sin salir de la app.
  </div>
  {% if not consultas %}<p class="text-muted">No hay mails todavía.</p>{% endif %}
  {% for c in consultas %}
  <div class="card mb-3 shadow-sm {{ 'card-respondida' if c.respondida else 'card-pendiente' }}">
    <div class="card-body">
      <div class="d-flex justify-content-between align-items-start mb-1">
        <div>
          <strong>{{ c.usuario.nombre if c.usuario and c.usuario.nombre else c.email_remitente }}</strong>
          {% if c.usuario %}<small class="text-muted">({{ c.email_remitente }})</small>{% endif %}
          {% if not c.respondida %}<span class="badge bg-warning text-dark ms-2">PENDIENTE</span>{% endif %}
          {% if c.respondida %}<span class="badge bg-success ms-2">Respondida</span>{% endif %}
        </div>
        <small class="text-muted text-nowrap ms-2">{{ c.created_at|local_dt }}</small>
      </div>
      {% if c.retencion %}
      <p class="mb-1 small text-primary">
        <i class="bi bi-reply"></i> Responde a un mensaje de retención — Segmento {{ c.retencion.segmento }},
        enviado por {{ c.retencion.canal }} el {{ c.retencion.created_at|local_dt('%d/%m/%Y') }}
      </p>
      {% endif %}
      {% if c.usuario %}
      <p class="mb-1 small"><a href="{{ url_for('admin.seguimiento_detalle', uid=c.usuario.id) }}">Ver perfil completo en Seguimiento →</a></p>
      {% endif %}
      {% if c.asunto %}<p class="mb-1 small text-muted"><strong>Asunto:</strong> {{ c.asunto }}</p>{% endif %}
      <p class="mb-2 border rounded p-2 bg-white" style="white-space:pre-wrap">{{ c.mensaje }}</p>
      {% if c.respondida %}
      <p class="mb-0 small text-muted"><strong>Tu respuesta:</strong> {{ c.respuesta_admin }}</p>
      {% else %}
      <form method="POST" action="{{ url_for('admin.email_responder', cid=c.id) }}">
        <div class="input-group">
          <textarea name="respuesta" class="form-control" rows="2" placeholder="Escribí la respuesta..." required></textarea>
          <button type="submit" class="btn btn-success">Enviar</button>
        </div>
      </form>
      {% endif %}
    </div>
  </div>
  {% endfor %}
</div></body></html>
""", consultas=filas, user=g.user)


@bp.route('/email/<int:cid>/responder', methods=['POST'])
@admin_required
def email_responder(cid):
    texto = (request.form.get('respuesta') or '').strip()
    if not texto:
        flash('Escribí un mensaje antes de enviar.', 'error')
        return redirect(url_for('admin.email_inbox'))

    db = get_db()
    consulta = db.execute("SELECT * FROM email_consultas_entrantes WHERE id=?", (cid,)).fetchone()
    if not consulta:
        db.close()
        flash('Consulta no encontrada.', 'error')
        return redirect(url_for('admin.email_inbox'))

    from routes.email_bot import enviar_respuesta_email
    asunto_resp = f"Re: {consulta['asunto']}" if consulta['asunto'] else "Re: PresupuestoPRO"
    ok, detalle = enviar_respuesta_email(consulta['email_remitente'], texto, asunto_resp)
    if ok:
        db.execute(
            "UPDATE email_consultas_entrantes SET respondida=1, respuesta_admin=? WHERE id=?",
            (texto, cid)
        )
        db.commit()
        flash('Respuesta enviada.', 'success')
    else:
        flash(f'No se pudo enviar: {detalle}.', 'error')
    db.close()
    return redirect(url_for('admin.email_inbox'))


# PRECIOS MATERIALES
# Fix 06/08/2026 (encontrado mientras Daniel probaba el aumento del 6.5%,
# reportó que varios materiales le aparecían en $0 después de aplicarlo):
# 'Piedra Partida (Calc. ó Granít.)', 'Granza (mediana)' y 'Perlitas Telgopor
# (75 Lts)' NO son el sub_nombre real que usa analisis_sub HOY -- no aparecen
# con ese texto exacto en la migración más reciente que tocó estos 3
# materiales (2q, LISTA_MATERIALES_V3_formulafix.xlsx). Confirmado levantando
# la base real (init_db+migrate_db) y mirando qué nombre quedó efectivamente
# cargado: 'Piedra partida' (p minúscula, $93.000), 'Granza' (sin paréntesis,
# $36.000), 'Perlitas Telgopor' (sin el "(75 Lts)", $106,67 -- el factor de
# bolsa de 75Lt para el precio comercial sigue andando igual, `_info_comercial`
# matchea por la palabra "perlitas", no por el nombre completo). Resultado:
# esta pantalla mostraba $0 para los 3 SIEMPRE (no por el 6.5% aplicado hoy)
# porque el lookup por nombre nunca encontraba la fila real -- y peor, si
# Daniel hubiera escrito un precio y guardado, el UPDATE ("WHERE
# sub_nombre=?") tampoco encontraba ninguna fila para actualizar, así que el
# precio real usado en los cálculos de presupuesto nunca cambiaba, aunque la
# pantalla pareciera aceptar el cambio sin error. Corregido acá a los
# nombres reales -- no se tocó ninguna fila de la base, es solo la etiqueta
# con la que esta pantalla busca/edita cada material.
#
# Al confirmar esos 3, se hizo una auditoría completa (script aparte, no
# queda en el repo) comparando CADA nombre de esta lista contra los
# sub_nombre reales de analisis_sub: de 90 materiales, 32 no matcheaban.
# 27 eran el mismo problema (tildes/ñ faltantes: "Cano"→"Caño",
# "latex"→"látex", "ceramico"→"cerámico", "neumatico"→"neumático", etc.) y
# se corrigieron igual que los 3 de arriba. 2 más ("Llaves de Paso Agua" y
# "Llaves de Paso Gas") en realidad comparten la MISMA fila real
# ("Llaves de Paso", una sola, sin distinguir agua/gas) — se corrigieron los
# 2 rótulos a ese nombre real (ya era así en los datos, esto no cambia el
# comportamiento, solo hace que la pantalla deje de mostrar $0 en los dos).
# 'Rev Text.' -- Daniel encontró (vía Costo/m2 de "Revest. Texturado") que
# SÍ tiene precio real, $4.400, así que no era un material sin cargar como
# se había concluido antes. Causa real: la migración 2n ("358 materiales
# resincronizados contra PRESUPUESTO COCHERA.xlsx", database.py línea 1874)
# renombró este material de 'Rev Text.' a 'DeckAr' (nombre comercial real
# del producto) DENTRO de la receta del ítem -- ese resync pasó por encima
# del nombre viejo que _LISTA_PRECIOS seguía usando. El precio SÍ se vio
# afectado por el aumento del 6.5% de hoy (4166.67 → 4400 redondeado a $100,
# coincide exacto con lo que Daniel vio en Costo/m2) porque
# `precios_aumento()` trabaja contra TODO analisis_sub, no contra esta
# lista -- lo único que faltaba corregir era el rótulo para que esta
# pantalla también lo encuentre y lo pueda editar.
#
# Reversión 06/08/2026: Daniel decidió que 'DeckAr' NO se mantiene como
# nombre -- es un nombre comercial de la zona de Rosario, no se reconoce en
# el resto del país (la app apunta a expansión regional). Se vuelve a
# 'Rev Text.' en TODAS partes: acá (rótulo de pantalla) y en la base real,
# vía la migración 3h en database.py (UPDATE analisis_sub SET
# sub_nombre='Rev Text.' WHERE sub_nombre='DeckAr'). Costo/m2 y Presupuestos
# no necesitan ningún cambio de código -- ambos leen sub_nombre en vivo
# desde analisis_sub, así que siguen el nuevo nombre automáticamente en
# cuanto corre la migración.
_LISTA_PRECIOS = [
    ('CORRALÓN - Áridos y Cemento', [
        'Cemento portland bolsas', 'Cemento Albañilería', 'Cal Hidráulica',
        'Cal aérea Milagro', 'Hidrófugo', 'Arena común', 'Tierra Colorada',
        'Piedra partida', 'Granza', 'Hormigon elaborado colado',
        'Perlitas Telgopor',
    ]),
    ('CORRALÓN - Ladrillos y Mampostería', [
        'Ladrillos comunes', 'Ladrillos vista',
        'Ladrillo hueco 8x18x33cm', 'Ladrillo hueco 12X18X33cm',
        'Ladrillo hueco 18X18X33cm', 'Ladrillo hueco Portante 12x18x33cm',
        'Ladrillo hueco Portante 18x18x33cm',
    ]),
    ('CORRALÓN - Hierros y Ferretería', [
        'Hierro redondo d=10mm', 'Alambre negro',
        'Clavos 2"', 'Clavos 2" 1/2', 'Clavos 3"', 'Clavos 4"',
    ]),
    ('CORRALÓN - Viguetas', [
        'Viga Vipret 4m.', 'Ladrillo telgopor 12*38*1m',
    ]),
    ('Maderera', [
        'Palito 1"x1"', 'Metal desplegado', 'Saligna   1"x2"', 'Saligna 1"x4"',
        'Saligna 3"x3"', 'Pino encofrado 1"', 'Tirantes 2x6', 'Pino tabla machimbre',
        'Escurridores 1/2 x 2', 'Issolant', 'Clavadores 2 x 2', 'Chapas Techo',
        'Tornillo c/arand goma', 'Chapas Cerco', 'Zócalo de madera', 'Tarugo 6',
        'Tornillo',
    ]),
    ('Instalaciones - Electricas', [
        'Caño Corrugado 1"', 'Caño Corrugado 3/4"', 'Cajas Metalicas',
        'Cable 2,5 mm', 'Cable 1,5 mm',
    ]),
    ('Instalaciones - Sanitarias', [
        'Caño Awaduct 110', 'Caño Awaduct 63', 'Caño Awaduct 50',
        'Caño Awaduct 40', 'Accesorios Desagues',
    ]),
    ('Instalaciones - Agua F/C', [
        'Caño TF 25', 'Caño TF 20', 'Accesorios TF', 'Llaves de Paso',
    ]),
    ('Instalaciones - Gas', [
        'Caño Epoxi 3/4', 'Caño epoxi 1/2', 'Accesorios Gas', 'Llaves de Paso',
    ]),
    ('Revestimientos y Pisos', [
        'Klaukol', 'Pastina',
        'Rvto.cerámico 1', 'Rvto.cerámico 2', 'Rvto.cerámico 3 (porcellanato)',
        'Piso cerámico 1', 'Piso cerámico 2', 'Piso cerámico 3 (porcellanato)',
        'Mosaico calcáreo', 'Loseta cemento 60x40cm', 'Baldosa cerámica azotea',
        'Zócalo cerámico 1', 'Zócalo cerámico 2', 'Zócalo cerámico 3 (Porcellanato)',
    ]),
    ('Pinturas y Terminaciones', [
        'Pintura látex exterior', 'Pintura látex interior', 'Pintura látex cielos',
        'Esmalte albalux', 'Pintura especial 1', 'Pintura especial 2',
        'Pintura satinol', 'Color pintura cal', 'Enduido sintético',
    ]),
    ('Materiales Especiales', [
        'Super Iggam', 'Salpicrete', 'Rev Text.', 'Fondo Base',
    ]),
    ('Servicios y Varios', [
        'Transporte material suelto', 'Martillo neumático',
    ]),
]

@bp.route('/precios')
@admin_required
def precios():
    db = get_db()
    rows = db.execute(
        "SELECT sub_nombre, MAX(precio_ars) as precio_ars "
        "FROM analisis_sub WHERE es_material=1 GROUP BY sub_nombre"
    ).fetchall()
    cfg_jo = db.execute("SELECT valor FROM config WHERE clave='jornal_oficial_dia'").fetchone()
    cfg_ja = db.execute("SELECT valor FROM config WHERE clave='jornal_ayudante_dia'").fetchone()
    db.close()
    precios_dict = {r['sub_nombre']: r['precio_ars'] for r in rows}

    jornal_oficial_dia  = int(float(cfg_jo['valor'])) if cfg_jo else 80000
    jornal_ayudante_dia = int(float(cfg_ja['valor'])) if cfg_ja else 40000

    # ⚠️ Cemento portland, Cemento Albañilería, Cal aérea, Klaukol, Salpicrete y
    # Super Iggam YA están en analisis_sub como precio por bolsa (migraciones 2j/2k/2l
    # en database.py). No van en este diccionario: si se los vuelve a multiplicar acá
    # por el factor de bolsa queda una DOBLE CONVERSIÓN y el precio comercial sale
    # 25-30 veces más caro de lo real (mismo bug que había en paso6_materiales).
    COMERCIAL = {
        'cal hidr':     (25,   'bolsa 25kg'),   # sin migrar — sigue en $/kg
        'cal viv':      (25,   'bolsa 25kg'),   # sin migrar — sigue en $/kg
        'perlitas':     (75,   'bolsa 75Lt'),   # sin migrar — sigue en $/lt
        'revear':       (30,   'balde 30kg'),   # sin migrar — sigue en $/kg
        'hierro':       (7.44, 'barra 12m'),    # sin migrar — sigue en $/kg
        'pastina':      (5,    'bolsa 5kg'),    # sin migrar — sigue en $/kg
    }

    def _info_comercial(nombre):
        n = nombre.lower()
        for kw, (factor, unidad) in COMERCIAL.items():
            if kw in n:
                return factor, unidad
        return 1, ''

    sectores = []
    for sector, nombres in _LISTA_PRECIOS:
        items = []
        for n in nombres:
            precio_calc = precios_dict.get(n, 0)
            factor, unidad_com = _info_comercial(n)
            precio_com = round(precio_calc * factor) if factor != 1 else precio_calc
            items.append({
                'nombre':      n,
                'precio':      precio_calc,
                'precio_com':  precio_com,
                'factor':      factor,
                'unidad_com':  unidad_com,
            })
        sectores.append({'sector': sector, 'items': items})

    return render_template('admin/precios.html', sectores=sectores, user=g.user,
                           jornal_oficial_dia=jornal_oficial_dia,
                           jornal_ayudante_dia=jornal_ayudante_dia)

@bp.route('/precios/aumento', methods=['POST'])
@admin_required
def precios_aumento():
    """Aplica un aumento porcentual a TODA la lista de precios de materiales
    (analisis_sub.precio_ars, es_material=1) de una sola vez, en vez de tener
    que tocar ítem por ítem en Admin > Precios.

    Agregado 06/08/2026, pedido de Daniel (aumento del 6.5% con redondeo, sin
    decimales, a toda la lista). Antes esa pantalla solo permitía editar cada
    material a mano; para un ajuste general periódico (algo que en Argentina
    se repite seguido) hacía falta escribir un script/migración cada vez.
    Redondeo "hacia arriba desde .5" (el redondeo tradicional en español, no
    el bankers' rounding de Python round(), que redondea 0.5 al par más
    cercano y puede confundir en algo con plata de por medio).

    Fix 06/08/2026 (mismo día, 2do pedido): el redondeo era siempre al peso
    -- Daniel pidió redondear a los $100. Se deja elegible (`redondeo`: 1, 10
    o 100, default 100) por si el próximo aumento necesita otra escala, en
    vez de hardcodear un solo valor y tener que volver a tocar código."""
    try:
        pct = float((request.form.get('pct') or '0').replace(',', '.'))
    except (TypeError, ValueError):
        flash('Porcentaje inválido.', 'error')
        return redirect(url_for('admin.precios'))
    if pct == 0:
        flash('Ingresá un porcentaje distinto de 0.', 'error')
        return redirect(url_for('admin.precios'))

    try:
        redondeo = int(request.form.get('redondeo') or 100)
    except (TypeError, ValueError):
        redondeo = 100
    if redondeo not in (1, 10, 100):
        redondeo = 100

    db = get_db()
    filas = db.execute(
        "SELECT sub_nombre, MAX(precio_ars) as precio_ars "
        "FROM analisis_sub WHERE es_material=1 GROUP BY sub_nombre"
    ).fetchall()
    actualizados = 0
    for f in filas:
        bruto = f['precio_ars'] * (1 + pct / 100)
        nuevo = math.floor(bruto / redondeo + 0.5) * redondeo
        # Piso de seguridad: un material barato (ej. $40) con redondeo a
        # $100 puede caer a $0 (40*1.065=42.6, redondea para abajo) -- eso
        # lo dejaría gratis en los presupuestos, un bug peor que el
        # redondeo en sí. Si el precio original era > 0, nunca baja de
        # `redondeo`.
        if nuevo == 0 and bruto > 0:
            nuevo = redondeo
        db.execute(
            "UPDATE analisis_sub SET precio_ars=? WHERE sub_nombre=? AND es_material=1",
            (nuevo, f['sub_nombre'])
        )
        actualizados += 1
    db.commit()
    db.close()
    flash(f'Aumento del {pct:g}% aplicado a {actualizados} materiales (redondeado a ${redondeo}).', 'success')
    return redirect(url_for('admin.precios'))


@bp.route('/precios/actualizar', methods=['POST'])
@admin_required
def precios_actualizar():
    db = get_db()
    actualizados = 0
    jornal_cambio = False
    for key, val in request.form.items():
        if key.startswith('calc_'):
            sub_nombre = key[5:]
            try:
                precio_ars = float(val)
                if precio_ars >= 0:
                    db.execute(
                        "UPDATE analisis_sub SET precio_ars=? WHERE sub_nombre=?",
                        (precio_ars, sub_nombre)
                    )
                    actualizados += 1
            except:
                pass
        elif key in ('jornal_oficial_dia', 'jornal_ayudante_dia'):
            try:
                valor = float(val)
                if valor > 0:
                    db.execute(
                        "INSERT OR REPLACE INTO config (clave, valor) VALUES (?, ?)",
                        (key, str(int(valor)))
                    )
                    actualizados += 1
                    jornal_cambio = True
            except:
                pass
    # Fix 07/07/2026: si se tocó el jornal, recalcular precio_mo_ars de TODOS
    # los ítems con los jornales nuevos — antes esta pantalla guardaba el
    # jornal en `config` pero nada volvía a leerlo, así que ningún ítem
    # cambiaba de costo de MO. Ver database.py::recalcular_precio_mo_ars.
    if jornal_cambio:
        n_mo = recalcular_precio_mo_ars(db)
        actualizados += n_mo
    db.commit()
    db.close()
    flash(f'Precios actualizados ({actualizados} ítems).', 'success')
    return redirect(url_for('admin.precios'))

# TIPOS DE CAMBIO
@bp.route('/tipos-cambio', methods=['GET', 'POST'])
@admin_required
def tipos_cambio():
    db = get_db()
    if request.method == 'POST':
        for key, val in request.form.items():
            if key.startswith('tasa_'):
                pais = key.replace('tasa_', '')
                try:
                    db.execute(
                        "UPDATE tipos_cambio SET tasa=?, updated_at=datetime('now', 'localtime') WHERE pais=?",
                        (float(val), pais)
                    )
                except:
                    pass
        db.commit()
        flash('Tipos de cambio actualizados.', 'success')
        db.close()
        return redirect(url_for('admin.tipos_cambio'))
    tcs = db.execute("SELECT * FROM tipos_cambio ORDER BY pais").fetchall()
    db.close()
    return render_template('admin/tipos_cambio.html', tcs=tcs, user=g.user)


@bp.route('/tipos-cambio/fetch-web')
@admin_required
def tipos_cambio_fetch():
    db = get_db()
    errores = []
    actualizados = []

    try:
        req = urllib.request.Request(
            'https://dolarapi.com/v1/dolares/oficial',
            headers={'User-Agent': 'PresupuestoPRO/1.0'}
        )
        with urllib.request.urlopen(req, timeout=5) as r:
            data = json.loads(r.read())
        tasa_ar = round(float(data.get('venta', 0)), 2)
        if tasa_ar > 0:
            db.execute("UPDATE tipos_cambio SET tasa=?, updated_at=datetime('now', 'localtime') WHERE pais='AR'",
                       (tasa_ar,))
            actualizados.append(f"ARS: {tasa_ar:.2f}")
    except Exception as e:
        errores.append(f"ARS: {e}")

    try:
        req2 = urllib.request.Request(
            'https://open.er-api.com/v6/latest/USD',
            headers={'User-Agent': 'PresupuestoPRO/1.0'}
        )
        with urllib.request.urlopen(req2, timeout=5) as r:
            rates = json.loads(r.read()).get('rates', {})
        mapa = {'CL': 'CLP', 'UY': 'UYU', 'BR': 'BRL', 'PY': 'PYG'}
        for pais, moneda in mapa.items():
            tasa = rates.get(moneda)
            if tasa:
                tasa_r = round(float(tasa), 2)
                db.execute(
                    "UPDATE tipos_cambio SET tasa=?, updated_at=datetime('now', 'localtime') WHERE pais=?",
                    (tasa_r, pais)
                )
                actualizados.append(f"{pais}: {tasa_r}")
    except Exception as e:
        errores.append(f"LATAM: {e}")

    db.commit()
    db.close()

    if actualizados:
        flash(f"Cotizaciones actualizadas: {', '.join(actualizados)}", 'success')
    if errores:
        flash(f"Errores: {'; '.join(errores)}", 'error')

    return redirect(url_for('admin.tipos_cambio'))


# RENDIMIENTOS
@bp.route('/rendimientos')
@admin_required
def rendimientos():
    db = get_db()
    items = db.execute("SELECT * FROM items_obra ORDER BY rubro_num, id").fetchall()
    db.close()
    from utils.calculations import RUBROS_DEFAULT
    return render_template('admin/rendimientos.html', items=items,
                           rubros=RUBROS_DEFAULT, user=g.user)

@bp.route('/rendimientos/actualizar', methods=['POST'])
@admin_required
def rendimientos_actualizar():
    db = get_db()
    for key, val in request.form.items():
        if key.startswith('hof_') or key.startswith('hay_'):
            tipo, iid = key.split('_', 1)
            try:
                if tipo == 'hof':
                    db.execute("UPDATE items_obra SET hof=? WHERE id=?", (float(val), int(iid)))
                else:
                    db.execute("UPDATE items_obra SET hay=? WHERE id=?", (float(val), int(iid)))
            except Exception:
                pass
    # Fix 07/07/2026: al cambiar HOF/HAY hay que recalcular precio_mo_ars con
    # esos valores nuevos (usando el jornal vigente en `config`) — antes
    # quedaba desactualizado hasta la próxima migración manual (ver bug
    # documentado en PROYECTO.md, sesión 04/07, y database.py::recalcular_precio_mo_ars).
    recalcular_precio_mo_ars(db)
    db.commit()
    db.close()
    flash('Rendimientos actualizados correctamente.', 'success')
    return redirect(url_for('admin.rendimientos'))


# FIX DB
@bp.route('/fix-db')
def fix_db():
    db = get_db()
    log = []
    items_borrar = [
        'Ayuda gremios y varios',
        'Rvto. marmol',
        'Ho.Ado. tanque (90-13)',
        'H.Elab. tanque (90-13)',
        'Cemento: revoque tanque',
    ]
    for nombre in items_borrar:
        r = db.execute("DELETE FROM items_obra WHERE nombre=?", (nombre,))
        if r.rowcount:
            log.append(f"DEL items_obra: {nombre}")

    r = db.execute("UPDATE items_obra SET precio_mo_ars=5000 WHERE id IN (97,98,99,100,101)")
    log.append(f"UPD pintura MO x{r.rowcount}")

    ceramicos = {82:11550, 83:11550, 84:21800, 87:4050, 88:4050, 92:14650, 93:14650}
    for iid, mo in ceramicos.items():
        db.execute("UPDATE items_obra SET precio_mo_ars=? WHERE id=?", (mo, iid))
    log.append(f"UPD ceramicos MO x{len(ceramicos)}")

    db.commit()
    cnt = db.execute("SELECT COUNT(*) FROM items_obra").fetchone()[0]
    db.close()
    log.append(f"VERIFY items_obra:{cnt}")
    from flask import jsonify
    return jsonify({'ok': True, 'cambios': log})


# CONFIGURACION
@bp.route('/configuracion', methods=['GET', 'POST'])
@admin_required
def configuracion():
    db = get_db()
    if request.method == 'POST':
        for clave in ('pct_gg', 'pct_impuestos'):
            val = request.form.get(clave)
            if val:
                db.execute(
                    "INSERT OR REPLACE INTO config (clave, valor) VALUES (?,?)",
                    (clave, val)
                )
        # Fix 10/07/2026: validación de cuenta (email/WhatsApp) — checkbox no
        # tildado no manda el campo en el form, por eso se chequea presencia
        # en vez de leer el valor.
        verificacion_val = '1' if request.form.get('verificacion_activa') == 'on' else '0'
        verificacion_previa = db.execute(
            "SELECT valor FROM config WHERE clave='verificacion_activa'"
        ).fetchone()
        se_esta_prendiendo = verificacion_val == '1' and (not verificacion_previa or verificacion_previa['valor'] != '1')
        db.execute(
            "INSERT OR REPLACE INTO config (clave, valor) VALUES ('verificacion_activa', ?)",
            (verificacion_val,)
        )
        # Fix 10/07/2026: al PRENDER el switch (no antes), se marca como
        # validado a cualquiera que ya tuviera cuenta y todavía no hubiera
        # validado — el bloqueo arranca a regir recién de acá para adelante,
        # nunca retroactivo a alguien que se registró antes de que el switch
        # estuviera prendido (Daniel lo pidió explícitamente 10/07/2026).
        if se_esta_prendiendo:
            n_grandfather = db.execute(
                "UPDATE users SET email_verificado=1, phone_verificado=1 "
                "WHERE metodo_verificacion != '' AND (email_verificado=0 OR phone_verificado=0)"
            ).rowcount
            if n_grandfather:
                print(f"[admin.configuracion] verificacion_activa prendida: "
                      f"{n_grandfather} cuenta(s) existente(s) marcadas como ya validadas (no retroactivo)")

        # Fix 18/07/2026: switch aparte para mostrar "Por WhatsApp" en el
        # registro, desacoplado de si las variables de entorno están
        # cargadas (ver utils/verificacion.py::whatsapp_configurado).
        whatsapp_val = '1' if request.form.get('whatsapp_validacion_habilitada') == 'on' else '0'
        db.execute(
            "INSERT OR REPLACE INTO config (clave, valor) VALUES ('whatsapp_validacion_habilitada', ?)",
            (whatsapp_val,)
        )
        db.commit()
        db.close()
        if se_esta_prendiendo and n_grandfather:
            flash(f'Configuracion guardada. Validación activada — {n_grandfather} cuenta(s) que ya '
                  f'existían quedaron marcadas como validadas (no se les exige validar retroactivamente).',
                  'success')
        else:
            flash('Configuracion guardada.', 'success')
        return redirect(url_for('admin.dashboard'))
    cfg = {r['clave']: r['valor'] for r in db.execute("SELECT * FROM config").fetchall()}
    pendientes_validar = db.execute(
        "SELECT COUNT(*) c FROM users WHERE metodo_verificacion != '' "
        "AND ((metodo_verificacion='email' AND email_verificado=0) "
        "OR (metodo_verificacion='whatsapp' AND phone_verificado=0))"
    ).fetchone()['c']
    db.close()
    return render_template('admin/configuracion.html', cfg=cfg, user=g.user,
                           pendientes_validar=pendientes_validar)


# LEADS
@bp.route('/leads')
@admin_required
def leads():
    db = get_db()
    todos = db.execute("SELECT * FROM leads ORDER BY created_at DESC").fetchall()
    db.close()
    return render_template('admin/leads.html', leads=todos, user=g.user)

@bp.route('/leads/<int:lid>/estado', methods=['POST'])
@admin_required
def lead_estado(lid):
    estado = request.form.get('estado', 'nuevo')
    notas  = request.form.get('notas', '')
    db = get_db()
    db.execute("UPDATE leads SET estado=?, notas=? WHERE id=?", (estado, notas, lid))
    db.commit()
    db.close()
    flash('Lead actualizado.', 'success')
    return redirect(url_for('admin.leads'))
