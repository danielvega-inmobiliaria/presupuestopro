import sqlite3
import os
from werkzeug.security import generate_password_hash
from config import Config

def get_db():
    os.makedirs(os.path.dirname(Config.DATABASE), exist_ok=True)
    db = sqlite3.connect(Config.DATABASE, timeout=20)
    db.row_factory = sqlite3.Row
    # Fix 05/07/2026: WAL en vez de DELETE — permite lecturas concurrentes
    # mientras hay una escritura en curso (antes cada escritura bloqueaba el
    # archivo entero por un instante). Pensado de cara a un pico de tráfico
    # el día del lanzamiento.
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys = ON")
    return db


def _actualizar_mo_analisis(db):
    """Calcula y guarda items_obra.precio_mo_ars (costo MO por unidad) desde la
    hoja Análisis de PRESUPUESTO HOTMART.xlsx.
    Para ítems no presentes en el Análisis usa fallback: hof×$10000 + hay×$5000.
    Solo corre si Replanteo.precio_mo_ars == 0 (primer arranque o reset)."""
    import unicodedata, re
    xlsx_path = os.path.join(os.path.dirname(Config.DATABASE), 'PRESUPUESTO HOTMART.xlsx')
    xlsx_exists = os.path.exists(xlsx_path)

    # Verificar si ya fue actualizado (todos los items con hof/hay tienen precio_mo_ars)
    sin_mo = db.execute(
        "SELECT COUNT(*) FROM items_obra WHERE (hof > 0 OR hay > 0) AND (precio_mo_ars IS NULL OR precio_mo_ars = 0)"
    ).fetchone()[0]
    if sin_mo == 0:
        return  # Todos los ítems ya tienen precio_mo_ars

    if not xlsx_exists:
        # Sin Excel: usar fallback hof×$10000 + hay×$5000 para items sin valor
        db.execute("""
            UPDATE items_obra
            SET precio_mo_ars = ROUND(hof * 10000 + hay * 5000, 2)
            WHERE (precio_mo_ars IS NULL OR precio_mo_ars = 0) AND (hof > 0 OR hay > 0)
        """)
        db.commit()
        updated = db.execute("SELECT changes()").fetchone()[0]
        print(f"[migrate_db] precio_mo_ars fallback hof/hay: {updated} ítems actualizados")
        return

    try:
        import openpyxl
    except ImportError:
        print("[migrate_db] openpyxl no disponible — precio_mo_ars no actualizado")
        return

    def norm(name):
        nfkd = unicodedata.normalize('NFKD', str(name))
        s = ''.join(c for c in nfkd if not unicodedata.combining(c))
        return re.sub(r'\.\s+', '.', ' '.join(s.lower().split()).rstrip('.'))

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Análisis']

    # ── Fase 1: parsear items y sub-ítems ─────────────────────────────────────
    items_raw = {}   # norm(nombre) → {'subs': [(sub_nombre, cant, precio, es_mo)]}
    pending   = []

    for row in ws.iter_rows(min_row=4, max_row=700, values_only=True):
        nombre  = str(row[0]).strip() if row[0] else None
        if not nombre or nombre == 'None':
            continue
        unidad, cant, precio, importe = row[1], row[2], row[3], row[4]
        # ITEM HEADER: tiene unidad, pero cant=None y precio=None
        if unidad is not None and cant is None and precio is None and importe is not None:
            items_raw[norm(nombre)] = {'subs': list(pending)}
            pending = []
        elif cant is not None and precio is not None:
            n_lower = nombre.lower()
            es_mo = n_lower.startswith('oficial') or n_lower.startswith('ayudante')
            pending.append((nombre, float(cant), float(precio), es_mo))

    # ── Alias para resolver sub-ítems referenciados por nombre corto ──────────
    SUB_ALIAS = {
        'hormigon colado':    'hormigon comun colado (6,00hh)',
        'hormigon elaborado': 'hormigon elab.(4,00hh)',
        'armadura colocada':  'armadura colocada (0,20hh)',
        'encofrado':          'encofrado (2,60hh)',
    }

    # ── Fase 2: resolver MO per unit recursivamente ───────────────────────────
    memo = {}

    def get_mo(key, depth=0):
        if key in memo:
            return memo[key]
        if key not in items_raw or depth > 3:
            return 0.0
        mo = 0.0
        for (sub_nombre, cant, precio, es_mo) in items_raw[key]['subs']:
            if es_mo:
                mo += cant * precio
            else:
                sub_key = SUB_ALIAS.get(norm(sub_nombre), norm(sub_nombre))
                if sub_key in items_raw:
                    mo += cant * get_mo(sub_key, depth + 1)
        memo[key] = mo
        return mo

    for key in items_raw:
        get_mo(key)

    # ── Mapeo items_obra → clave en Análisis (donde los nombres difieren) ─────
    ANALISIS_MAP = {
        # Mampostería
        'mamp.ladrillo comun 15cm':           'mamp.ladrillo comun 0,15m',
        'mamp.ladrillo comun 30cm':           'mamp.ladrillo comun 0,30m',
        'mamp.ladrillo vista 15cm':           'mamp.ladrillo vista 0,15m',
        'mamp.ladrillo vista 30cm':           'mamp.ladrillo vista 0,30m',
        # Tabiques (en Análisis figuran como 'mamp.ladrillo ...')
        'tabique ladrillo comun 8cm':         'mamp.ladrillo comun 0,075m',
        'tabique ladrillo hueco 8x18x33':     'mamp.ladrillo hueco 8cm',
        'tabique ladrillo hueco 12x18x33':    'mamp.ladrillo hueco 12cm',
        'tabique ladrillo hueco 18x18x33':    'mamp.ladrillo hueco 18cm',
        'tabique hueco portante 12x19x33cm':  'mamp.ladrillo hueco portante 12x18x33cm',
        'tabique hueco portante 18x19x33cm':  'mamp.ladrillo hueco portante 18x18x33cm',
        # Dinteles — diferencia tipográfica "30 cm" vs "30cm"
        'dinteles pared 30cm':               'dinteles pared 30 cm',
        # Cemento → nombre real en Análisis
        'cemento: piso alisado':              'piso cto.alisado',
        'cemento: piso rodillado':            'piso cemento rodillado',
        'cemento: revoque':                   'revoque cemento',
        'cemento: toma de juntas':            'toma de juntas',
        'cemento: zocalo 10cm':               'zocalo cemento 10cm',
        'cemento: azotado':                   'azotado impermeable',
        'cemento: capa aisladora s/muros':    'capa aisladora hor.y vertical',
        # Instalaciones
        'instalacion desague':                'instalacion desagues',
        # Hormigón
        'hormigon comun':                     'hormigon comun colado (6,00hh)',
        'hormigon elaborado colado':          'hormigon elab.(4,00hh)',
        'armadura colocada':                  'armadura colocada (0,20hh)',
        'armadura colocada (h.elab.)':        'armadura colocada (0,10hh)',
        'encofrado':                          'encofrado (2,60hh)',
        'encofrado (h.elab.)':                'encofrado (2,60hh)',
        # Ho.Ado. compuestos
        'ho.ado.(horm+arm+enc)':              'ho.ao.segun cant.de horm.arm y enc',
        'ho.ado.viga fundacion (75-7)':       'ho.ado.viga fund (75-7)',
        'ho.ado.losa 15cm (75-6.66)':         'ho.ado.losa 15cm (75-7)',
        'ho.ado.tabique 10cm (110-20)':       'ho.ado.tab.10cm (110-20)',
        'ho.ado.tabique 15cm (110-13)':       'ho.ado.tab.15cm (110-13,33)',
        'ho.ado.tanque (90-13)':              'ho.ado.tanque (90-13,33)',
        # H.Elab. compuestos
        'h.elab.(horm+arm+enc)':              'ho.ao.segun cant.de horm.arm y enc',
        'h.elab.viga fundacion (75-7)':       'h.elab.viga fund (75-7)',
        'h.elab.losa 10cm (80-10)':           'h.elab.losa 10cm.(80-10)',
        'h.elab.losa 15cm (75-6.66)':         'h.elab.losa 15cm.(75-7)',
        'h.elab.losa premoldeada':            'ho.ado.losa premoldeada',
        'h.elab.tabique 10cm (110-20)':       'h.elab.tab.10cm.(110-14)',
        'h.elab.tabique 15cm (110-13)':       'h.elab.tab.15cm.(110-7)',
        'h.elab.tanque (90-13)':              'h.elab.tanque (90-14)',
        # Otros
        'carpeta bajo piso':                  'carpeta b/piso',
        'banquina ho pobre 10cm':             'banquina h pobre 10cm',
        'revest.texturado':                   'revest texturado',
        'demolicion hormigon armado':         'demolicion hormigon ado',
        'rebajo de cordon':                   'rebaje de cordon',
        'excavacion a maquina':               'excavacion maquina',
        'dinteles pared 30cm':                'dinteles pared 30 cm',
        'estructura cielo suspendido (3-6m)': 'estructura cielo suspendido (3 a 6m)',
        # NOTA: 'ayuda gremios y varios' matchea directo tras norm() → no necesita entrada
        # NOTA: 'tabique ladrillo comun 8cm', 'tabique bloque 12/18' → usa fallback hof/hay
    }

    PRICE_OF = 10000   # Oficial: $10.000/h (de hoja Mat y MO)
    PRICE_AY = 5000    # Ayudante: $5.000/h

    # ── Actualizar precio_mo_ars en items_obra ────────────────────────────────
    items = db.execute("SELECT id, nombre, hof, hay FROM items_obra").fetchall()
    updated = 0
    for item in items:
        item_id, nombre, hof, hay = item[0], item[1], item[2] or 0.0, item[3] or 0.0
        key    = norm(nombre)
        mapped = ANALISIS_MAP.get(key, key)

        if mapped in memo:
            mo = memo[mapped]
        else:
            # Fallback: hof/hay (por unidad) × tasas fijas de Mat y MO
            mo = hof * PRICE_OF + hay * PRICE_AY

        db.execute(
            "UPDATE items_obra SET precio_mo_ars=? WHERE id=?",
            (round(mo, 2), item_id)
        )
        updated += 1

    db.commit()
    print(f"[migrate_db] precio_mo_ars actualizado: {updated} ítems desde Análisis")


def _actualizar_rendimientos_excel(db):
    """Actualiza items_obra.hof y items_obra.hay desde PRESUPUESTO HOTMART.xlsx.
    Solo corre si el archivo existe. Se detecta si los valores están desactualizados
    comparando Techo de chapas (debe ser hof=3.5 según el Excel actual)."""
    import unicodedata, re
    xlsx_path = os.path.join(os.path.dirname(Config.DATABASE), 'PRESUPUESTO HOTMART.xlsx')
    if not os.path.exists(xlsx_path):
        return

    # Detectar si ya está actualizado (Techo de chapas hof debe ser 3.5)
    r = db.execute("SELECT hof FROM items_obra WHERE nombre='Techo de chapas'").fetchone()
    if r and abs(r[0] - 3.5) < 0.01:
        return  # Ya actualizado

    try:
        import openpyxl
    except ImportError:
        print("[migrate_db] openpyxl no disponible — rendimientos no actualizados")
        return

    def norm(name):
        nfkd = unicodedata.normalize('NFKD', str(name))
        s = ''.join(c for c in nfkd if not unicodedata.combining(c))
        return re.sub(r'\.\s+', '.', ' '.join(s.lower().split()).rstrip('.'))

    NAME_MAP = {
        'hormigon comun':                    'hormigon comun (6,00hh)',
        'hormigon elaborado colado':         'ho.ado.(s/cant.de horm,arm y enc)',
        'armadura colocada':                 'armadura colocada (0,20hh)',
        'armadura colocada (h.elab.)':       'armadura colocada (0,20hh)',
        'encofrado':                         'encofrado (2,60hh)',
        'encofrado (h.elab.)':               'encofrado (2,60hh)',
        'ho.ado. (horm+arm+enc)':           'ho.ado.(s/cant.de horm,arm y enc)',
        'h.elab. (horm+arm+enc)':           'ho.ado.(s/cant.de horm,arm y enc)',
        'ho.ado.viga fundacion (75-7)':     'ho.ado.viga fund (75-7)',
        'ho.ado.losa 15cm (75-6.66)':       'ho.ado.losa 15cm (75-6,66)',
        'ho.ado.tabique 10cm (110-20)':     'ho.ado.tab.10cm (110-20)',
        'ho.ado.tabique 15cm (110-13)':     'ho.ado.tab.15cm (110-13,33)',
        'ho.ado.tanque (90-13)':            'ho.ado.tanque (90-13,33)',
        'h.elab.viga fundacion (75-7)':     'h.elab.viga fund (75-7)',
        'h.elab.losa 10cm (80-10)':         'h.elab.losa 10cm.(80-10)',
        'h.elab.losa 15cm (75-6.66)':       'h.elab.losa 15cm.(75-6,66)',
        'h.elab.tabique 10cm (110-20)':     'h.elab.tab.10cm.(110-20)',
        'h.elab.tabique 15cm (110-13)':     'h.elab.tab.15cm.(110-13,33)',
        'h.elab.tanque (90-13)':            'h.elab.tanque (90-13,33)',
        'tabique ladrillo comun 8cm':       'tabique ladrillo comun 8cm (08)',
        'tabique ladrillo hueco 8x18x33':   'tabique ladrillo hueco 8x18x33 (10)',
        'tabique ladrillo hueco 12x18x33':  'tabique ladrillo hueco 12x18x33 (15)',
        'tabique ladrillo hueco 18x18x33':  'tabique ladrillo hueco 18x18x33 (21)',
        'tabique bloque 12x19x33cm':        'tabique bloque 12x19x33cm (15)',
        'tabique bloque 18x19x33cm':        'tabique bloque 18x19x33cm (21)',
        'instalacion desague':              'instalacion desagues',
        'carpeta bajo piso':                'carpeta b/piso',
        'banquina ho pobre 10cm':           'banquina h pobre 10cm',
        'revest.texturado':                 'revest texturado',
    }

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['Items']

    def to_f(v):
        try: return float(v) if v and v != 'no' else 0.0
        except: return 0.0

    ref_hof = [to_f(ws['I23'].value), to_f(ws['I24'].value), to_f(ws['I25'].value)]
    ref_hay = [to_f(ws['J23'].value), to_f(ws['J24'].value), to_f(ws['J25'].value)]

    rend = {}
    for row in ws.iter_rows(min_row=5, max_row=150, values_only=True):
        nombre = row[0]
        if not nombre or nombre == 'ítems': continue
        n = str(nombre).strip()
        i_hof, j_hay = to_f(row[8]), to_f(row[9])
        k, l, m = to_f(row[10]), to_f(row[11]), to_f(row[12])
        if i_hof > 0 or j_hay > 0:
            rend[norm(n)] = (i_hof, j_hay)
        elif k > 0 or l > 0 or m > 0:
            rend[norm(n)] = (
                ref_hof[0]*k + ref_hof[1]*l + ref_hof[2]*m,
                ref_hay[0]*k + ref_hay[1]*l + ref_hay[2]*m,
            )

    items_rows = db.execute("SELECT id, nombre FROM items_obra").fetchall()
    updated = 0
    for row in items_rows:
        item_id, nombre = row[0], row[1]
        key = norm(nombre)
        mapped = NAME_MAP.get(key, key)
        if mapped in rend:
            hof, hay = rend[mapped]
            db.execute("UPDATE items_obra SET hof=?, hay=? WHERE id=?",
                       (round(hof, 4), round(hay, 4), item_id))
            updated += 1

    db.commit()
    print(f"[migrate_db] Rendimientos HOF/HAY actualizados desde Excel: {updated} ítems")


def migrate_db():
    """Agrega columnas nuevas y actualiza rendimientos HOF/HAY desde Excel."""
    db = get_db()
    try:
        # 1. Asegurar que la tabla presupuestos exista (por si acaso)
        cols_rows = db.execute("PRAGMA table_info(presupuestos)").fetchall()
        if not cols_rows:
            db.close()
            return  # init_db se encargara de crear la tabla
        cols = [r[1] for r in cols_rows]

        # Mapa completo de columnas que pueden faltar en DBs viejas
        columnas_nuevas = {
            'modo':               "ALTER TABLE presupuestos ADD COLUMN modo TEXT DEFAULT 'mo_mat'",
            'superficie':         "ALTER TABLE presupuestos ADD COLUMN superficie REAL DEFAULT 0",
            'rubros_json':        "ALTER TABLE presupuestos ADD COLUMN rubros_json TEXT DEFAULT '[]'",
            'subcontratos_json':  "ALTER TABLE presupuestos ADD COLUMN subcontratos_json TEXT DEFAULT '[]'",
            'indirectos_json':    "ALTER TABLE presupuestos ADD COLUMN indirectos_json TEXT DEFAULT '{}'",
            'materiales_json':    "ALTER TABLE presupuestos ADD COLUMN materiales_json TEXT DEFAULT '[]'",
            'hh_total':           "ALTER TABLE presupuestos ADD COLUMN hh_total REAL DEFAULT 0",
            'n_oficiales':        "ALTER TABLE presupuestos ADD COLUMN n_oficiales INTEGER DEFAULT 2",
            'n_ayudantes':        "ALTER TABLE presupuestos ADD COLUMN n_ayudantes INTEGER DEFAULT 1",
            'jornal_oficial':     "ALTER TABLE presupuestos ADD COLUMN jornal_oficial REAL DEFAULT 0",
            'jornal_ayudante':    "ALTER TABLE presupuestos ADD COLUMN jornal_ayudante REAL DEFAULT 0",
            'dias_obra':          "ALTER TABLE presupuestos ADD COLUMN dias_obra INTEGER DEFAULT 0",
            'pct_gg':             "ALTER TABLE presupuestos ADD COLUMN pct_gg REAL DEFAULT 20",
            'pct_impuestos':      "ALTER TABLE presupuestos ADD COLUMN pct_impuestos REAL DEFAULT 7",
            'pct_anticipo':       "ALTER TABLE presupuestos ADD COLUMN pct_anticipo REAL DEFAULT 30",
            'pct_final':          "ALTER TABLE presupuestos ADD COLUMN pct_final REAL DEFAULT 20",
            'frecuencia_pago':    "ALTER TABLE presupuestos ADD COLUMN frecuencia_pago TEXT DEFAULT 'mensual'",
            'tipo_cambio':        "ALTER TABLE presupuestos ADD COLUMN tipo_cambio REAL DEFAULT 1",
            'costo_directo':      "ALTER TABLE presupuestos ADD COLUMN costo_directo REAL DEFAULT 0",
            'total_subcontratos': "ALTER TABLE presupuestos ADD COLUMN total_subcontratos REAL DEFAULT 0",
            'total_indirectos':   "ALTER TABLE presupuestos ADD COLUMN total_indirectos REAL DEFAULT 0",
            'total_mo':           "ALTER TABLE presupuestos ADD COLUMN total_mo REAL DEFAULT 0",
            'total_materiales':   "ALTER TABLE presupuestos ADD COLUMN total_materiales REAL DEFAULT 0",
            'total_presupuesto':  "ALTER TABLE presupuestos ADD COLUMN total_presupuesto REAL DEFAULT 0",
            'descripcion_trabajos': "ALTER TABLE presupuestos ADD COLUMN descripcion_trabajos TEXT DEFAULT ''",
            'status':             "ALTER TABLE presupuestos ADD COLUMN status TEXT DEFAULT 'completo'",
            'wizard_step':        "ALTER TABLE presupuestos ADD COLUMN wizard_step INTEGER DEFAULT 8",
            'session_json':       "ALTER TABLE presupuestos ADD COLUMN session_json TEXT DEFAULT '{}'",
            'updated_at':         "ALTER TABLE presupuestos ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP",
        }
        nuevas = []
        for col, sql in columnas_nuevas.items():
            if col not in cols:
                nuevas.append(sql)
        for sql in nuevas:
            db.execute(sql)
        if nuevas:
            db.commit()
            print("[migrate_db] Columnas agregadas: {}".format(len(nuevas)))

        # 0. Crear tabla suscripciones si no existe (DBs anteriores a MP)
        db.execute("""
            CREATE TABLE IF NOT EXISTS suscripciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER REFERENCES users(id),
                mp_preapproval_id TEXT UNIQUE,
                plan_nombre TEXT DEFAULT 'mensual',
                monto_ars REAL DEFAULT 0,
                estado TEXT DEFAULT 'pending',
                fecha_inicio DATE,
                fecha_fin DATE,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Agregar columna mp_preapproval_id a users si no existe
        cols_users = [r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()]
        if 'mp_preapproval_id' not in cols_users:
            db.execute("ALTER TABLE users ADD COLUMN mp_preapproval_id TEXT DEFAULT ''")
            db.commit()
            print("[migrate_db] users.mp_preapproval_id agregado")

        # Crear tabla contactos si no existe
        db.execute("""
            CREATE TABLE IF NOT EXISTS contactos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT,
                apellido TEXT,
                telefono TEXT,
                email TEXT,
                ciudad TEXT,
                provincia TEXT,
                mensaje TEXT,
                leido INTEGER DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Agregar columnas extra a contactos si ya existía sin ellas
        cols_contactos = [r[1] for r in db.execute("PRAGMA table_info(contactos)").fetchall()]
        for col, tipo in [('apellido','TEXT'), ('telefono','TEXT'), ('ciudad','TEXT'), ('provincia','TEXT')]:
            if col not in cols_contactos:
                db.execute(f"ALTER TABLE contactos ADD COLUMN {col} {tipo} DEFAULT ''")
        if 'contestado' not in cols_contactos:
            db.execute("ALTER TABLE contactos ADD COLUMN contestado INTEGER DEFAULT 0")
        # Agregar columnas a users para datos de perfil
        cols_users2 = [r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()]
        for col, tipo in [('apellido','TEXT'), ('telefono','TEXT'), ('ciudad','TEXT'), ('provincia','TEXT')]:
            if col not in cols_users2:
                db.execute(f"ALTER TABLE users ADD COLUMN {col} {tipo} DEFAULT ''")
        db.commit()

        # Crear tabla sugerencias si no existe (05/07/2026 — feedback de usuarios logueados)
        db.execute("""
            CREATE TABLE IF NOT EXISTS sugerencias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER REFERENCES users(id),
                mensaje TEXT,
                leido INTEGER DEFAULT 0,
                respondida INTEGER DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        db.commit()

        # 1a-extra. Crear tabla empresa_perfil si no existe (para DBs antiguas)
        db.execute("""
            CREATE TABLE IF NOT EXISTS empresa_perfil (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER UNIQUE REFERENCES users(id),
                nombre TEXT DEFAULT '',
                contacto TEXT DEFAULT '',
                telefono TEXT DEFAULT '',
                email TEXT DEFAULT '',
                slogan TEXT DEFAULT '',
                logo_data TEXT DEFAULT '',
                logo_filename TEXT DEFAULT '',
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Agregar columna contacto si no existe (DBs anteriores)
        cols_ep = [r[1] for r in db.execute("PRAGMA table_info(empresa_perfil)").fetchall()]
        if 'contacto' not in cols_ep:
            db.execute("ALTER TABLE empresa_perfil ADD COLUMN contacto TEXT DEFAULT ''")
            db.commit()
            print("[migrate_db] empresa_perfil.contacto agregado")
        db.commit()

        # Fix 05/07/2026: tabla password_reset_tokens — routes/auth.py (recuperar()
        # y restablecer()) la usa desde siempre, pero nunca se creó acá ni en
        # init_db(). En producción esto rompía "Olvidé mi contraseña" con un
        # Error Interno del Servidor (INSERT/SELECT contra una tabla inexistente).
        # Reportado por Daniel al probar el login en stopro.com.ar.
        db.execute("""
            CREATE TABLE IF NOT EXISTS password_reset_tokens (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id),
                token TEXT UNIQUE NOT NULL,
                expires_at DATETIME NOT NULL,
                used INTEGER DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        db.commit()

        # 1b. Columnas nuevas en items_obra (precio_mo_ars, orden)
        cols_io = [r[1] for r in db.execute("PRAGMA table_info(items_obra)").fetchall()]
        if 'precio_mo_ars' not in cols_io:
            db.execute("ALTER TABLE items_obra ADD COLUMN precio_mo_ars REAL DEFAULT 0")
            db.commit()
            print("[migrate_db] Columna items_obra.precio_mo_ars agregada")
        if 'orden' not in cols_io:
            db.execute("ALTER TABLE items_obra ADD COLUMN orden INTEGER DEFAULT 0")
            db.commit()
            print("[migrate_db] Columna items_obra.orden agregada")

        # 1b-extra. Eliminar items retirados de la lista
        for _nombre_eliminar in ('Casilla rodante obrador', 'Cartel de obra',
                                 'Desmonte a maquina', 'Excavacion a maquina',
                                 'Suelo cemento 6%',
                                 'Ho.Ado. zapata (30-6)',
                                 'H.Elab. zapata (30-6)',
                                 'Hormigon comun',
                                 'Armadura colocada',
                                 'Encofrado',
                                 'Ho.Ado. (Horm+Arm+Enc)',
                                 'Hormigon elaborado colado',
                                 'Armadura colocada (H.Elab.)',
                                 'Encofrado (H.Elab.)',
                                 'H.Elab. (Horm+Arm+Enc)'):
            db.execute("DELETE FROM items_obra WHERE nombre=?", (_nombre_eliminar,))
        db.commit()

        # 1c. Rename Tabique bloque → Tabique hueco portante
        db.execute("""UPDATE items_obra SET nombre='Tabique hueco portante 12x19x33cm'
                      WHERE nombre='Tabique bloque 12x19x33cm'""")
        db.execute("""UPDATE items_obra SET nombre='Tabique hueco portante 18x19x33cm'
                      WHERE nombre='Tabique bloque 18x19x33cm'""")
        db.commit()

        # 1d. Reasignar rubro_num / rubro_nombre / orden según nuevo esquema de tareas
        NUEVA_ASIGNACION = {
            # nombre_item: (rubro_num, rubro_nombre, orden)
            # 01 - Preliminares
            'Cerco de obra':                       ('01','Preliminares', 30),
            'Replanteo':                           ('01','Preliminares', 40),
            'Picado de revoques':                  ('01','Preliminares', 50),
            'Rebajo de cordon':                    ('01','Preliminares', 60),
            'Demolicion solado 20cm':              ('01','Preliminares', 70),
            'Demolicion mamposteria':              ('01','Preliminares', 80),
            'Demolicion hormigon armado':          ('01','Preliminares', 90),
            # 02 - Fundaciones
            'Excavacion zanjas a mano':            ('02','Fundaciones', 10),
            'Excavacion bases a mano':             ('02','Fundaciones', 20),
            'Excavacion de Pilotes':               ('02','Fundaciones', 30),
            'Relleno y Compactacion':              ('02','Fundaciones', 40),
            'Relleno de arena':                    ('02','Fundaciones', 50),
            'Retiro de tierra':                    ('02','Fundaciones', 60),
            'Zapata Ho Pobre':                     ('02','Fundaciones', 80),
            'Mamp. ladrillo comun cimientos':      ('02','Fundaciones', 90),
            'Mamp. ladrillo comun submuracion':    ('02','Fundaciones',100),
            # 03 - Hormigón Armado (Ho.Ado.)
            'Ho.Ado. base (50-3)':                 ('03','Hormigón Armado (Ho.Ado.)',  10),
            'Ho.Ado. zapata (60-6)':               ('03','Hormigón Armado (Ho.Ado.)',  30),
            'Ho.Ado. viga fundacion (75-7)':       ('03','Hormigón Armado (Ho.Ado.)',  40),
            'Ho.Ado. columna (105-11)':            ('03','Hormigón Armado (Ho.Ado.)',  50),
            'Ho.Ado. viga (120-16)':               ('03','Hormigón Armado (Ho.Ado.)',  60),
            'Ho.Ado. tabique 10cm (110-20)':       ('03','Hormigón Armado (Ho.Ado.)',  70),
            'Ho.Ado. tabique 15cm (110-13)':       ('03','Hormigón Armado (Ho.Ado.)',  80),
            'Ho.Ado. tanque (90-13)':              ('03','Hormigón Armado (Ho.Ado.)',  90),
            'Ho.Ado. losa 10cm (80-10)':           ('03','Hormigón Armado (Ho.Ado.)', 100),
            'Ho.Ado. losa 15cm (75-6.66)':         ('03','Hormigón Armado (Ho.Ado.)', 110),
            'Ho.Ado. losa premoldeada':            ('03','Hormigón Armado (Ho.Ado.)', 120),
            'Ho.Ado. pavimento 15cm':              ('03','Hormigón Armado (Ho.Ado.)', 130),
            # 04 - Hormigón Elaborado (H.Elab.)
            'H.Elab. base (50-3)':                 ('04','Hormigón Elaborado (H.Elab.)',  10),
            'H.Elab. zapata (60-6)':               ('04','Hormigón Elaborado (H.Elab.)',  30),
            'H.Elab. viga fundacion (75-7)':       ('04','Hormigón Elaborado (H.Elab.)',  40),
            'H.Elab. columna (105-11)':            ('04','Hormigón Elaborado (H.Elab.)',  50),
            'H.Elab. viga (120-16)':               ('04','Hormigón Elaborado (H.Elab.)',  60),
            'H.Elab. tabique 10cm (110-20)':       ('04','Hormigón Elaborado (H.Elab.)',  70),
            'H.Elab. tabique 15cm (110-13)':       ('04','Hormigón Elaborado (H.Elab.)',  80),
            'H.Elab. tanque (90-13)':              ('04','Hormigón Elaborado (H.Elab.)',  90),
            'H.Elab. losa 10cm (80-10)':           ('04','Hormigón Elaborado (H.Elab.)', 100),
            'H.Elab. losa 15cm (75-6.66)':         ('04','Hormigón Elaborado (H.Elab.)', 110),
            'H.Elab. losa premoldeada':            ('04','Hormigón Elaborado (H.Elab.)', 120),
            'H.Elab. pavimento 15cm':              ('04','Hormigón Elaborado (H.Elab.)', 130),
            # 05 - Cemento
            'Cemento: capa aisladora s/muros':     ('05','Cemento',  10),
            'Carpeta bajo piso':                   ('05','Cemento',  20),
            'Cemento: carpeta azotea':             ('05','Cemento',  30),
            'Cemento: toma de juntas':             ('05','Cemento',  40),
            'Cemento: azotado':                    ('05','Cemento',  50),
            'Cemento: revoque':                    ('05','Cemento',  60),
            'Cemento: zocalo 10cm':                ('05','Cemento',  80),
            'Cemento: piso rodillado':             ('05','Cemento',  90),
            'Cemento: piso alisado':               ('05','Cemento', 100),
            # 06 - Mampostería
            'Mamp. ladrillo comun 15cm':           ('06','Mampostería',  10),
            'Mamp. ladrillo comun 30cm':           ('06','Mampostería',  20),
            'Mamp. ladrillo vista 15cm':           ('06','Mampostería',  30),
            'Mamp. ladrillo vista 30cm':           ('06','Mampostería',  40),
            'Tabique ladrillo comun 8cm':          ('06','Mampostería',  50),
            'Tabique ladrillo hueco 8x18x33':      ('06','Mampostería',  60),
            'Tabique ladrillo hueco 12x18x33':     ('06','Mampostería',  70),
            'Tabique ladrillo hueco 18x18x33':     ('06','Mampostería',  80),
            'Tabique hueco portante 12x19x33cm':   ('06','Mampostería',  90),
            'Tabique hueco portante 18x19x33cm':   ('06','Mampostería', 100),
            'Dinteles pared 15cm':                 ('06','Mampostería', 110),
            'Dinteles pared 30cm':                 ('06','Mampostería', 120),
            # 07 - Contrapisos
            'Contrapiso cascotes 15cm':            ('07','Contrapisos',  10),
            'Contrapiso cascotes 10cm':            ('07','Contrapisos',  20),
            'Banquina Ho Pobre 10cm':              ('07','Contrapisos',  30),
            'Contrapiso c/perlitas 10cm':          ('07','Contrapisos',  40),
            'Contrapiso c/perlitas 5cm':           ('07','Contrapisos',  50),
            'Ho de Pendiente c/perlitas 10cm':     ('07','Contrapisos',  60),
            # 08 - Revoques
            'R. exterior cal (az+gr+f)':           ('08','Revoques',  10),
            'R. interior cal (gr+f) fratas':       ('08','Revoques',  20),
            'R. interior cal (gr+f) fieltro':      ('08','Revoques',  30),
            'R. grueso interior b/rvto.':          ('08','Revoques',  40),
            'Cielorraso s/losa (gr+f)':            ('08','Revoques',  50),
            'Cielo s/metal despl. (az+gr+f)':      ('08','Revoques',  60),
            'Estructura cielo suspendido (3-6m)':  ('08','Revoques',  70),
            # 09 - Revestimientos
            'Enlucido S.Iggam':                    ('09','Revestimientos',  10),
            'Enlucido Salpicrete':                 ('09','Revestimientos',  20),
            'Revest. Texturado':                   ('09','Revestimientos',  30),
            # 10 - Techos
            'Techo de chapas':                     ('10','Techos',  10),
            # 11 - Instalaciones
            'Colocacion de Aberturas':             ('11','Instalaciones',  10),
            'Instalacion Desague':                 ('11','Instalaciones',  20),
            'Instalacion Agua F/C':                ('11','Instalaciones',  30),
            'Instalacion Gas':                     ('11','Instalaciones',  40),
            'Instalacion Electrica':               ('11','Instalaciones',  50),
            # 12 - Cerámicos y Porcellanatos
            'Piso ceramico azotea':                ('12','Cerámicos y Porcellanatos',  10),
            'Piso ceramico 1':                     ('12','Cerámicos y Porcellanatos',  20),
            'Piso ceramico 2':                     ('12','Cerámicos y Porcellanatos',  30),
            'Piso ceramico 3 (porcellanato)':      ('12','Cerámicos y Porcellanatos',  40),
            'Piso loseta cemento':                 ('12','Cerámicos y Porcellanatos',  50),
            'Piso calcareo vereda':                ('12','Cerámicos y Porcellanatos',  60),
            'Zocalo ceramico 1':                   ('12','Cerámicos y Porcellanatos',  70),
            'Zocalo ceramico 2':                   ('12','Cerámicos y Porcellanatos',  80),
            'Zocalo ceramico 3':                   ('12','Cerámicos y Porcellanatos',  90),
            'Zocalo de pino':                      ('12','Cerámicos y Porcellanatos', 100),
            'Rvto. ceramico 1':                    ('12','Cerámicos y Porcellanatos', 110),
            'Rvto. ceramico 2':                    ('12','Cerámicos y Porcellanatos', 120),
            'Rvto. ceramico 3 porcellanato':       ('12','Cerámicos y Porcellanatos', 130),
            'Rvto. marmol':                        ('12','Cerámicos y Porcellanatos', 140),
            # 13 - Pintura
            'Pintura cal':                         ('13','Pintura',  10),
            'Pintura latex exterior':              ('13','Pintura',  20),
            'Pintura latex interior':              ('13','Pintura',  30),
            'Pintura latex cielos':                ('13','Pintura',  40),
            'Pintura especial 1':                  ('13','Pintura',  50),
            'Pintura especial 2':                  ('13','Pintura',  60),
            'Enduido s/paredes':                   ('13','Pintura',  70),
            'Pintura satinol muros':               ('13','Pintura',  80),
            'Pintura s/carp. madera':              ('13','Pintura',  90),
            'Pintura s/carp. metalica':            ('13','Pintura', 100),
            'Ayuda gremios y varios':              ('13','Pintura', 110),
        }
        for nombre, (rn, rnombre, ord_) in NUEVA_ASIGNACION.items():
            db.execute(
                "UPDATE items_obra SET rubro_num=?, rubro_nombre=?, orden=? WHERE nombre=?",
                (rn, rnombre, ord_, nombre)
            )
        db.commit()
        print("[migrate_db] Rubros / orden actualizados según esquema de tareas")

        # 2. Actualizar HOF/HAY en items_obra si la mayoría son 0
        cero = db.execute("SELECT COUNT(*) as c FROM items_obra WHERE hof=0 AND hay=0").fetchone()['c']
        total = db.execute("SELECT COUNT(*) as c FROM items_obra").fetchone()['c']
        if total > 0 and cero / total > 0.7:
            # Rendimientos reales por nombre de item
            RENDIMIENTOS = {
                'Cerco de obra': (0.3, 0.1),
                'Replanteo': (0.05, 0.02),
                'Picado de revoques': (0.4, 0.2),
                'Rebajo de cordon': (0.5, 0.2),
                'Demolicion solado 20cm': (0.3, 0.15),
                'Demolicion mamposteria': (8.0, 4.0),
                'Demolicion hormigon armado': (0, 4.0),
                'Relleno y Compactacion': (1.5, 0.5),
                'Suelo cemento 6%': (2.0, 1.0),
                'Relleno de arena': (1.0, 0.5),
                'Retiro de tierra': (0.5, 0.5),
                'Excavacion zanjas a mano': (4.0, 0),
                'Excavacion bases a mano': (5.0, 0),
                'Excavacion de Pilotes': (3.0, 0),
                'Ho.Ado. pavimento 15cm': (1.5, 0.7),
                'Ho.Ado. base (50-3)': (25.0, 12.0),
                'Ho.Ado. zapata (60-6)': (30.0, 15.0),
                'Ho.Ado. columna (105-11)': (50.0, 25.0),
                'Ho.Ado. viga fundacion (75-7)': (40.0, 20.0),
                'Ho.Ado. viga (120-16)': (55.0, 27.0),
                'Ho.Ado. losa 10cm (80-10)': (45.0, 22.0),
                'Ho.Ado. losa 15cm (75-6.66)': (40.0, 20.0),
                'Ho.Ado. losa premoldeada': (1.5, 0.7),
                'Ho.Ado. tabique 10cm (110-20)': (55.0, 27.0),
                'Ho.Ado. tabique 15cm (110-13)': (50.0, 25.0),
                'Ho.Ado. tanque (90-13)': (45.0, 22.0),
                'Zapata Ho Pobre': (8.0, 4.0),
                'Mamp. ladrillo comun cimientos': (18.0, 9.0),
                'Mamp. ladrillo comun submuracion': (20.0, 10.0),
                'Mamp. ladrillo comun 15cm': (18.0, 9.0),
                'Mamp. ladrillo comun 30cm': (22.0, 11.0),
                'Mamp. ladrillo vista 15cm': (22.0, 11.0),
                'Mamp. ladrillo vista 30cm': (26.0, 13.0),
                'Tabique ladrillo comun 8cm': (1.0, 0.5),
                'Tabique ladrillo hueco 8x18x33': (1.0, 0.5),
                'Tabique ladrillo hueco 12x18x33': (1.2, 0.6),
                'Tabique ladrillo hueco 18x18x33': (1.4, 0.7),
                'Tabique hueco portante 12x19x33cm': (0.8, 0.4),
                'Tabique hueco portante 18x19x33cm': (0.9, 0.45),
                'Dinteles pared 15cm': (0.5, 0.25),
                'Dinteles pared 30cm': (0.7, 0.35),
                'Cemento: azotado': (0.15, 0.05),
                'Cemento: capa aisladora s/muros': (0.3, 0.1),
                'Cemento: carpeta azotea': (0.25, 0.1),
                'Contrapiso cascotes 15cm': (0.35, 0.15),
                'Contrapiso cascotes 10cm': (0.3, 0.12),
                'Banquina Ho Pobre 10cm': (0.3, 0.12),
                'Ho de Pendiente c/perlitas 10cm': (0.4, 0.15),
                'Contrapiso c/perlitas 10cm': (0.35, 0.15),
                'Contrapiso c/perlitas 5cm': (0.25, 0.1),
                'Piso ceramico azotea': (0.7, 0.35),
                'Techo de chapas': (0.5, 0.25),
                'Instalacion Desague': (40.0, 20.0),
                'Instalacion Agua F/C': (30.0, 15.0),
                'Instalacion Gas': (20.0, 10.0),
                'Instalacion Electrica': (4.0, 2.0),
                'Cemento: piso rodillado': (0.3, 0.1),
                'Cemento: piso alisado': (0.4, 0.15),
                'Cemento: toma de juntas': (0.15, 0.05),
                'Cemento: revoque': (0.45, 0.2),
                'Cemento: zocalo 10cm': (0.2, 0.05),
                'Enlucido S.Iggam': (0.25, 0.1),
                'Enlucido Salpicrete': (0.2, 0.08),
                'Revest. Texturado': (0.3, 0.1),
                'R. exterior cal (az+gr+f)': (0.5, 0.25),
                'R. interior cal (gr+f) fratas': (0.45, 0.2),
                'R. interior cal (gr+f) fieltro': (0.5, 0.22),
                'R. grueso interior b/rvto.': (0.25, 0.1),
                'Cielorraso s/losa (gr+f)': (0.6, 0.3),
                'Cielo s/metal despl. (az+gr+f)': (0.7, 0.35),
                'Carpeta bajo piso': (0.3, 0.1),
                'Piso ceramico 1': (0.7, 0.35),
                'Piso ceramico 2': (0.75, 0.35),
                'Piso ceramico 3 (porcellanato)': (0.9, 0.45),
                'Piso loseta cemento': (0.65, 0.3),
                'Piso calcareo vereda': (0.6, 0.3),
                'Zocalo ceramico 1': (0.3, 0.1),
                'Zocalo ceramico 2': (0.3, 0.1),
                'Zocalo ceramico 3': (0.35, 0.12),
                'Zocalo de pino': (0.2, 0.05),
                'Rvto. marmol': (1.2, 0.6),
                'Rvto. ceramico 1': (0.8, 0.4),
                'Rvto. ceramico 2': (0.85, 0.4),
                'Rvto. ceramico 3 porcellanato': (1.0, 0.5),
                'Colocacion de Aberturas': (3.0, 1.5),
                'Pintura cal': (0.12, 0.04),
                'Pintura latex exterior': (0.15, 0.05),
                'Pintura latex interior': (0.15, 0.05),
                'Pintura latex cielos': (0.18, 0.06),
                'Pintura especial 1': (0.2, 0.07),
                'Pintura especial 2': (0.22, 0.08),
                'Enduido s/paredes': (0.2, 0.05),
                'Pintura satinol muros': (0.2, 0.07),
                'Pintura s/carp. madera': (0.25, 0.05),
                'Pintura s/carp. metalica': (0.2, 0.05),
                'Estructura cielo suspendido (3-6m)': (1.5, 0.75),
                'Ayuda gremios y varios': (1.0, 0.5),
                'H.Elab. pavimento 15cm': (0.755, 0.755),
                'H.Elab. base (50-3)': (8.3, 8.3),
                'H.Elab. zapata (60-6)': (13.6, 13.6),
                'H.Elab. columna (105-11)': (23.85, 23.85),
                'H.Elab. viga fundacion (75-7)': (15.95, 15.95),
                'H.Elab. viga (120-16)': (32.6, 32.6),
                'H.Elab. losa 10cm (80-10)': (21.0, 21.0),
                'H.Elab. losa 15cm (75-6.66)': (15.406, 15.406),
                'H.Elab. losa premoldeada': (0.12, 0.12),
                'H.Elab. tabique 10cm (110-20)': (38.5, 38.5),
                'H.Elab. tabique 15cm (110-13)': (27.828, 27.828),
                'H.Elab. tanque (90-13)': (26.828, 26.828),
            }
            for nombre, (hof, hay) in RENDIMIENTOS.items():
                db.execute(
                    "UPDATE items_obra SET hof=?, hay=? WHERE nombre=? AND hof=0 AND hay=0",
                    (hof, hay, nombre)
                )
            db.commit()
            print(f"[migrate_db] Rendimientos HOF/HAY actualizados en items_obra")
        # 2b. Corregir es_material en analisis_sub para equipos/transporte
        servicios_costo = [
            'Transporte material suelto', 'Transporte casilla',
            'Martillo neumático', 'Máquina excavadora', 'Máquina topadora',
        ]
        for svc in servicios_costo:
            db.execute(
                "UPDATE analisis_sub SET es_material=1 WHERE sub_nombre=? AND es_material=0",
                (svc,)
            )
        db.commit()

        # 2c. Corregir HOF/HAY de Demolicion hormigon armado (solo Ayudante=4, HOF=0)
        db.execute(
            "UPDATE items_obra SET hof=0, hay=4 WHERE nombre='Demolicion hormigon armado'"
        )
        db.commit()

        # 2d. Actualizar analisis_sub: renombres y precios corregidos
        db.execute("UPDATE analisis_sub SET sub_nombre='Piedra partida' WHERE sub_nombre='Canto rodado'")
        db.execute("UPDATE analisis_sub SET precio_ars=2284.95 WHERE sub_nombre='Hierro redondo d=10mm'")
        db.execute("UPDATE analisis_sub SET sub_nombre='Junta (Pavimento Hormigón)' WHERE sub_nombre='Junta (encofrado equivalente)'")
        db.commit()

        # 2e. Corregir HOF/HAY de H.Elab. — valores derivados del Excel (HOF=HAY por simetría)
        # Fórmula Excel: días_item = HOF_total / (8 × n_of); HOF igual a HAY para H.Elab.
        helab_rendimientos = {
            'H.Elab. pavimento 15cm':         (0.755, 0.755),
            'H.Elab. base (50-3)':             (8.3,   8.3),
            'H.Elab. zapata (60-6)':           (13.6,  13.6),
            'H.Elab. columna (105-11)':        (23.85, 23.85),
            'H.Elab. viga fundacion (75-7)':   (15.95, 15.95),
            'H.Elab. viga (120-16)':           (32.6,  32.6),
            'H.Elab. losa 10cm (80-10)':       (21.0,  21.0),
            'H.Elab. losa 15cm (75-6.66)':     (15.406,15.406),
            'H.Elab. losa premoldeada':        (0.12,  0.12),
            'H.Elab. tabique 10cm (110-20)':   (38.5,  38.5),
            'H.Elab. tabique 15cm (110-13)':   (27.828,27.828),
            'H.Elab. tanque (90-13)':          (26.828,26.828),
        }
        for nombre, (hof, hay) in helab_rendimientos.items():
            db.execute("UPDATE items_obra SET hof=?, hay=? WHERE nombre=?", (hof, hay, nombre))
        db.commit()
        print("[migrate_db] 2e: HOF/HAY H.Elab. corregidos")

        # 2f. Correcciones losa ceramica y precio Hormigón elaborado colado ($145.000/m3)
        db.execute("""
            UPDATE analisis_sub SET precio_ars=145000
            WHERE sub_nombre IN ('Hormigón elaborado colado','Hormigon elaborado colado')
        """)
        # Losa ceramica: cambiar 'Hormigón colado' → 'Hormigón elaborado colado'
        db.execute("""
            UPDATE analisis_sub SET sub_nombre='Hormigón elaborado colado', precio_ars=145000
            WHERE item_nombre='H.Elab.losa ceramica' AND sub_nombre='Hormigón colado'
        """)
        db.commit()
        print("[migrate_db] 2f: Precios Hormigón elaborado colado y losa ceramica corregidos")
        print("[migrate_db] 2e: HOF/HAY H.Elab. corregidos")

        # 2g. Agregar columna m2_factor a items_obra (para calculadora Costo/m2)
        cols_io2 = [r[1] for r in db.execute("PRAGMA table_info(items_obra)").fetchall()]
        if 'm2_factor' not in cols_io2:
            db.execute("ALTER TABLE items_obra ADD COLUMN m2_factor REAL DEFAULT NULL")
            db.commit()
            print("[migrate_db] 2g: Columna items_obra.m2_factor agregada")

        # Renombres de items (para mejor legibilidad en la app)
        _renombres = [
            ('Relleno y Compactacion',              'Relleno y Compactacion C/15cm'),
            ('Piso ceramico 2',                     'Piso Ceramico'),
            ('Piso ceramico 3 (porcellanato)',       'Piso Porcellanato'),
            ('Zocalo ceramico 2',                   'Zocalo Ceramico'),
            ('Zocalo ceramico 3',                   'Zocalo Porcellanato'),
            ('Rvto. ceramico 2',                    'Rvto. Ceramico'),
            ('Rvto. ceramico 3 porcellanato',       'Rvto. Porcellanato'),
            ('Demolicion mamposteria',              'Demolicion mamposteria 0,15'),
        ]
        for viejo, nuevo in _renombres:
            db.execute("UPDATE items_obra SET nombre=? WHERE nombre=?", (nuevo, viejo))
        db.commit()

        # Agregar Demolicion mamposteria 0,30 si no existe
        existe_030 = db.execute(
            "SELECT id FROM items_obra WHERE nombre='Demolicion mamposteria 0,30'"
        ).fetchone()
        if not existe_030:
            src = db.execute(
                "SELECT rubro_num, rubro_nombre, unidad, precio_ars, hof, hay "
                "FROM items_obra WHERE nombre='Demolicion mamposteria 0,15'"
            ).fetchone()
            if src:
                db.execute(
                    "INSERT INTO items_obra (rubro_num, rubro_nombre, nombre, unidad, precio_ars, hof, hay, orden, m2_factor) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (src['rubro_num'], src['rubro_nombre'],
                     'Demolicion mamposteria 0,30', src['unidad'],
                     src['precio_ars'], src['hof'], src['hay'], 85, 0.30)
                )
                db.commit()
                print("[migrate_db] 2g: Item 'Demolicion mamposteria 0,30' agregado")

        # Setear m2_factor para los items que se convierten de m3 a m2
        _m2_factors = {
            'Demolicion mamposteria 0,15':     0.15,
            'Demolicion mamposteria 0,30':     0.30,
            'Relleno y Compactacion C/15cm':   0.15,
            'Ho.Ado. tabique 10cm (110-20)':   0.10,
            'Ho.Ado. tabique 15cm (110-13)':   0.15,
            'Ho.Ado. losa 10cm (80-10)':       0.10,
            'Ho.Ado. losa 15cm (75-6.66)':     0.15,
            'H.Elab. tabique 10cm (110-20)':   0.10,
            'H.Elab. tabique 15cm (110-13)':   0.15,
            'H.Elab. losa 10cm (80-10)':       0.10,
            'H.Elab. losa 15cm (75-6.66)':     0.15,
            'Mamp. ladrillo comun 15cm':       0.15,
            'Mamp. ladrillo comun 30cm':       0.30,
            'Mamp. ladrillo vista 15cm':       0.15,
            'Mamp. ladrillo vista 30cm':       0.30,
        }
        for nombre, factor in _m2_factors.items():
            db.execute("UPDATE items_obra SET m2_factor=? WHERE nombre=?", (factor, nombre))
        db.commit()
        print("[migrate_db] 2g: m2_factor seteado en items_obra")

        # 3. Crear tabla analisis_sub si no existe y popularla
        db.execute("""
            CREATE TABLE IF NOT EXISTS analisis_sub (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_nombre TEXT NOT NULL,
                sub_nombre TEXT NOT NULL,
                cant_por_unit REAL NOT NULL,
                precio_ars REAL DEFAULT 0,
                es_material INTEGER DEFAULT 1
            )
        """)
        db.execute("CREATE INDEX IF NOT EXISTS idx_analisis_item ON analisis_sub(item_nombre)")
        analisis_count = db.execute("SELECT COUNT(*) as c FROM analisis_sub").fetchone()['c']
        if analisis_count == 0:
            try:
                from utils.analisis_data import ANALISIS
                rows = []
                for item_nombre, subs in ANALISIS.items():
                    for s in subs:
                        rows.append((item_nombre, s['n'], s['c'], s['p'], s['m']))
                db.executemany(
                    "INSERT INTO analisis_sub (item_nombre, sub_nombre, cant_por_unit, precio_ars, es_material) VALUES (?,?,?,?,?)",
                    rows
                )
                db.commit()
                print(f"[migrate_db] analisis_sub: {len(rows)} filas insertadas")
            except Exception as e2:
                print(f"[migrate_db] analisis_sub error: {e2}")

        # 4. Actualizar rendimientos HOF/HAY desde PRESUPUESTO HOTMART.xlsx
        _actualizar_rendimientos_excel(db)

        # 5. Calcular precio_mo_ars (MO cost per unit desde Análisis sheet)
        _actualizar_mo_analisis(db)

        # ── 2h. Corrección masiva HOF/HAY + repoblado analisis_sub (datos reales Excel) ─
        ya_2h = db.execute("SELECT valor FROM config WHERE clave='2h_done'").fetchone()
        if not ya_2h:
            # HOF/HAY correctos del Excel (hoja Items, columnas I y J — valores del autor)
            # Para ítems compuestos (Ho.Ado./H.Elab.) se calculó con la fórmula del Excel:
            #   HOF = ref_hof_horm*hor + ref_hof_arm*arm + ref_hof_enc*enc
            # ref Ho.Ado: horm=[1.5,4.5], arm=[0.1,0.1], enc=[1.6,1.0]
            # ref H.Elab: horm=[1.0,3.0], arm=[0.05,0.05], enc=[1.6,1.0]
            HOF_HAY_2H = {
                # 01 Preliminares
                'Cerco de obra':                     (0.75,  0.0),
                'Replanteo':                         (0.75,  0.75),
                'Picado de revoques':                (0.0,   0.45),
                'Rebajo de cordon':                  (3.0,   0.0),
                'Demolicion solado 20cm':            (0.0,   0.8),
                'Demolicion mamposteria 0,15':       (0.0,   4.0),
                'Demolicion mamposteria 0,30':       (0.0,   4.0),
                'Demolicion hormigon armado':        (0.0,   4.0),
                # 02 Fundaciones
                'Excavacion zanjas a mano':          (0.0,   4.6),
                'Excavacion bases a mano':           (0.0,   6.0),
                'Excavacion de Pilotes':             (0.0,   4.5),
                'Relleno y Compactacion C/15cm':     (0.0,   5.0),
                'Relleno de arena':                  (0.0,   2.0),
                'Retiro de tierra':                  (0.0,   2.0),
                'Zapata Ho Pobre':                   (0.8,   4.0),
                'Mamp. ladrillo comun cimientos':    (8.0,   4.0),
                'Mamp. ladrillo comun submuracion':  (13.7,  6.85),
                # 03 Ho.Ado. — calculados con fórmula Excel
                'Ho.Ado. base (50-3)':               (11.3,  12.5),
                'Ho.Ado. zapata (60-6)':             (17.1,  16.5),
                'Ho.Ado. viga fundacion (75-7)':     (20.2,  19.0),
                'Ho.Ado. columna (105-11)':          (29.6,  26.0),
                'Ho.Ado. viga (120-16)':             (39.1,  32.5),
                'Ho.Ado. tabique 10cm (110-20)':     (44.5,  35.5),
                'Ho.Ado. tabique 15cm (110-13)':     (33.828,28.83),
                'Ho.Ado. tanque (90-13)':            (31.828,26.83),
                'Ho.Ado. losa 10cm (80-10)':         (25.5,  22.5),
                'Ho.Ado. losa 15cm (75-6.66)':       (19.656,18.66),
                'Ho.Ado. losa premoldeada':          (1.3,   2.25),
                'Ho.Ado. pavimento 15cm':            (0.955, 1.225),
                # 04 H.Elab. — calculados con fórmula Excel
                'H.Elab. base (50-3)':               (8.3,   8.5),
                'H.Elab. zapata (60-6)':             (13.6,  12.0),
                'H.Elab. viga fundacion (75-7)':     (15.95, 13.75),
                'H.Elab. columna (105-11)':          (23.85, 19.25),
                'H.Elab. viga (120-16)':             (32.6,  25.0),
                'H.Elab. tabique 10cm (110-20)':     (38.5,  28.5),
                'H.Elab. tabique 15cm (110-13)':     (27.828,21.83),
                'H.Elab. tanque (90-13)':            (26.828,20.83),
                'H.Elab. losa 10cm (80-10)':         (21.0,  17.0),
                'H.Elab. losa 15cm (75-6.66)':       (15.406,13.41),
                'H.Elab. losa premoldeada':          (0.75,  2.25),
                'H.Elab. pavimento 15cm':            (0.755, 0.875),
                # 05 Cemento
                'Cemento: capa aisladora s/muros':   (1.2,   0.6),
                'Carpeta bajo piso':                 (0.5,   0.25),
                'Cemento: carpeta azotea':           (0.6,   0.3),
                'Cemento: toma de juntas':           (1.2,   0.6),
                'Cemento: azotado':                  (0.3,   0.15),
                'Cemento: revoque':                  (1.6,   0.8),
                'Cemento: zocalo 10cm':              (0.5,   0.25),
                'Cemento: piso rodillado':           (0.76,  0.38),
                'Cemento: piso alisado':             (1.5,   0.75),
                # 06 Mampostería
                'Mamp. ladrillo comun 15cm':         (11.5,  5.75),
                'Mamp. ladrillo comun 30cm':         (9.4,   4.7),
                'Mamp. ladrillo vista 15cm':         (17.2,  8.6),
                'Mamp. ladrillo vista 30cm':         (14.0,  7.0),
                'Tabique ladrillo comun 8cm':        (1.2,   0.6),
                'Tabique ladrillo hueco 8x18x33':    (0.7,   0.35),
                'Tabique ladrillo hueco 12x18x33':   (0.8,   0.4),
                'Tabique ladrillo hueco 18x18x33':   (0.9,   0.45),
                'Tabique hueco portante 12x19x33cm': (0.8,   0.4),
                'Tabique hueco portante 18x19x33cm': (0.9,   0.45),
                'Dinteles pared 15cm':               (0.65,  0.33),
                'Dinteles pared 30cm':               (0.9,   0.45),
                # 07 Contrapisos
                'Contrapiso cascotes 15cm':          (0.8,   0.4),
                'Contrapiso cascotes 10cm':          (0.6,   0.3),
                'Banquina Ho Pobre 10cm':            (0.6,   0.3),
                'Contrapiso c/perlitas 10cm':        (0.6,   0.3),
                'Contrapiso c/perlitas 5cm':         (0.3,   0.15),
                'Ho de Pendiente c/perlitas 10cm':   (0.6,   0.3),
                # 08 Revoques
                'R. exterior cal (az+gr+f)':         (1.16,  0.58),
                'R. interior cal (gr+f) fratas':     (1.04,  0.52),
                'R. interior cal (gr+f) fieltro':    (1.14,  0.57),
                'R. grueso interior b/rvto.':        (0.66,  0.33),
                'Cielorraso s/losa (gr+f)':          (1.46,  0.73),
                'Cielo s/metal despl. (az+gr+f)':    (1.64,  0.82),
                'Estructura cielo suspendido (3-6m)':(1.6,   0.8),
                # 09 Revestimientos
                'Enlucido S.Iggam':                  (1.2,   0.6),
                'Enlucido Salpicrete':               (0.46,  0.23),
                'Revest. Texturado':                 (1.2,   0.6),
                # 10 Techos
                'Techo de chapas':                   (3.5,   3.5),
                # 11 Instalaciones
                'Colocacion de Aberturas':           (2.2,   1.1),
                'Instalacion Desague':               (24.5,  24.5),
                'Instalacion Agua F/C':               (31.67, 31.67),
                'Instalacion Gas':                   (27.0,  27.0),
                'Instalacion Electrica':             (5.7,   5.7),
                # 12 Cerámicos
                'Piso ceramico azotea':              (1.49,  0.3),
                'Piso ceramico 1':                   (1.05,  0.21),
                'Piso Ceramico':                     (1.05,  0.21),
                'Piso Porcellanato':                 (1.98,  0.4),
                'Piso loseta cemento':               (0.8,   0.16),
                'Piso calcareo vereda':              (1.17,  0.23),
                'Zocalo ceramico 1':                 (0.37,  0.07),
                'Zocalo Ceramico':                   (0.37,  0.07),
                'Zocalo Porcellanato':               (0.37,  0.07),
                'Zocalo de pino':                    (1.22,  0.24),
                'Rvto. ceramico 1':                  (1.33,  0.27),
                'Rvto. Ceramico':                    (1.33,  0.27),
                'Rvto. Porcellanato':                (2.21,  0.44),
                'Rvto. marmol':                      (3.8,   1.9),
                # 13 Pintura
                'Pintura cal':                       (0.25,  0.0),
                'Pintura latex exterior':            (0.5,   0.0),
                'Pintura latex interior':            (0.5,   0.0),
                'Pintura latex cielos':              (0.5,   0.0),
                'Pintura especial 1':                (0.5,   0.0),
                'Pintura especial 2':                (0.5,   0.0),
                'Enduido s/paredes':                 (0.5,   0.0),
                'Pintura satinol muros':             (0.5,   0.0),
                'Pintura s/carp. madera':            (1.0,   0.0),
                'Pintura s/carp. metalica':          (0.9,   0.0),
                'Ayuda gremios y varios':            (0.5,   0.5),
            }
            for nombre, (hof, hay) in HOF_HAY_2H.items():
                db.execute("UPDATE items_obra SET hof=?, hay=? WHERE nombre=?",
                           (hof, hay, nombre))
            db.execute("""
                UPDATE items_obra SET precio_mo_ars = ROUND(hof*10000 + hay*5000, 2)
                WHERE hof > 0 OR hay > 0
            """)
            db.commit()
            print("[migrate_db] 2h: HOF/HAY actualizados desde datos Excel")

            # Repoblar analisis_sub con nombres de ítems actuales y materiales reales
            # Precios: Jun-2026 del corralón (alineados con precios_v1)
            # Materiales compuestos (Hormigón, Armadura, Encofrado) usan precio referencia
            db.execute("DELETE FROM analisis_sub")
            db.commit()

            SUBS_2H = [
                # ─── 01 Preliminares ─────────────────────────────────────────────────
                ('Cerco de obra','Saligna   1"x2"',         0.7,   1037.97, 1),
                ('Cerco de obra','Saligna 1"x4"',           1.0,    886.08, 1),
                ('Cerco de obra','Clavos 2"',               0.05,  5600.0,  1),
                ('Replanteo','Transporte material suelto',  0.06, 19000.0,  1),
                ('Rebajo de cordon','Transporte material suelto', 0.4, 19000.0, 1),
                ('Demolicion solado 20cm','Transporte material suelto', 2.0, 19000.0, 1),
                ('Demolicion mamposteria 0,15','Transporte material suelto', 2.0, 19000.0, 1),
                ('Demolicion mamposteria 0,15','Martillo neumático', 4.0, 25000.0, 1),
                ('Demolicion mamposteria 0,30','Transporte material suelto', 2.0, 19000.0, 1),
                ('Demolicion mamposteria 0,30','Martillo neumático', 4.0, 25000.0, 1),
                ('Demolicion hormigon armado','Máquina excavadora', 0.1, 25000.0, 1),
                ('Demolicion hormigon armado','Transporte material suelto', 1.5, 19000.0, 1),
                # ─── 02 Fundaciones ──────────────────────────────────────────────────
                ('Relleno y Compactacion C/15cm','Máquina excavadora', 0.12, 25000.0, 1),
                ('Relleno y Compactacion C/15cm','Tierra Colorada',    1.5,  20000.0, 1),
                ('Relleno de arena','Arena común',  0.7, 31000.0, 1),
                ('Mamp. ladrillo comun cimientos','Ladrillos comunes',   400.0,  230.0, 1),
                ('Mamp. ladrillo comun cimientos','Cemento Albañilería',  75.0,  296.0, 1),
                ('Mamp. ladrillo comun cimientos','Arena común',           0.5, 31000.0, 1),
                ('Mamp. ladrillo comun submuracion','Ladrillos comunes',  400.0,  230.0, 1),
                ('Mamp. ladrillo comun submuracion','Cemento Albañilería', 67.0,  296.0, 1),
                ('Mamp. ladrillo comun submuracion','Arena común',          0.5, 31000.0, 1),
                # ─── 03 Hormigón Armado (Ho.Ado.) ────────────────────────────────────
                # Nota: 'Hormigón colado' = hormigón común colado $181.100/m3 (precio ref.)
                ('Ho.Ado. base (50-3)','Hormigón colado',       1.0, 181100.0, 1),
                ('Ho.Ado. base (50-3)','Armadura colocada',    50.0,   3969.0, 1),
                ('Ho.Ado. base (50-3)','Encofrado',             3.0,  24918.0, 1),
                ('Ho.Ado. zapata (60-6)','Hormigón colado',     1.0, 181100.0, 1),
                ('Ho.Ado. zapata (60-6)','Armadura colocada',  60.0,   3969.0, 1),
                ('Ho.Ado. zapata (60-6)','Encofrado',           6.0,  24918.0, 1),
                ('Ho.Ado. viga fundacion (75-7)','Hormigón colado',    1.0, 181100.0, 1),
                ('Ho.Ado. viga fundacion (75-7)','Armadura colocada', 75.0,   3969.0, 1),
                ('Ho.Ado. viga fundacion (75-7)','Encofrado',          7.0,  24918.0, 1),
                ('Ho.Ado. columna (105-11)','Hormigón colado',       1.0, 181100.0, 1),
                ('Ho.Ado. columna (105-11)','Armadura colocada',   105.0,   3969.0, 1),
                ('Ho.Ado. columna (105-11)','Encofrado',            11.0,  24918.0, 1),
                ('Ho.Ado. viga (120-16)','Hormigón colado',         1.0, 181100.0, 1),
                ('Ho.Ado. viga (120-16)','Armadura colocada',     120.0,   3969.0, 1),
                ('Ho.Ado. viga (120-16)','Encofrado',              16.0,  24918.0, 1),
                ('Ho.Ado. tabique 10cm (110-20)','Hormigón colado',     1.0, 181100.0, 1),
                ('Ho.Ado. tabique 10cm (110-20)','Armadura colocada', 110.0,   3969.0, 1),
                ('Ho.Ado. tabique 10cm (110-20)','Encofrado',          20.0,  24918.0, 1),
                ('Ho.Ado. tabique 15cm (110-13)','Hormigón colado',     1.0, 181100.0, 1),
                ('Ho.Ado. tabique 15cm (110-13)','Armadura colocada', 110.0,   3969.0, 1),
                ('Ho.Ado. tabique 15cm (110-13)','Encofrado',          13.33, 24918.0, 1),
                ('Ho.Ado. tanque (90-13)','Hormigón colado',      1.0, 181100.0, 1),
                ('Ho.Ado. tanque (90-13)','Armadura colocada',   90.0,   3969.0, 1),
                ('Ho.Ado. tanque (90-13)','Encofrado',           13.33, 24918.0, 1),
                ('Ho.Ado. losa 10cm (80-10)','Hormigón colado',     1.0, 181100.0, 1),
                ('Ho.Ado. losa 10cm (80-10)','Armadura colocada',  80.0,   3969.0, 1),
                ('Ho.Ado. losa 10cm (80-10)','Encofrado',          10.0,  24918.0, 1),
                ('Ho.Ado. losa 15cm (75-6.66)','Hormigón colado',    1.0, 181100.0, 1),
                ('Ho.Ado. losa 15cm (75-6.66)','Armadura colocada', 75.0,   3969.0, 1),
                ('Ho.Ado. losa 15cm (75-6.66)','Encofrado',          6.66, 24918.0, 1),
                ('Ho.Ado. losa premoldeada','Hormigón colado',        0.07, 181100.0, 1),
                ('Ho.Ado. losa premoldeada','Armadura colocada',      1.0,   3969.0, 1),
                ('Ho.Ado. losa premoldeada','Ladrillo Telgopor 12*38*1m', 2.0, 6200.0, 1),
                ('Ho.Ado. losa premoldeada','Viga Vipret 4m.',        2.0,  3862.5, 1),
                ('Ho.Ado. pavimento 15cm','Hormigón colado',          1.0, 181100.0, 1),
                ('Ho.Ado. pavimento 15cm','Armadura colocada',       50.0,   3969.0, 1),
                ('Ho.Ado. pavimento 15cm','Encofrado',                3.0,  24918.0, 1),
                # ─── 04 Hormigón Elaborado (H.Elab.) ─────────────────────────────────
                ('H.Elab. base (50-3)','Hormigon elaborado colado',   1.0, 190000.0, 1),
                ('H.Elab. base (50-3)','Armadura colocada',          50.0,   3220.0, 1),
                ('H.Elab. base (50-3)','Encofrado',                   3.0,  24918.0, 1),
                ('H.Elab. zapata (60-6)','Hormigon elaborado colado', 1.0, 190000.0, 1),
                ('H.Elab. zapata (60-6)','Armadura colocada',        60.0,   3220.0, 1),
                ('H.Elab. zapata (60-6)','Encofrado',                 6.0,  24918.0, 1),
                ('H.Elab. viga fundacion (75-7)','Hormigon elaborado colado', 1.0, 190000.0, 1),
                ('H.Elab. viga fundacion (75-7)','Armadura colocada',        75.0,   3220.0, 1),
                ('H.Elab. viga fundacion (75-7)','Encofrado',                 7.0,  24918.0, 1),
                ('H.Elab. columna (105-11)','Hormigon elaborado colado', 1.0, 190000.0, 1),
                ('H.Elab. columna (105-11)','Armadura colocada',       105.0,   3220.0, 1),
                ('H.Elab. columna (105-11)','Encofrado',               11.0,  24918.0, 1),
                ('H.Elab. viga (120-16)','Hormigon elaborado colado',   1.0, 190000.0, 1),
                ('H.Elab. viga (120-16)','Armadura colocada',          120.0,   3220.0, 1),
                ('H.Elab. viga (120-16)','Encofrado',                  16.0,  24918.0, 1),
                ('H.Elab. tabique 10cm (110-20)','Hormigon elaborado colado', 1.0, 190000.0, 1),
                ('H.Elab. tabique 10cm (110-20)','Armadura colocada',        110.0,   3220.0, 1),
                ('H.Elab. tabique 10cm (110-20)','Encofrado',               20.0,  24918.0, 1),
                ('H.Elab. tabique 15cm (110-13)','Hormigon elaborado colado', 1.0, 190000.0, 1),
                ('H.Elab. tabique 15cm (110-13)','Armadura colocada',        110.0,   3220.0, 1),
                ('H.Elab. tabique 15cm (110-13)','Encofrado',               13.33, 24918.0, 1),
                ('H.Elab. tanque (90-13)','Hormigon elaborado colado', 1.0, 190000.0, 1),
                ('H.Elab. tanque (90-13)','Armadura colocada',         90.0,   3220.0, 1),
                ('H.Elab. tanque (90-13)','Encofrado',                 13.33, 24918.0, 1),
                ('H.Elab. losa 10cm (80-10)','Hormigon elaborado colado', 1.0, 190000.0, 1),
                ('H.Elab. losa 10cm (80-10)','Armadura colocada',         80.0,   3220.0, 1),
                ('H.Elab. losa 10cm (80-10)','Encofrado',                 10.0,  24918.0, 1),
                ('H.Elab. losa 15cm (75-6.66)','Hormigon elaborado colado', 1.0, 190000.0, 1),
                ('H.Elab. losa 15cm (75-6.66)','Armadura colocada',         75.0,   3220.0, 1),
                ('H.Elab. losa 15cm (75-6.66)','Encofrado',                  6.66, 24918.0, 1),
                ('H.Elab. losa premoldeada','Hormigon elaborado colado',      0.07, 190000.0, 1),
                ('H.Elab. losa premoldeada','Armadura colocada',              1.0,   3220.0, 1),
                ('H.Elab. losa premoldeada','Ladrillo Telgopor 12*38*1m',    2.0,  6200.0, 1),
                ('H.Elab. losa premoldeada','Viga Vipret 4m.',               2.0,  3862.5, 1),
                ('H.Elab. pavimento 15cm','Hormigon elaborado colado',        1.0, 190000.0, 1),
                ('H.Elab. pavimento 15cm','Armadura colocada',               50.0,   3220.0, 1),
                ('H.Elab. pavimento 15cm','Encofrado',                        3.0,  24918.0, 1),
                # ─── 05 Cemento ──────────────────────────────────────────────────────
                ('Cemento: capa aisladora s/muros','Cemento portland bolsas', 15.0,  332.0, 1),
                ('Cemento: capa aisladora s/muros','Arena común',              0.028,31000.0,1),
                ('Carpeta bajo piso','Baldosa cerámica azotea', 1.05, 33000.0, 1),
                ('Carpeta bajo piso','Cemento portland bolsas', 5.0,    332.0, 1),
                ('Carpeta bajo piso','Cal aérea Milagro',        4.5,   372.0, 1),
                ('Carpeta bajo piso','Arena común',              0.03, 31000.0, 1),
                ('Cemento: carpeta azotea','Cemento portland bolsas', 1.5,  332.0, 1),
                ('Cemento: carpeta azotea','Arena común',              0.005,31000.0,1),
                ('Cemento: toma de juntas','Cemento portland bolsas', 21.8,  332.0, 1),
                ('Cemento: toma de juntas','Arena común',              0.03, 31000.0, 1),
                ('Cemento: toma de juntas','Hidrófugo',               0.75,  2900.0, 1),
                ('Cemento: azotado','Cemento portland bolsas', 11.0,  332.0, 1),
                ('Cemento: azotado','Arena común',              0.025,31000.0,1),
                ('Cemento: azotado','Hidrófugo',                0.25, 2900.0, 1),
                ('Cemento: revoque','Cemento portland bolsas', 21.8,  332.0, 1),
                ('Cemento: revoque','Arena común',              0.03, 31000.0, 1),
                ('Cemento: revoque','Hidrófugo',                0.75, 2900.0, 1),
                ('Cemento: zocalo 10cm','Super Iggam', 12.0, 833.33, 1),
                ('Cemento: piso rodillado','Cemento portland bolsas', 17.0,  332.0, 1),
                ('Cemento: piso rodillado','Arena común',              0.028,31000.0,1),
                ('Cemento: piso alisado','Cemento portland bolsas', 13.5,  332.0, 1),
                ('Cemento: piso alisado','Arena común',              0.025,31000.0,1),
                # ─── 06 Mampostería ───────────────────────────────────────────────────
                ('Mamp. ladrillo comun 15cm','Ladrillos comunes',   400.0,  230.0, 1),
                ('Mamp. ladrillo comun 15cm','Cemento Albañilería',  67.0,  296.0, 1),
                ('Mamp. ladrillo comun 15cm','Arena común',           0.5, 31000.0, 1),
                # Comun 30cm: mismas proporciones que 15cm (corrección del Excel)
                ('Mamp. ladrillo comun 30cm','Ladrillos comunes',   400.0,  230.0, 1),
                ('Mamp. ladrillo comun 30cm','Cemento Albañilería',  67.0,  296.0, 1),
                ('Mamp. ladrillo comun 30cm','Arena común',           0.5, 31000.0, 1),
                ('Mamp. ladrillo vista 15cm','Ladrillos vista',     430.0,  390.0, 1),
                ('Mamp. ladrillo vista 15cm','Cemento Albañilería',  75.0,  296.0, 1),
                ('Mamp. ladrillo vista 15cm','Arena común',           0.5, 31000.0, 1),
                ('Mamp. ladrillo vista 30cm','Ladrillos vista',     430.0,  390.0, 1),
                ('Mamp. ladrillo vista 30cm','Cemento Albañilería',  75.0,  296.0, 1),
                ('Mamp. ladrillo vista 30cm','Arena común',           0.5, 31000.0, 1),
                ('Tabique ladrillo comun 8cm','Ladrillo hueco 8x18x33cm', 17.0, 720.0,  1),
                ('Tabique ladrillo comun 8cm','Cemento Albañilería',       7.0,  296.0, 1),
                ('Tabique ladrillo comun 8cm','Arena común',               0.05,31000.0,1),
                ('Tabique ladrillo hueco 8x18x33','Ladrillo hueco 12X18X33cm', 17.0, 830.0,  1),
                ('Tabique ladrillo hueco 8x18x33','Cemento Albañilería',        7.0,  296.0, 1),
                ('Tabique ladrillo hueco 8x18x33','Arena común',                0.05,31000.0,1),
                ('Tabique ladrillo hueco 12x18x33','Ladrillo hueco 18X18X33cm', 17.0, 1160.0, 1),
                ('Tabique ladrillo hueco 12x18x33','Cemento Albañilería',        7.0,   296.0, 1),
                ('Tabique ladrillo hueco 12x18x33','Arena común',                0.05, 31000.0,1),
                ('Tabique ladrillo hueco 18x18x33','Ladrillo hueco Portante 12x18x33cm',17.0,1160.0,1),
                ('Tabique ladrillo hueco 18x18x33','Cemento Albañilería',                7.0,  296.0,1),
                ('Tabique ladrillo hueco 18x18x33','Arena común',                        0.05,31000.0,1),
                ('Tabique ladrillo hueco 18x18x33','Hierro redondo d=10mm',              2.0, 2553.76,1),
                ('Tabique hueco portante 12x19x33cm','Ladrillo hueco Portante 18x18x33cm',17.0,1400.0,1),
                ('Tabique hueco portante 12x19x33cm','Cemento Albañilería',               7.5,  296.0,1),
                ('Tabique hueco portante 12x19x33cm','Arena común',                       0.05,31000.0,1),
                ('Tabique hueco portante 12x19x33cm','Hierro redondo d=10mm',             3.0, 2553.76,1),
                ('Tabique hueco portante 18x19x33cm','Ladrillos comunes',  10.0,  230.0, 1),
                ('Tabique hueco portante 18x19x33cm','Cemento Albañilería', 2.35, 296.0, 1),
                ('Tabique hueco portante 18x19x33cm','Arena común',         0.02, 31000.0,1),
                ('Tabique hueco portante 18x19x33cm','Hierro redondo d=10mm',3.0, 2553.76,1),
                ('Dinteles pared 15cm','Ladrillos comunes',    20.0,  230.0,  1),
                ('Dinteles pared 15cm','Cemento Albañilería',   4.7,  296.0,  1),
                ('Dinteles pared 15cm','Arena común',           0.04,31000.0, 1),
                ('Dinteles pared 15cm','Hierro redondo d=10mm', 5.0, 2553.76, 1),
                ('Dinteles pared 30cm','Cemento portland bolsas', 2.7,  332.0, 1),
                ('Dinteles pared 30cm','Arena común',             0.006,31000.0,1),
                ('Dinteles pared 30cm','Hidrófugo',               0.13, 2900.0,1),
                # ─── 07 Contrapisos ───────────────────────────────────────────────────
                ('Contrapiso cascotes 15cm','Cemento portland bolsas',  1.5,  332.0,  1),
                ('Contrapiso cascotes 15cm','Cemento Albañilería',      14.29,296.0,  1),
                ('Contrapiso cascotes 15cm','Arena común',              0.05, 31000.0, 1),
                ('Contrapiso cascotes 15cm','Granza',                   0.08, 36000.0, 1),
                ('Contrapiso cascotes 10cm','Cemento portland bolsas',  1.5,  332.0,  1),
                ('Contrapiso cascotes 10cm','Cemento Albañilería',      14.29,296.0,  1),
                ('Contrapiso cascotes 10cm','Arena común',              0.05, 31000.0, 1),
                ('Contrapiso cascotes 10cm','Granza',                   0.08, 36000.0, 1),
                ('Banquina Ho Pobre 10cm','Perlitas Telgopor (75 Lts)', 106.0, 106.67, 1),
                ('Banquina Ho Pobre 10cm','Cemento Albañilería',         10.4,  296.0, 1),
                ('Banquina Ho Pobre 10cm','Arena común',                0.014,31000.0, 1),
                ('Banquina Ho Pobre 10cm','Granza',                     0.014,36000.0, 1),
                ('Contrapiso c/perlitas 10cm','Perlitas Telgopor (75 Lts)', 70.0, 106.67, 1),
                ('Contrapiso c/perlitas 10cm','Cemento portland bolsas',     6.5,  332.0, 1),
                ('Contrapiso c/perlitas 10cm','Arena común',             0.014,31000.0, 1),
                ('Contrapiso c/perlitas 10cm','Granza',                  0.014,36000.0, 1),
                ('Contrapiso c/perlitas 5cm','Cemento portland bolsas',  13.5, 332.0,  1),
                ('Contrapiso c/perlitas 5cm','Arena común',              0.025,31000.0, 1),
                ('Ho de Pendiente c/perlitas 10cm','Perlitas Telgopor (75 Lts)',106.0,106.67,1),
                ('Ho de Pendiente c/perlitas 10cm','Cemento portland bolsas',   10.5,  332.0,1),
                ('Ho de Pendiente c/perlitas 10cm','Arena común',               0.02,31000.0,1),
                ('Ho de Pendiente c/perlitas 10cm','Granza',                    0.02,36000.0,1),
                # ─── 08 Revoques ─────────────────────────────────────────────────────
                ('R. exterior cal (az+gr+f)','Cemento portland bolsas', 0.85, 332.0, 1),
                ('R. exterior cal (az+gr+f)','Cemento Albañilería',     4.71, 296.0, 1),
                ('R. exterior cal (az+gr+f)','Arena común',             0.025,31000.0,1),
                ('R. interior cal (gr+f) fratas','Cemento portland bolsas', 0.85, 332.0, 1),
                ('R. interior cal (gr+f) fratas','Cemento Albañilería',     4.71, 296.0, 1),
                ('R. interior cal (gr+f) fratas','Arena común',             0.025,31000.0,1),
                ('R. interior cal (gr+f) fieltro','Cemento portland bolsas', 0.7,  332.0, 1),
                ('R. interior cal (gr+f) fieltro','Cemento Albañilería',     2.1,  296.0, 1),
                ('R. interior cal (gr+f) fieltro','Arena común',             0.015,31000.0,1),
                ('R. grueso interior b/rvto.','Cemento portland bolsas', 32.7,  332.0, 1),
                ('R. grueso interior b/rvto.','Arena común',             0.045,31000.0, 1),
                ('Cielorraso s/losa (gr+f)','Cemento portland bolsas', 1.5,  332.0, 1),
                ('Cielorraso s/losa (gr+f)','Cemento Albañilería',     4.71, 296.0, 1),
                ('Cielorraso s/losa (gr+f)','Arena común',             0.03, 31000.0,1),
                ('Cielo s/metal despl. (az+gr+f)','Metal desplegado',  1.15, 3333.33,1),
                ('Cielo s/metal despl. (az+gr+f)','Palito 1"x1"',      7.0,   500.0, 1),
                ('Cielo s/metal despl. (az+gr+f)','Clavos 2"',         0.15, 5600.0, 1),
                ('Estructura cielo suspendido (3-6m)','Cemento portland bolsas', 2.0,  332.0, 1),
                ('Estructura cielo suspendido (3-6m)','Cemento Albañilería',    20.0,  296.0, 1),
                ('Estructura cielo suspendido (3-6m)','Arena común',            0.075,31000.0,1),
                ('Estructura cielo suspendido (3-6m)','Granza',                 0.12, 36000.0,1),
                # ─── 09 Revestimientos ───────────────────────────────────────────────
                ('Enlucido S.Iggam','Salpicrete',  6.0, 2666.67, 1),
                ('Enlucido Salpicrete','Fondo Base',3.0, 3800.0, 1),
                ('Revest. Texturado','Cemento portland bolsas', 4.25,  332.0, 1),
                ('Revest. Texturado','Cemento Albañilería',     4.71,  296.0, 1),
                ('Revest. Texturado','Hidrófugo',               0.19, 2900.0, 1),
                ('Revest. Texturado','Arena común',             0.04, 31000.0, 1),
                # ─── 10 Techos ───────────────────────────────────────────────────────
                ('Techo de chapas','Chapas Techo',     1.1,  14800.0, 1),
                ('Techo de chapas','Clavadores 2 x 2', 2.0,  1007.59, 1),
                ('Techo de chapas','Tirantes 2x6',     0.3, 27166.67, 1),
                ('Techo de chapas','Issolant',          1.1,  4950.0,  1),
                ('Techo de chapas','Clavos 2" 1/2',    0.05, 5600.0,  1),
                ('Techo de chapas','Tornillo c/arand goma', 10.0, 118.0, 1),
                # ─── 11 Instalaciones ────────────────────────────────────────────────
                ('Colocacion de Aberturas','Cemento portland bolsas', 4.0, 332.0,  1),
                ('Colocacion de Aberturas','Arena común',             0.02,31000.0,1),
                ('Instalacion Desague','Caño Awaduct 110',   11.2, 8125.0,  1),
                ('Instalacion Desague','Caño Awaduct 63',     2.4, 6250.0,  1),
                ('Instalacion Desague','Caño Awaduct 50',     1.0, 5300.0,  1),
                ('Instalacion Desague','Caño Awaduct 40',     3.2, 2900.0,  1),
                ('Instalacion Desague','Accesorios Desagues',10.0, 8500.0,  1),
                ('Instalacion Desague','Cemento portland bolsas', 4.0, 332.0, 1),
                ('Instalacion Desague','Arena común',         0.02,31000.0, 1),
                ('Instalacion Agua F/C','Caño Epoxi 3/4',    8.0, 12890.63, 1),
                ('Instalacion Agua F/C','Caño epoxi 1/2',    2.2,  6406.25, 1),
                ('Instalacion Agua F/C','Accesorios Gas',    12.0,  3500.0,  1),
                ('Instalacion Agua F/C','Llaves de Paso Gas',1.0,  25000.0,  1),
                ('Instalacion Agua F/C','Cemento portland bolsas', 2.0, 332.0, 1),
                ('Instalacion Agua F/C','Arena común',        0.03,31000.0, 1),
                ('Instalacion Gas','Caño TF 25',   8.0,  5000.0, 1),
                ('Instalacion Gas','Caño TF 20',  14.0,  2625.0, 1),
                ('Instalacion Gas','Accesorios TF',27.0, 3000.0, 1),
                ('Instalacion Gas','Llaves de Paso Agua', 3.0, 25000.0, 1),
                ('Instalacion Gas','Cemento portland bolsas', 2.0, 332.0, 1),
                ('Instalacion Gas','Arena común',   0.026,31000.0,1),
                ('Instalacion Electrica','Caño Corrugado 1"',  1.0,  376.0, 1),
                ('Instalacion Electrica','Caño Corrugado 3/4"',2.0,  260.0, 1),
                ('Instalacion Electrica','Cajas Metalicas',    1.0,  850.0, 1),
                ('Instalacion Electrica','Cable 2,5 mm',       2.5,  850.0, 1),
                ('Instalacion Electrica','Cable 1,5 mm',       5.5,  510.0, 1),
                ('Instalacion Electrica','Cemento portland bolsas', 2.0, 332.0, 1),
                ('Instalacion Electrica','Arena común',        0.01,31000.0,1),
                # ─── 12 Cerámicos ─────────────────────────────────────────────────────
                ('Piso ceramico azotea','Baldosa cerámica azotea', 1.05, 33000.0, 1),
                ('Piso ceramico azotea','Klaukol',                  4.0,   720.0, 1),
                ('Piso ceramico azotea','Pastina',                  0.2,  5500.0, 1),
                ('Piso ceramico 1','Piso cerámico 1',               1.05,15000.0, 1),
                ('Piso ceramico 1','Klaukol',                        4.0,  720.0, 1),
                ('Piso ceramico 1','Pastina',                        0.2, 5500.0, 1),
                ('Piso Ceramico','Piso cerámico 2',                 1.05,25000.0, 1),
                ('Piso Ceramico','Klaukol',                          4.0,  720.0, 1),
                ('Piso Ceramico','Pastina',                          0.2, 5500.0, 1),
                ('Piso Porcellanato','Piso cerámico 3 (porcellanato)', 1.05,35000.0,1),
                ('Piso Porcellanato','Klaukol',                          4.0,  720.0,1),
                ('Piso Porcellanato','Pastina',                          0.2, 5500.0,1),
                ('Piso loseta cemento','Mosaico calcáreo',          1.1,  25000.0, 1),
                ('Piso loseta cemento','Cemento portland bolsas',   5.0,    332.0, 1),
                ('Piso loseta cemento','Cal aérea Milagro',          4.5,   372.0, 1),
                ('Piso loseta cemento','Arena común',                0.03,31000.0, 1),
                ('Piso calcareo vereda','Loseta cemento 60x40cm',  1.05, 5000.0, 1),
                ('Piso calcareo vereda','Klaukol',                  0.2,  720.0, 1),
                ('Piso calcareo vereda','Pastina',                  0.02, 5500.0,1),
                ('Zocalo ceramico 1','Zócalo cerámico 1',           1.05, 1250.0, 1),
                ('Zocalo ceramico 1','Klaukol',                      0.2,   720.0, 1),
                ('Zocalo ceramico 1','Pastina',                      0.02, 5500.0, 1),
                ('Zocalo Ceramico','Zócalo cerámico 2',              1.05, 2083.33,1),
                ('Zocalo Ceramico','Klaukol',                         0.2,   720.0,1),
                ('Zocalo Ceramico','Pastina',                         0.02, 5500.0,1),
                ('Zocalo Porcellanato','Zócalo cerámico 3 (Porcellanato)',1.05,2916.67,1),
                ('Zocalo Porcellanato','Klaukol',                          0.2,  720.0,1),
                ('Zocalo Porcellanato','Pastina',                          0.02,5500.0,1),
                ('Zocalo de pino','Zócalo de madera',               1.05, 3500.0, 1),
                ('Rvto. ceramico 1','Rvto.cerámico 1',              1.1, 15000.0, 1),
                ('Rvto. ceramico 1','Klaukol',                       2.0,   720.0, 1),
                ('Rvto. ceramico 1','Pastina',                       0.2,  5500.0, 1),
                ('Rvto. Ceramico','Rvto.cerámico 2',                 1.1, 25000.0, 1),
                ('Rvto. Ceramico','Klaukol',                          2.0,   720.0, 1),
                ('Rvto. Ceramico','Pastina',                          0.2,  5500.0, 1),
                ('Rvto. Porcellanato','Rvto.cerámico 3 (porcellanato)', 1.1, 35000.0,1),
                ('Rvto. Porcellanato','Klaukol',                          2.0,   720.0,1),
                ('Rvto. Porcellanato','Pastina',                          0.2,  5500.0,1),
                ('Rvto. marmol','Rvto.cerámico 1',  1.1, 15000.0, 1),
                ('Rvto. marmol','Klaukol',            2.0,   720.0, 1),
                ('Rvto. marmol','Pastina',            0.2,  5500.0, 1),
                # ─── 13 Pintura ───────────────────────────────────────────────────────
                ('Pintura cal','Pintura cal hidráulica',       0.36, 5817.0, 1),
                ('Pintura latex exterior','Pintura látex exterior', 0.36, 7250.0, 1),
                ('Pintura latex interior','Pintura látex interior', 0.36, 4750.0, 1),
                ('Pintura latex cielos','Pintura látex cielos',    0.36, 7200.0, 1),
                ('Pintura especial 1','Pintura especial 1',         0.36,21250.0, 1),
                ('Pintura especial 2','Pintura especial 2',         0.36,22500.0, 1),
                ('Enduido s/paredes','Enduido sintético',           0.6,  7500.0, 1),
                ('Pintura satinol muros','Pintura satinol',         0.36,22500.0, 1),
                ('Pintura s/carp. madera','Esmalte albalux',       0.36,23175.0, 1),
                ('Pintura s/carp. metalica','Esmalte albalux',     0.36,23175.0, 1),
            ]

            db.executemany(
                "INSERT INTO analisis_sub "
                "(item_nombre, sub_nombre, cant_por_unit, precio_ars, es_material) "
                "VALUES (?,?,?,?,?)",
                SUBS_2H
            )
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave, valor) VALUES ('2h_done','2026-06-30')")
            db.execute("INSERT OR REPLACE INTO config (clave, valor) VALUES ('precios_updated_at','2026-06-30')")
            db.commit()
            print(f"[migrate_db] 2h: analisis_sub repoblado ({len(SUBS_2H)} filas) + HOF/HAY corregidos")

        # ── 2i. Columnas fecha_actualizacion e historial_json en presupuestos ──
        ya_2i = db.execute("SELECT valor FROM config WHERE clave='2i_done'").fetchone()
        if not ya_2i:
            try:
                db.execute("ALTER TABLE presupuestos ADD COLUMN fecha_actualizacion TEXT")
            except Exception:
                pass
            try:
                db.execute("ALTER TABLE presupuestos ADD COLUMN historial_json TEXT DEFAULT '[]'")
            except Exception:
                pass
            # precios_updated_at por si 2h ya corrió antes sin este valor
            db.execute("INSERT OR IGNORE INTO config (clave, valor) VALUES ('precios_updated_at','2026-06-30')")
            db.execute("INSERT OR REPLACE INTO config (clave, valor) VALUES ('2i_done','2026-06-30')")
            db.commit()
            print("[migrate_db] 2i: columnas fecha_actualizacion + historial_json agregadas")

        # ── 2j. analisis_sub → precios por unidad comercial de venta ─────────
        # Convierte precios almacenados por unidad mínima a la unidad con la que
        # se compra en corralón. El total (cant × precio) se mantiene idéntico.
        #   Ladrillos comunes/vista : unidad → millar (×1000)
        #   Cemento Albañilería     : kg     → bolsa 25 kg (×25)
        #   Cemento portland bolsas : kg     → bolsa 50 kg (×50)
        #   Cal aérea Milagro       : kg     → bolsa 25 kg (×25)
        #   Hierro redondo          : por kg → sin cambio
        #   Arena/Granza/m³         : ya están en m³ → sin cambio
        ya_2j = db.execute("SELECT valor FROM config WHERE clave='2j_done'").fetchone()
        if not ya_2j:
            db.execute("""
                UPDATE analisis_sub
                SET cant_por_unit = ROUND(cant_por_unit / 1000.0, 6),
                    precio_ars    = ROUND(precio_ars * 1000, 2)
                WHERE sub_nombre IN ('Ladrillos comunes','Ladrillos vista')
                  AND es_material = 1
            """)
            db.execute("""
                UPDATE analisis_sub
                SET cant_por_unit = ROUND(cant_por_unit / 25.0, 4),
                    precio_ars    = ROUND(precio_ars * 25, 2)
                WHERE sub_nombre = 'Cemento Albañilería'
                  AND es_material = 1
            """)
            db.execute("""
                UPDATE analisis_sub
                SET cant_por_unit = ROUND(cant_por_unit / 50.0, 4),
                    precio_ars    = ROUND(precio_ars * 50, 2)
                WHERE sub_nombre = 'Cemento portland bolsas'
                  AND es_material = 1
            """)
            db.execute("""
                UPDATE analisis_sub
                SET cant_por_unit = ROUND(cant_por_unit / 25.0, 4),
                    precio_ars    = ROUND(precio_ars * 25, 2)
                WHERE sub_nombre LIKE 'Cal a%'
                  AND es_material = 1
            """)
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2j_done','2026-06-30')")
            db.commit()
            print("[migrate_db] 2j: analisis_sub convertido a unidades comerciales")

        # ── 2k. Unidades comerciales — resto de materiales ───────────────────
        #   Klaukol             : kg → bolsa 30 kg  (×30)
        #   Hidrófugo           : L  → bidón 5 L    (×5)
        #   Super Iggam         : kg → bolsa 20 kg  (×20)
        #   Salpicrete          : kg → bolsa 20 kg  (×20)
        #   Fondo Base          : L  → bidón 20 L   (×20)
        #   Enduido sintético   : kg → balde 30 kg  (×30)
        #   Pintura látex (ext/int/cielos) : L → balde 20 L (×20)
        #   Pintura especial 1/2, satinol  : L → balde 20 L (×20)
        #   Esmalte albalux     : L → lata 4 L      (×4)
        #   Pintura cal         : kg → bolsa 25 kg  (×25)
        ya_2k = db.execute("SELECT valor FROM config WHERE clave='2k_done'").fetchone()
        if not ya_2k:
            conversiones_2k = [
                ("Klaukol",                     30,  30),
                ("Hidrófugo",                    5,   5),
                ("Super Iggam",                 20,  20),
                ("Salpicrete",                  20,  20),
                ("Fondo Base",                  20,  20),
                ("Enduido sintético",           30,  30),
                ("Pintura látex exterior",      20,  20),
                ("Pintura látex interior",      20,  20),
                ("Pintura látex cielos",        20,  20),
                ("Pintura especial 1",          20,  20),
                ("Pintura especial 2",          20,  20),
                ("Pintura satinol",             20,  20),
                ("Esmalte albalux",              4,   4),
                ("Pintura cal hidráulica",      25,  25),
            ]
            for sub, div_cant, mul_precio in conversiones_2k:
                db.execute("""
                    UPDATE analisis_sub
                    SET cant_por_unit = ROUND(cant_por_unit / ?, 6),
                        precio_ars    = ROUND(precio_ars * ?, 2)
                    WHERE sub_nombre = ? AND es_material = 1
                """, (div_cant, mul_precio, sub))
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2k_done','2026-06-30')")
            db.commit()
            print("[migrate_db] 2k: unidades comerciales aplicadas (Klaukol, Hidrófugo, pinturas, etc.)")

        # ── 2l. Corrección de factores comerciales incorrectos en 2j y 2k ──────
        #   Referencia: LISTA_MATERIALES_V3_formulafix.xlsx (Lista 169, 04/06/2026)
        #   Fórmula: cant_nueva = cant_actual x (factor_viejo / factor_nuevo)
        #            precio_nuevo = precio_actual x (factor_nuevo / factor_viejo)
        ya_2l = db.execute("SELECT valor FROM config WHERE clave='2l_done'").fetchone()
        if not ya_2l:
            correcciones_2l = [
                # (sub_nombre,                  factor_viejo, factor_nuevo)
                ("Cemento portland bolsas",           50,  25),  # bolsa 25kg, no 50
                ("Klaukol",                           30,  25),  # bolsa 25kg, no 30
                ("Hidrófugo",                    5,  10),  # balde 10kg, no bidon 5L
                ("Super Iggam",                       20,  30),  # bolsa 30kg, no 20
                ("Salpicrete",                        20,  30),  # bolsa 30kg, no 20
                ("Fondo Base",                        20,  25),  # balde 25kg, no bidon 20L
                ("Enduido sintético",            30,   4),  # lata 4kg, no balde 30
                ("Pintura látex cielos",         20,  10),  # balde 10L, no 20
                ("Pintura especial 1",                20,   4),  # lata 4L, no balde 20
                ("Pintura especial 2",                20,   4),  # lata 4L, no balde 20
                ("Pintura satinol",                   20,   4),  # lata 4L, no balde 20
                ("Pintura cal hidráulica",       25,   1),  # revertir -- sin presentacion en Excel
            ]
            for sub, f_viejo, f_nuevo in correcciones_2l:
                db.execute("""
                    UPDATE analisis_sub
                    SET cant_por_unit = ROUND(cant_por_unit * (? * 1.0 / ?), 6),
                        precio_ars    = ROUND(precio_ars    * (? * 1.0 / ?), 2)
                    WHERE sub_nombre = ? AND es_material = 1
                """, (f_viejo, f_nuevo, f_nuevo, f_viejo, sub))
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2l_done','2026-06-30')")
            db.commit()
            print("[migrate_db] 2l: factores comerciales corregidos (11 materiales + revert cal)")

        # ── 2m. Corrección HOF/HAY de las 4 instalaciones (estaban cruzadas) ────
        #   Verificado contra PRESUPUESTO COCHERA.xlsx (hoja Análisis, bloque propio
        #   de cada instalación — Oficial/Ayudante Plomería/Electricista/Gasista):
        #   Desagües=24,5/24,5 · Agua F/C=31,67/31,67 · Gas=27/27 · Eléctrica=5,7/5,7.
        #   La migración 2h había cargado estos 4 valores desde la columna de
        #   referencia (Items!I:J), que está desalineada respecto al desglose real.
        ya_2m = db.execute("SELECT valor FROM config WHERE clave='2m_done'").fetchone()
        if not ya_2m:
            HOF_HAY_2M = {
                'Instalacion Desague':   (24.5,  24.5),
                'Instalacion Agua F/C':  (31.67, 31.67),
                'Instalacion Gas':       (27.0,  27.0),
                'Instalacion Electrica': (5.7,   5.7),
            }
            for nombre, (hof, hay) in HOF_HAY_2M.items():
                db.execute("UPDATE items_obra SET hof=?, hay=? WHERE nombre=?",
                           (hof, hay, nombre))
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2m_done','2026-07-04')")
            db.commit()
            print("[migrate_db] 2m: HOF/HAY de instalaciones corregidos (Desagues/AguaFC/Gas/Electrica)")

        # ── 2n. Resync completo de analisis_sub contra PRESUPUESTO COCHERA.xlsx ──
        #   Verificado ítem por ítem (04/07/2026) contra la hoja "Análisis" del
        #   Excel real de Ezequiel Petrini (que trae el maestro completo, no solo
        #   la obra puntual). Las cantidades/precios de materiales que traía
        #   analisis_sub desde 2h estaban desactualizadas para la gran mayoría de
        #   los ítems (79 de 93 comparados tenían al menos una diferencia).
        #   Reemplaza TODOS los materiales (es_material=1) de los ítems que están
        #   en el Excel; los 9 ítems que no están en el Excel (variantes propias
        #   de la app: Piso/Rvto/Zocalo Cerámico "2"/"3", Relleno y Compactación
        #   C/15cm, Demolición mampostería 0,15/0,30) quedan sin tocar.
        ya_2n = db.execute("SELECT valor FROM config WHERE clave='2n_done'").fetchone()
        if not ya_2n:
            SUBS_2N = [
                ('Casilla rodante obrador', 'Transporte casilla', 1.0, 150000.0, 1),
                ('Cartel de obra', 'Cartel de obra', 1.0, 80000.0, 1),
                ('Cerco de obra', 'Saligna 3"x3"', 3.0, 1766.67, 1),
                ('Cerco de obra', 'Chapas Cerco', 2.0, 14800.0, 1),
                ('Cerco de obra', 'Clavos 2"', 0.06, 5600.0, 1),
                ('Replanteo', 'Saligna   1"x2"', 0.7, 1037.97, 1),
                ('Replanteo', 'Saligna 1"x4"', 1.0, 886.08, 1),
                ('Replanteo', 'Clavos 2"', 0.05, 5600.0, 1),
                ('Picado de revoques', 'Transporte material suelto', 0.06, 19000.0, 1),
                ('Demolicion solado 20cm', 'Transporte material suelto', 0.4, 19000.0, 1),
                ('Demolicion mamposteria', 'Transporte material suelto', 2.0, 19000.0, 1),
                ('Demolicion hormigon armado', 'Transporte material suelto', 2.0, 19000.0, 1),
                ('Demolicion hormigon armado', 'Martillo neumático', 4.0, 15000.0, 1),
                ('Desmonte a maquina', 'Máquina topadora', 0.1, 50000.0, 1),
                ('Desmonte a maquina', 'Transporte material suelto', 1.5, 19000.0, 1),
                ('Excavacion maquina', 'Máquina excavadora', 0.05, 50000.0, 1),
                ('Excavacion maquina', 'Transporte material suelto', 1.5, 19000.0, 1),
                ('Relleno y Compactación', 'Tierra Colorada', 1.58, 20000.0, 1),
                ('Suelo cemento 6%', 'Máquina topadora', 0.12, 50000.0, 1),
                ('Suelo cemento 6%', 'Tierra Colorada', 1.5, 20000.0, 1),
                ('Suelo cemento 6%', 'Cemento portland bolsas', 162.0, 332.0, 1),
                ('Relleno de arena', 'Arena común', 1.0, 31000.0, 1),
                ('Hormigon común colado (6,00hh)', 'Cemento portland bolsas', 330.0, 332.0, 1),
                ('Hormigon común colado (6,00hh)', 'Arena común', 0.7, 31000.0, 1),
                ('Hormigon común colado (6,00hh)', 'Canto rodado', 0.7, 93000.0, 1),
                ('Armadura colocada (0,20hh)', 'Hierro redondo d=10mm', 1.05, 2553.76, 1),
                ('Armadura colocada (0,20hh)', 'Alambre negro', 0.01, 7000.0, 1),
                ('Encofrado (2,60hh)', 'Pino encofrado 1"', 0.35, 8860.76, 1),
                ('Encofrado (2,60hh)', 'Clavos', 0.3, 5600.0, 1),
                ('Ho.Ao. según cant.de horm.arm y enc', 'Hormigón colado', 1.0, 215000.0, 1),
                ('Ho.Ao. según cant.de horm.arm y enc', 'Armadura colocada', 90.0, 3501.45, 1),
                ('Ho.Ao. según cant.de horm.arm y enc', 'Encofrado', 10.0, 25781.27, 1),
                ('Ho.Ado. pavimento 15cm', 'Hormigón colado', 0.15, 233860.0, 1),
                ('Ho.Ado. pavimento 15cm', 'Armadura colocada', 2.5, 4251.45, 1),
                ('Ho.Ado. pavimento 15cm', 'Junta (encofrado equivalente)', 0.3, 25781.27, 1),
                ('Ho.Ado. base (50-3)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. base (50-3)', 'Armadura colocada', 50.0, 4251.45, 1),
                ('Ho.Ado. base (50-3)', 'Encofrado', 3.0, 25781.27, 1),
                ('Ho.Ado.zapata (30-6)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado.zapata (30-6)', 'Armadura colocada', 30.0, 4251.45, 1),
                ('Ho.Ado.zapata (30-6)', 'Encofrado', 6.0, 25781.27, 1),
                ('Ho.Ado. zapata (60-6)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. zapata (60-6)', 'Armadura colocada', 60.0, 4251.45, 1),
                ('Ho.Ado. zapata (60-6)', 'Encofrado', 6.0, 25781.27, 1),
                ('Ho.Ado. columna (105-11)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. columna (105-11)', 'Armadura colocada', 105.0, 4251.45, 1),
                ('Ho.Ado. columna (105-11)', 'Encofrado', 11.0, 25781.27, 1),
                ('Ho.Ado. viga fundacion (75-7)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. viga fundacion (75-7)', 'Armadura colocada', 75.0, 4251.45, 1),
                ('Ho.Ado. viga fundacion (75-7)', 'Encofrado', 7.0, 25781.27, 1),
                ('Ho.Ado. viga (120-16)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. viga (120-16)', 'Armadura colocada', 120.0, 4251.45, 1),
                ('Ho.Ado. viga (120-16)', 'Encofrado', 16.0, 25781.27, 1),
                ('Ho.Ado. losa 10cm (80-10)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. losa 10cm (80-10)', 'Armadura colocada', 80.0, 4251.45, 1),
                ('Ho.Ado. losa 10cm (80-10)', 'Encofrado', 10.0, 25781.27, 1),
                ('Ho.Ado. losa 15cm (75-6.66)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. losa 15cm (75-6.66)', 'Armadura colocada', 75.0, 4251.45, 1),
                ('Ho.Ado. losa 15cm (75-6.66)', 'Encofrado', 6.66, 25781.27, 1),
                ('Ho.Ado. losa premoldeada', 'Hormigón colado', 0.07, 233860.0, 1),
                ('Ho.Ado. losa premoldeada', 'Armadura colocada', 1.0, 4251.45, 1),
                ('Ho.Ado. losa premoldeada', 'Ladrillo telgopor 12*38*1m', 2.0, 6200.0, 1),
                ('Ho.Ado. losa premoldeada', 'Viga Vipret 4m.', 2.0, 3862.5, 1),
                ('Ho.Ado. tabique 10cm (110-20)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. tabique 10cm (110-20)', 'Armadura colocada', 110.0, 4251.45, 1),
                ('Ho.Ado. tabique 10cm (110-20)', 'Encofrado', 20.0, 25781.27, 1),
                ('Ho.Ado. tabique 15cm (110-13)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. tabique 15cm (110-13)', 'Armadura colocada', 110.0, 4251.45, 1),
                ('Ho.Ado. tabique 15cm (110-13)', 'Encofrado', 13.33, 25781.27, 1),
                ('Ho.Ado. tanque (90-13)', 'Hormigón colado', 1.0, 233860.0, 1),
                ('Ho.Ado. tanque (90-13)', 'Armadura colocada', 90.0, 4251.45, 1),
                ('Ho.Ado. tanque (90-13)', 'Encofrado', 13.33, 25781.27, 1),
                ('Zapata Ho Pobre', 'Cemento portland bolsas', 225.0, 332.0, 1),
                ('Zapata Ho Pobre', 'Cal Hidráulica', 350.0, 332.0, 1),
                ('Zapata Ho Pobre', 'Arena común', 1.68, 31000.0, 1),
                ('Zapata Ho Pobre', 'Granza', 2.25, 36000.0, 1),
                ('Pintura cal', 'Cal aérea Milagro', 0.2, 372.0, 1),
                ('Pintura cal', 'Color pintura cal', 0.05, 20000.0, 1),
                ('Pintura latex exterior', 'Pintura látex exterior', 0.36, 7250.0, 1),
                ('Pintura latex interior', 'Pintura látex interior', 0.36, 4750.0, 1),
                ('Pintura latex cielos', 'Pintura látex cielos', 0.36, 7200.0, 1),
                ('Pintura especial 1', 'Pintura especial 1', 0.36, 14500.0, 1),
                ('Pintura especial 2', 'Pintura especial 2', 0.36, 16250.0, 1),
                ('Enduido s/paredes', 'Enduido sintético', 0.6, 7500.0, 1),
                ('Pintura satinol muros', 'Pintura satinol', 0.36, 22500.0, 1),
                ('Pintura s/carp. madera', 'Esmalte albalux', 0.36, 23175.0, 1),
                ('Pintura s/carp. metalica', 'Esmalte albalux', 0.36, 23175.0, 1),
                ('Hormigon elab. (4,00hh)', 'Hormigón elaborado colado', 1.0, 190000.0, 1),
                ('Armadura colocada (0,10hh)', 'Hierro redondo d=10mm', 1.05, 2553.76, 1),
                ('Armadura colocada (0,10hh)', 'Alambre negro', 0.01, 7000.0, 1),
                ('H.Elab. pavimento 15cm', 'Hormigón elaborado colado', 0.15, 215000.0, 1),
                ('H.Elab. pavimento 15cm', 'Armadura colocada', 2.5, 3501.45, 1),
                ('H.Elab. pavimento 15cm', 'Junta (equivalente de encofrado)', 0.3, 25781.27, 1),
                ('H.Elab. base (50-3)', 'Hormigón elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. base (50-3)', 'Armadura colocada', 50.0, 3501.45, 1),
                ('H.Elab. base (50-3)', 'Encofrado', 3.0, 25781.27, 1),
                ('H.Elab.zapata (30-6)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab.zapata (30-6)', 'Armadura colocada', 30.0, 3501.45, 1),
                ('H.Elab.zapata (30-6)', 'Encofrado', 6.0, 25781.27, 1),
                ('H.Elab. zapata (60-6)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. zapata (60-6)', 'Armadura colocada', 60.0, 3501.45, 1),
                ('H.Elab. zapata (60-6)', 'Encofrado', 6.0, 25781.27, 1),
                ('H.Elab. columna (105-11)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. columna (105-11)', 'Armadura colocada', 105.0, 3501.45, 1),
                ('H.Elab. columna (105-11)', 'Encofrado', 11.0, 25781.27, 1),
                ('H.Elab. viga fundacion (75-7)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. viga fundacion (75-7)', 'Armadura colocada', 75.0, 3501.45, 1),
                ('H.Elab. viga fundacion (75-7)', 'Encofrado', 7.0, 25781.27, 1),
                ('H.Elab. viga (120-16)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. viga (120-16)', 'Armadura colocada', 120.0, 3501.45, 1),
                ('H.Elab. viga (120-16)', 'Encofrado', 16.0, 25781.27, 1),
                ('H.Elab. losa 10cm (80-10)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. losa 10cm (80-10)', 'Armadura colocada', 80.0, 3501.45, 1),
                ('H.Elab. losa 10cm (80-10)', 'Encofrado', 10.0, 25781.27, 1),
                ('H.Elab. losa 15cm (75-6.66)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. losa 15cm (75-6.66)', 'Armadura colocada', 75.0, 3501.45, 1),
                ('H.Elab. losa 15cm (75-6.66)', 'Encofrado', 6.66, 25781.27, 1),
                ('H.Elab. losa premoldeada', 'Hormigón colado', 0.07, 215000.0, 1),
                ('H.Elab. losa premoldeada', 'Armadura colocada', 1.0, 3501.45, 1),
                ('H.Elab. losa premoldeada', 'Ladrillo Telgopor', 2.0, 6200.0, 1),
                ('H.Elab. losa premoldeada', 'Viga Vipret 4m.', 2.0, 3862.5, 1),
                ('H.Elab. tabique 10cm (110-20)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. tabique 10cm (110-20)', 'Armadura colocada', 110.0, 3501.45, 1),
                ('H.Elab. tabique 10cm (110-20)', 'Encofrado', 20.0, 25781.27, 1),
                ('H.Elab. tabique 15cm (110-13)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. tabique 15cm (110-13)', 'Armadura colocada', 110.0, 3501.45, 1),
                ('H.Elab. tabique 15cm (110-13)', 'Encofrado', 13.33, 25781.27, 1),
                ('H.Elab. tanque (90-13)', 'Hormigon elaborado colado', 1.0, 215000.0, 1),
                ('H.Elab. tanque (90-13)', 'Armadura colocada', 90.0, 3501.45, 1),
                ('H.Elab. tanque (90-13)', 'Encofrado', 13.33, 25781.27, 1),
                ('Retiro de tierra', 'Transporte material suelto', 1.0, 19000.0, 1),
                ('Mamp. ladrillo comun cimientos', 'Ladrillos comunes', 400.0, 230.0, 1),
                ('Mamp. ladrillo comun cimientos', 'Cemento Albañilería', 67.0, 296.0, 1),
                ('Mamp. ladrillo comun cimientos', 'Arena común', 0.5, 31000.0, 1),
                ('Mamp. ladrillo comun submuracion', 'Ladrillos comunes', 400.0, 230.0, 1),
                ('Mamp. ladrillo comun submuracion', 'Cemento Albañilería', 75.0, 296.0, 1),
                ('Mamp. ladrillo comun submuracion', 'Arena común', 0.5, 31000.0, 1),
                ('Mamp. ladrillo comun 15cm', 'Ladrillos comunes', 400.0, 230.0, 1),
                ('Mamp. ladrillo comun 15cm', 'Cemento Albañilería', 67.0, 296.0, 1),
                ('Mamp. ladrillo comun 15cm', 'Arena común', 0.5, 31000.0, 1),
                ('Mamp. ladrillo comun 30cm', 'Ladrillos comunes', 400.0, 230.0, 1),
                ('Mamp. ladrillo comun 30cm', 'Cemento Albañilería', 67.0, 296.0, 1),
                ('Mamp. ladrillo comun 30cm', 'Arena común', 0.5, 31000.0, 1),
                ('Mamp. ladrillo vista 15cm', 'Ladrillos vista', 430.0, 390.0, 1),
                ('Mamp. ladrillo vista 15cm', 'Cemento Albañilería', 75.0, 296.0, 1),
                ('Mamp. ladrillo vista 15cm', 'Arena común', 0.5, 31000.0, 1),
                ('Mamp. ladrillo vista 30cm', 'Ladrillos vista', 430.0, 390.0, 1),
                ('Mamp. ladrillo vista 30cm', 'Cemento Albañilería', 75.0, 296.0, 1),
                ('Mamp. ladrillo vista 30cm', 'Arena común', 0.5, 31000.0, 1),
                ('Tabique ladrillo comun 8cm', 'Ladrillos comunes', 30.0, 230.0, 1),
                ('Tabique ladrillo comun 8cm', 'Cemento Albañilería', 7.5, 296.0, 1),
                ('Tabique ladrillo comun 8cm', 'Arena común', 0.05, 31000.0, 1),
                ('Tabique ladrillo hueco 8x18x33', 'Ladrillo hueco 8x18x33cm', 17.0, 720.0, 1),
                ('Tabique ladrillo hueco 8x18x33', 'Cemento Albañilería', 7.0, 296.0, 1),
                ('Tabique ladrillo hueco 8x18x33', 'Arena común', 0.05, 31000.0, 1),
                ('Tabique ladrillo hueco 12x18x33', 'Ladrillo hueco 12X18X33cm', 17.0, 830.0, 1),
                ('Tabique ladrillo hueco 12x18x33', 'Cemento Albañilería', 7.0, 296.0, 1),
                ('Tabique ladrillo hueco 12x18x33', 'Arena común', 0.05, 31000.0, 1),
                ('Tabique ladrillo hueco 18x18x33', 'Ladrillo hueco 18X18X33cm', 17.0, 1160.0, 1),
                ('Tabique ladrillo hueco 18x18x33', 'Cemento Albañilería', 7.0, 296.0, 1),
                ('Tabique ladrillo hueco 18x18x33', 'Arena común', 0.05, 31000.0, 1),
                ('Tabique hueco portante 12x19x33cm', 'Ladrillo hueco Portante 12x18x33cm', 17.0, 1160.0, 1),
                ('Tabique hueco portante 12x19x33cm', 'Cemento Albañilería', 7.0, 296.0, 1),
                ('Tabique hueco portante 12x19x33cm', 'Arena común', 0.05, 31000.0, 1),
                ('Tabique hueco portante 12x19x33cm', 'Hierro redondo mampostería', 2.0, 916.67, 1),
                ('Tabique hueco portante 18x19x33cm', 'Ladrillo hueco Portante 18x18x33cm', 17.0, 1400.0, 1),
                ('Tabique hueco portante 18x19x33cm', 'Cemento Albañilería', 7.5, 296.0, 1),
                ('Tabique hueco portante 18x19x33cm', 'Arena común', 0.05, 31000.0, 1),
                ('Tabique hueco portante 18x19x33cm', 'Hierro redondo mampostería', 3.0, 916.67, 1),
                ('Dinteles pared 15cm', 'Ladrillos comunes', 10.0, 230.0, 1),
                ('Dinteles pared 15cm', 'Cemento portland bolsas', 2.35, 332.0, 1),
                ('Dinteles pared 15cm', 'Arena común', 0.02, 31000.0, 1),
                ('Dinteles pared 15cm', 'Hierro redondo mampostería', 3.0, 916.67, 1),
                ('Dinteles pared 30cm', 'Ladrillos comunes', 20.0, 230.0, 1),
                ('Dinteles pared 30cm', 'Cemento portland bolsas', 4.7, 332.0, 1),
                ('Dinteles pared 30cm', 'Arena común', 0.04, 31000.0, 1),
                ('Dinteles pared 30cm', 'Hierro redondo mampostería', 5.0, 916.67, 1),
                ('Cemento: azotado', 'Cemento portland bolsas', 2.7, 332.0, 1),
                ('Cemento: azotado', 'Arena común', 0.006, 31000.0, 1),
                ('Cemento: azotado', 'Hidrófugo', 0.13, 2900.0, 1),
                ('Cemento: capa aisladora s/muros', 'Cemento portland bolsas', 11.0, 332.0, 1),
                ('Cemento: capa aisladora s/muros', 'Arena común', 0.025, 31000.0, 1),
                ('Cemento: capa aisladora s/muros', 'Hidrófugo', 0.25, 2900.0, 1),
                ('Cemento: piso rodillado', 'Cemento portland bolsas', 15.0, 332.0, 1),
                ('Cemento: piso rodillado', 'Arena común', 0.028, 31000.0, 1),
                ('Cemento: piso alisado', 'Cemento portland bolsas', 17.0, 332.0, 1),
                ('Cemento: piso alisado', 'Arena común', 0.028, 31000.0, 1),
                ('Cemento: carpeta azotea', 'Cemento portland bolsas', 13.5, 332.0, 1),
                ('Cemento: carpeta azotea', 'Arena común', 0.025, 31000.0, 1),
                ('Cemento: toma de juntas', 'Cemento portland bolsas', 1.5, 332.0, 1),
                ('Cemento: toma de juntas', 'Arena común', 0.005, 31000.0, 1),
                ('Cemento: revoque', 'Cemento portland bolsas', 21.8, 332.0, 1),
                ('Cemento: revoque', 'Arena común', 0.03, 31000.0, 1),
                ('Cemento: revoque', 'Hidrófugo', 0.75, 2900.0, 1),
                ('Revoque tanques', 'Cemento portland bolsas', 21.8, 332.0, 1),
                ('Revoque tanques', 'Arena común', 0.03, 31000.0, 1),
                ('Revoque tanques', 'Hidrófugo', 0.75, 2900.0, 1),
                ('Cemento: zocalo 10cm', 'Cemento portland bolsas', 0.8, 332.0, 1),
                ('Cemento: zocalo 10cm', 'Arena común', 0.01, 31000.0, 1),
                ('Enlucido S.Iggam', 'Super Iggam', 12.0, 833.33, 1),
                ('Enlucido Salpicrete', 'Salpicrete', 6.0, 2666.67, 1),
                ('Revest. Texturado', 'Fondo Base', 0.25, 3800.0, 1),
                ('Revest. Texturado', 'DeckAr', 3.0, 4166.67, 1),
                ('R. exterior cal (az+gr+f)', 'Cemento portland bolsas', 4.25, 332.0, 1),
                ('R. exterior cal (az+gr+f)', 'Cemento Albañilería', 4.71, 296.0, 1),
                ('R. exterior cal (az+gr+f)', 'Hidrófugo', 0.19, 2900.0, 1),
                ('R. exterior cal (az+gr+f)', 'Arena común', 0.04, 31000.0, 1),
                ('R. interior cal (gr+f) fratas', 'Cemento portland bolsas', 0.85, 332.0, 1),
                ('R. interior cal (gr+f) fratas', 'Cemento Albañilería', 4.71, 296.0, 1),
                ('R. interior cal (gr+f) fratas', 'Arena común', 0.025, 31000.0, 1),
                ('R. interior cal (gr+f) fieltro', 'Cemento portland bolsas', 0.85, 332.0, 1),
                ('R. interior cal (gr+f) fieltro', 'Cemento Albañilería', 4.71, 296.0, 1),
                ('R. interior cal (gr+f) fieltro', 'Arena común', 0.025, 31000.0, 1),
                ('R. grueso interior b/rvto.', 'Cemento portland bolsas', 0.7, 332.0, 1),
                ('R. grueso interior b/rvto.', 'Cemento Albañilería', 2.1, 296.0, 1),
                ('R. grueso interior b/rvto.', 'Arena común', 0.015, 31000.0, 1),
                ('Colocacion de Aberturas', 'Cemento portland bolsas', 32.7, 332.0, 1),
                ('Colocacion de Aberturas', 'Arena común', 0.045, 31000.0, 1),
                ('Instalacion Desague', 'Caño Awaduct 110', 11.2, 8125.0, 1),
                ('Instalacion Desague', 'Caño Awaduct 63', 2.4, 6250.0, 1),
                ('Instalacion Desague', 'Caño Awaduct 50', 1.0, 5300.0, 1),
                ('Instalacion Desague', 'Caño Awaduct 40', 3.2, 2900.0, 1),
                ('Instalacion Desague', 'Accesorios Desagues', 10.0, 8500.0, 1),
                ('Instalacion Desague', 'Cemento portland bolsas', 4.0, 332.0, 1),
                ('Instalacion Desague', 'Arena común', 0.02, 31000.0, 1),
                ('Instalacion Electrica', 'Caño Corrugado 1"', 1.0, 244.4, 1),
                ('Instalacion Electrica', 'Caño Corrugado 3/4"', 2.0, 169.0, 1),
                ('Instalacion Electrica', 'Cajas Metalicas', 1.0, 850.0, 1),
                ('Instalacion Electrica', 'Cable 2,5 mm', 2.5, 850.0, 1),
                ('Instalacion Electrica', 'Cable 1,5 mm', 5.5, 510.0, 1),
                ('Instalacion Electrica', 'Cemento portland bolsas', 2.0, 332.0, 1),
                ('Instalacion Electrica', 'Arena común', 0.01, 31000.0, 1),
                ('Instalacion Agua F/C', 'Caño TF 25', 8.0, 3250.0, 1),
                ('Instalacion Agua F/C', 'Caño TF 20', 14.0, 1706.25, 1),
                ('Instalacion Agua F/C', 'Accesorios TF', 27.0, 3000.0, 1),
                ('Instalacion Agua F/C', 'Llaves de Paso', 3.0, 25000.0, 1),
                ('Instalacion Agua F/C', 'Cemento portland bolsas', 2.0, 332.0, 1),
                ('Instalacion Agua F/C', 'Arena común', 0.0264, 31000.0, 1),
                ('Instalacion Gas', 'Caño Epoxi 3/4', 8.0, 12890.62, 1),
                ('Instalacion Gas', 'Caño epoxi 1/2', 2.2, 6406.25, 1),
                ('Instalacion Gas', 'Accesorios Gas', 12.0, 3500.0, 1),
                ('Instalacion Gas', 'Llaves de Paso', 1.0, 25000.0, 1),
                ('Instalacion Gas', 'Cemento portland bolsas', 2.0, 332.0, 1),
                ('Instalacion Gas', 'Arena común', 0.03, 31000.0, 1),
                ('Cielorraso s/losa (gr+f)', 'Cemento portland bolsas', 0.5, 332.0, 1),
                ('Cielorraso s/losa (gr+f)', 'Cemento Albañilería', 4.71, 296.0, 1),
                ('Cielorraso s/losa (gr+f)', 'Arena común', 0.025, 31000.0, 1),
                ('Cielo s/metal despl. (az+gr+f)', 'Cemento portland bolsas', 1.5, 332.0, 1),
                ('Cielo s/metal despl. (az+gr+f)', 'Cemento Albañilería', 4.71, 296.0, 1),
                ('Cielo s/metal despl. (az+gr+f)', 'Arena común', 0.03, 31000.0, 1),
                ('Estructura cielo suspendido (3-6m)', 'Metal desplegado', 1.15, 3333.33, 1),
                ('Estructura cielo suspendido (3-6m)', 'Palito 1"x1"', 7.0, 8860.76, 1),
                ('Estructura cielo suspendido (3-6m)', 'Clavos', 0.15, 5600.0, 1),
                ('Contrapiso cascotes 15cm', 'Cemento portland bolsas', 2.0, 332.0, 1),
                ('Contrapiso cascotes 15cm', 'Cemento Albañilería', 20.0, 296.0, 1),
                ('Contrapiso cascotes 15cm', 'Arena común', 0.075, 31000.0, 1),
                ('Contrapiso cascotes 15cm', 'Granza', 0.12, 36000.0, 1),
                ('Contrapiso cascotes 10cm', 'Cemento portland bolsas', 1.5, 332.0, 1),
                ('Contrapiso cascotes 10cm', 'Cemento Albañilería', 14.29, 296.0, 1),
                ('Contrapiso cascotes 10cm', 'Arena común', 0.05, 31000.0, 1),
                ('Contrapiso cascotes 10cm', 'Granza', 0.08, 36000.0, 1),
                ('Banquina Ho Pobre 10cm', 'Cemento portland bolsas', 1.5, 332.0, 1),
                ('Banquina Ho Pobre 10cm', 'Cemento Albañilería', 14.29, 296.0, 1),
                ('Banquina Ho Pobre 10cm', 'Arena común', 0.05, 31000.0, 1),
                ('Banquina Ho Pobre 10cm', 'Granza', 0.08, 36000.0, 1),
                ('Ho de Pendiente c/perlitas 10cm', 'Perlitas Telgopor', 106.0, 106.67, 1),
                ('Ho de Pendiente c/perlitas 10cm', 'Cemento portland bolsas', 10.4, 332.0, 1),
                ('Ho de Pendiente c/perlitas 10cm', 'Arena común', 0.014, 31000.0, 1),
                ('Ho de Pendiente c/perlitas 10cm', 'Granza', 0.014, 36000.0, 1),
                ('Contrapiso c/perlitas 10cm', 'Perlitas Telgopor', 106.0, 106.67, 1),
                ('Contrapiso c/perlitas 10cm', 'Cemento portland bolsas', 10.5, 332.0, 1),
                ('Contrapiso c/perlitas 10cm', 'Arena común', 0.02, 31000.0, 1),
                ('Contrapiso c/perlitas 10cm', 'Granza', 0.02, 36000.0, 1),
                ('Contrapiso c/perlitas 5cm', 'Perlitas Telgopor', 70.0, 106.67, 1),
                ('Contrapiso c/perlitas 5cm', 'Cemento portland bolsas', 6.5, 332.0, 1),
                ('Contrapiso c/perlitas 5cm', 'Arena común', 0.014, 31000.0, 1),
                ('Contrapiso c/perlitas 5cm', 'Granza', 0.014, 36000.0, 1),
                ('Carpeta bajo piso', 'Cemento portland bolsas', 13.5, 332.0, 1),
                ('Carpeta bajo piso', 'Arena común', 0.025, 31000.0, 1),
                ('Piso ceramico azotea', 'Baldosa cerámica azotea', 1.05, 33000.0, 1),
                ('Piso ceramico azotea', 'Cemento portland bolsas', 5.0, 332.0, 1),
                ('Piso ceramico azotea', 'Cal aérea Milagro', 4.5, 372.0, 1),
                ('Piso ceramico azotea', 'Arena común', 0.03, 31000.0, 1),
                ('Piso ceramico 1', 'Piso cerámico 1', 1.05, 15000.0, 1),
                ('Piso ceramico 1', 'Klaukol', 4.0, 600.0, 1),
                ('Piso ceramico 1', 'Pastina', 0.2, 5500.0, 1),
                ('Piso ceramico 2', 'Piso cerámico 2', 1.05, 25000.0, 1),
                ('Piso ceramico 2', 'Klaukol', 4.0, 600.0, 1),
                ('Piso ceramico 2', 'Pastina', 0.2, 5500.0, 1),
                ('Piso ceramico 3 (porcellanato)', 'Piso cerámico 3 (porcellanato)', 1.05, 35000.0, 1),
                ('Piso ceramico 3 (porcellanato)', 'Klaukol', 4.0, 600.0, 1),
                ('Piso ceramico 3 (porcellanato)', 'Pastina', 0.2, 5500.0, 1),
                ('Piso loseta cemento', 'Loseta cemento 60x40cm', 1.05, 22000.0, 1),
                ('Piso loseta cemento', 'Cemento portland bolsas', 5.0, 332.0, 1),
                ('Piso loseta cemento', 'Cal aérea Milagro', 4.5, 372.0, 1),
                ('Piso loseta cemento', 'Arena común', 0.03, 31000.0, 1),
                ('Piso calcareo vereda', 'Mosaico calcáreo', 1.1, 25000.0, 1),
                ('Piso calcareo vereda', 'Cemento portland bolsas', 5.0, 332.0, 1),
                ('Piso calcareo vereda', 'Cal aérea Milagro', 4.5, 372.0, 1),
                ('Piso calcareo vereda', 'Arena común', 0.03, 31000.0, 1),
                ('Zocalo ceramico 1', 'Zócalo cerámico 1', 1.05, 1250.0, 1),
                ('Zocalo ceramico 1', 'Klaukol', 0.2, 600.0, 1),
                ('Zocalo ceramico 1', 'Pastina', 0.02, 5500.0, 1),
                ('Zocalo ceramico 2', 'Zócalo cerámico 2', 1.05, 2083.33, 1),
                ('Zocalo ceramico 2', 'Klaukol', 0.2, 600.0, 1),
                ('Zocalo ceramico 2', 'Pastina', 0.02, 5500.0, 1),
                ('Zocalo ceramico 3', 'Zócalo cerámico 3 (Porcellanato)', 1.05, 2916.67, 1),
                ('Zocalo ceramico 3', 'Pastina', 0.02, 5500.0, 1),
                ('Zocalo ceramico 3', 'Klaukol', 0.2, 600.0, 1),
                ('Zocalo de pino', 'Zócalo de madera', 1.05, 3500.0, 1),
                ('Zocalo de pino', 'Tarugo 6', 1.05, 62.0, 1),
                ('Zocalo de pino', 'Tornillo', 1.05, 45.0, 1),
                ('Rvto. marmol', 'Cemento portland bolsas', 8.0, 332.0, 1),
                ('Rvto. marmol', 'Arena común', 0.03, 31000.0, 1),
                ('Rvto. marmol', 'Mármol revestimientos', 1.0, 200000.0, 1),
                ('Rvto. ceramico 1', 'Rvto.cerámico 1', 1.1, 15000.0, 1),
                ('Rvto. ceramico 1', 'Klaukol', 2.0, 600.0, 1),
                ('Rvto. ceramico 1', 'Pastina', 0.2, 5500.0, 1),
                ('Rvto. ceramico 2', 'Rvto.cerámico 2', 1.1, 25000.0, 1),
                ('Rvto. ceramico 2', 'Klaukol', 2.0, 600.0, 1),
                ('Rvto. ceramico 2', 'Pastina', 0.2, 5500.0, 1),
                ('Rvto. ceramico 3 porcellanato', 'Rvto.cerámico 3 (porcellanato)', 1.1, 35000.0, 1),
                ('Rvto. ceramico 3 porcellanato', 'Klaukol', 2.0, 600.0, 1),
                ('Rvto. ceramico 3 porcellanato', 'Pastina', 0.2, 5500.0, 1),
                ('Techo de chapas', 'Tirantes 2x6', 1.41, 27166.67, 1),
                ('Techo de chapas', 'Pino tabla machimbre', 1.1, 11600.0, 1),
                ('Techo de chapas', 'Escurridores 1/2 x 2', 2.82, 1700.0, 1),
                ('Techo de chapas', 'Issolant', 1.1, 4950.0, 1),
                ('Techo de chapas', 'Clavadores 2 x 2', 1.41, 1007.59, 1),
                ('Techo de chapas', 'Chapas Techo', 1.05, 14800.0, 1),
                ('Techo de chapas', 'Clavos 2" 1/2', 0.05, 5600.0, 1),
                ('Techo de chapas', 'Clavos 3"', 0.05, 7200.0, 1),
                ('Techo de chapas', 'Clavos 4"', 0.05, 7200.0, 1),
                ('Techo de chapas', 'Tornillo c/arand goma', 6.0, 118.0, 1),
                ('Demolicion mamposteria 0,15', 'Transporte material suelto', 2.0, 19000.0, 1),
                ('Demolicion mamposteria 0,15', 'Martillo neumático', 4.0, 25000.0, 1),
                ('Zocalo Porcellanato', 'Zócalo cerámico 3 (Porcellanato)', 1.05, 2916.67, 1),
                ('Zocalo Porcellanato', 'Klaukol', 0.2, 720.0, 1),
                ('Zocalo Porcellanato', 'Pastina', 0.02, 5500.0, 1),
                ('Rvto. Ceramico', 'Rvto.cerámico 2', 1.1, 25000.0, 1),
                ('Rvto. Ceramico', 'Klaukol', 2.0, 720.0, 1),
                ('Rvto. Ceramico', 'Pastina', 0.2, 5500.0, 1),
                ('Relleno y Compactacion C/15cm', 'Máquina excavadora', 0.12, 25000.0, 1),
                ('Relleno y Compactacion C/15cm', 'Tierra Colorada', 1.5, 20000.0, 1),
                ('Piso Porcellanato', 'Piso cerámico 3 (porcellanato)', 1.05, 35000.0, 1),
                ('Piso Porcellanato', 'Klaukol', 4.0, 720.0, 1),
                ('Piso Porcellanato', 'Pastina', 0.2, 5500.0, 1),
                ('Piso Ceramico', 'Piso cerámico 2', 1.05, 25000.0, 1),
                ('Piso Ceramico', 'Klaukol', 4.0, 720.0, 1),
                ('Piso Ceramico', 'Pastina', 0.2, 5500.0, 1),
                ('Demolicion mamposteria 0,30', 'Transporte material suelto', 2.0, 19000.0, 1),
                ('Demolicion mamposteria 0,30', 'Martillo neumático', 4.0, 25000.0, 1),
                ('Zocalo Ceramico', 'Zócalo cerámico 2', 1.05, 2083.33, 1),
                ('Zocalo Ceramico', 'Klaukol', 0.2, 720.0, 1),
                ('Zocalo Ceramico', 'Pastina', 0.02, 5500.0, 1),
                ('Rvto. Porcellanato', 'Rvto.cerámico 3 (porcellanato)', 1.1, 35000.0, 1),
                ('Rvto. Porcellanato', 'Klaukol', 2.0, 720.0, 1),
                ('Rvto. Porcellanato', 'Pastina', 0.2, 5500.0, 1),
            ]
            nombres_2n = sorted(set(t[0] for t in SUBS_2N))
            for item_nombre in nombres_2n:
                db.execute("DELETE FROM analisis_sub WHERE item_nombre=? AND es_material=1", (item_nombre,))
            db.commit()
            db.executemany(
                "INSERT INTO analisis_sub (item_nombre, sub_nombre, cant_por_unit, precio_ars, es_material) "
                "VALUES (?,?,?,?,?)",
                SUBS_2N
            )
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2n_done','2026-07-04')")
            db.commit()
            print(f"[migrate_db] 2n: {len(SUBS_2N)} materiales resincronizados contra PRESUPUESTO COCHERA.xlsx")

        # ── 2o. Corrección de ítems renombrados (_renombres) que 2n no alcanzó ──
        #   database.py renombra items_obra.nombre para varios ítems (ver lista
        #   `_renombres` más arriba: Relleno y Compactacion → "...C/15cm", Piso/
        #   Zocalo/Rvto ceramico "2"/"3" → "Piso/Zocalo/Rvto Ceramico/Porcellanato").
        #   La migración 2n comparó contra el Excel usando el nombre ORIGINAL
        #   (pre-renombre), así que para estos ítems quedaron sin corregir bajo su
        #   nombre real post-renombre. Encontrado por el usuario en "Relleno y
        #   Compactación": la app traía "Máquina excavadora" (que el Excel no
        #   tiene para este ítem) y menos Tierra Colorada de la que corresponde.
        ya_2o = db.execute("SELECT valor FROM config WHERE clave='2o_done'").fetchone()
        if not ya_2o:
            SUBS_2O = [
                # Relleno y Compactación: SIN máquina excavadora, Tierra Colorada correcta
                ('Relleno y Compactacion C/15cm', 'Tierra Colorada', 1.58, 20000.0, 1),
                # Klaukol estaba a precio viejo (720) en vez del real (600) en estos 6
                ('Piso Ceramico', 'Piso cerámico 2', 1.05, 25000.0, 1),
                ('Piso Ceramico', 'Klaukol', 4.0, 600.0, 1),
                ('Piso Ceramico', 'Pastina', 0.2, 5500.0, 1),
                ('Piso Porcellanato', 'Piso cerámico 3 (porcellanato)', 1.05, 35000.0, 1),
                ('Piso Porcellanato', 'Klaukol', 4.0, 600.0, 1),
                ('Piso Porcellanato', 'Pastina', 0.2, 5500.0, 1),
                ('Zocalo Ceramico', 'Zócalo cerámico 2', 1.05, 2083.33, 1),
                ('Zocalo Ceramico', 'Klaukol', 0.2, 600.0, 1),
                ('Zocalo Ceramico', 'Pastina', 0.02, 5500.0, 1),
                ('Zocalo Porcellanato', 'Zócalo cerámico 3 (Porcellanato)', 1.05, 2916.67, 1),
                ('Zocalo Porcellanato', 'Pastina', 0.02, 5500.0, 1),
                ('Zocalo Porcellanato', 'Klaukol', 0.2, 600.0, 1),
                ('Rvto. Ceramico', 'Rvto.cerámico 2', 1.1, 25000.0, 1),
                ('Rvto. Ceramico', 'Klaukol', 2.0, 600.0, 1),
                ('Rvto. Ceramico', 'Pastina', 0.2, 5500.0, 1),
                ('Rvto. Porcellanato', 'Rvto.cerámico 3 (porcellanato)', 1.1, 35000.0, 1),
                ('Rvto. Porcellanato', 'Klaukol', 2.0, 600.0, 1),
                ('Rvto. Porcellanato', 'Pastina', 0.2, 5500.0, 1),
            ]
            nombres_2o = sorted(set(t[0] for t in SUBS_2O))
            for item_nombre in nombres_2o:
                db.execute("DELETE FROM analisis_sub WHERE item_nombre=? AND es_material=1", (item_nombre,))
            db.commit()
            db.executemany(
                "INSERT INTO analisis_sub (item_nombre, sub_nombre, cant_por_unit, precio_ars, es_material) "
                "VALUES (?,?,?,?,?)",
                SUBS_2O
            )
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2o_done','2026-07-04')")
            db.commit()
            print(f"[migrate_db] 2o: {len(SUBS_2O)} materiales corregidos en items renombrados (Relleno y Compactacion, Piso/Zocalo/Rvto Ceramico/Porcellanato)")

        # ── 2p. Reaplicar conversión a unidad comercial post 2n/2o ──────────────
        #   2n y 2o insertaron los datos del Excel TAL CUAL (kg/L crudos), pero
        #   2j/2k/2l ya habían convertido 6 materiales a unidad comercial (bolsa)
        #   antes de que 2n/2o corrieran. Al hacer DELETE+INSERT, 2n/2o pisaron
        #   esa conversión y volvieron a dejar estos 6 materiales en kg crudo en
        #   TODOS los ítems que los usan — mismo bug de doble-conversión que ya
        #   se había corregido en routes/presupuesto.py y admin.py, pero ahora
        #   del lado de los DATOS en vez del código. Factores finales (los mismos
        #   que dejó 2l): Cemento portland/Albañilería/Cal aérea/Klaukol = 25,
        #   Salpicrete/Super Iggam = 30.
        ya_2p = db.execute("SELECT valor FROM config WHERE clave='2p_done'").fetchone()
        if not ya_2p:
            FACTORES_2P = [
                ('Cemento portland bolsas', 25),
                ('Cemento Albañilería',     25),
                ('Cal aérea Milagro',       25),
                ('Klaukol',                 25),
                ('Salpicrete',              30),
                ('Super Iggam',             30),
            ]
            for sub, factor in FACTORES_2P:
                db.execute("""
                    UPDATE analisis_sub
                    SET cant_por_unit = ROUND(cant_por_unit / ?, 6),
                        precio_ars    = ROUND(precio_ars    * ?, 2)
                    WHERE sub_nombre = ? AND es_material = 1
                """, (factor, factor, sub))
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2p_done','2026-07-04')")
            db.commit()
            print("[migrate_db] 2p: reconversion a unidad comercial post 2n/2o (6 materiales)")

        # ── 2q. Actualización de precios reales — LISTA_MATERIALES_V3_formulafix.xlsx ──
        #   Fuente: D:\ESCRITORIO\CLAUDE\02_CONSTRUCCION\PRESUPUESTOS\LISTA_MATERIALES_V3_formulafix.xlsx
        #   (Corralón El Cruce · Lista 169 · 04/06/2026). Para los 6 materiales que la app ya
        #   almacena en unidad comercial (bolsa, post 2j/2k/2l/2p) se usa la columna F de la V3
        #   ($ Precio comercial, precio de la bolsa entera). Para el resto (que sigue en su unidad
        #   de cálculo cruda: kg/m3/u/ml/lt/m2) se usa la columna G ($ por unidad de cálculo).
        #   88 materiales verificados 1 a 1 contra la V3 (incluye Baldosa cerámica azotea ← V3
        #   "Baldosa cerámica Alberdi", Zócalo de madera ← V3 "Zócalo de pino", Viga Vipret 4m. ←
        #   V3 "Vigueta 4,00 m", y Ladrillo hueco 18X18X33cm ← V3 "18x18x33 cm" (celda B22, corregida
        #   por Daniel — era un typo "25cm" en la V3), todos confirmados por Daniel el 04/07/2026).
        #   Queda sin actualizar (a propósito):
        #   - 'Hormigón colado' / 'Hormigon elaborado colado' / 'Hormigón elaborado colado':
        #     a propósito NO se actualizan acá. Son 2 materiales distintos usados como insumo
        #     interno en las recetas de analisis_sub (Hormigón colado = mezclado en obra, usado
        #     en ítems "Ho.Ado.*" a $233.860/m3; Hormigón elaborado colado = camión mixer, usado
        #     en ítems "H.Elab.*" a $215.000/m3 — 'Hormigon elaborado colado' sin tilde es el
        #     mismo material con inconsistencia de tilde). Decisión de Daniel (04/07/2026): estos
        #     cálculos internos se mantienen tal como los trae el Excel (COCHERA, migración 2n).
        #     El valor de la V3 ($190.000/m3, "Hormigón elaborado pto.obra") se usa solo como
        #     referencia de "costo de hormigón puesto en obra" en el Excel de precios comerciales,
        #     no reemplaza estos 2 insumos internos.
        #   NOTA: existe también _migrar_precios_v1() (llamada desde init_db, con estos mismos
        #   precios pero en base kg cruda para TODOS los materiales incl. los 6 de bolsa) que
        #   nunca corrió sobre una base ya existente porque solo se invoca desde init_db(), no
        #   desde migrate_db(). Esta migración 2q reemplaza esa función para las bases ya creadas.
        ya_2q = db.execute("SELECT valor FROM config WHERE clave='2q_done'").fetchone()
        if not ya_2q:
            PRECIOS_2Q = [
                ('Accesorios Desagues', 8500),
                ('Accesorios Gas', 3500),
                ('Accesorios TF', 3000),
                ('Alambre negro', 7000),
                ('Arena común', 31000),
                ('Baldosa cerámica azotea', 33000),
                ('Cable 1,5 mm', 510),
                ('Cable 2,5 mm', 850),
                ('Cajas Metalicas', 850),
                ('Cal Hidráulica', 332),
                ('Cal aérea Milagro', 9300),
                ('Caño Awaduct 110', 8125),
                ('Caño Awaduct 40', 2900),
                ('Caño Awaduct 50', 5300),
                ('Caño Awaduct 63', 6250),
                ('Caño Corrugado 1"', 376),
                ('Caño Corrugado 3/4"', 260),
                ('Caño Epoxi 3/4', 12890.63),
                ('Caño TF 20', 2625),
                ('Caño TF 25', 5000),
                ('Caño epoxi 1/2', 6406.25),
                ('Cemento Albañilería', 7400),
                ('Cemento portland bolsas', 8300),
                ('Chapas Cerco', 14800),
                ('Chapas Techo', 14800),
                ('Clavadores 2 x 2', 1007.59),
                ('Clavos 2"', 5600),
                ('Clavos 2" 1/2', 5600),
                ('Clavos 3"', 7200),
                ('Clavos 4"', 7200),
                ('Color pintura cal', 20000),
                ('Enduido sintético', 7500),
                ('Escurridores 1/2 x 2', 459.46),
                ('Esmalte albalux', 23175),
                ('Fondo Base', 3800),
                ('Granza', 36000),
                ('Hidrófugo', 2900),
                ('Hierro redondo d=10mm', 2553.76),
                ('Issolant', 4950),
                ('Klaukol', 18000),
                ('Ladrillo Telgopor 12*38*1m', 6200),
                ('Ladrillo hueco 12X18X33cm', 830),
                ('Ladrillo hueco 8x18x33cm', 720),
                ('Ladrillo hueco Portante 12x18x33cm', 1160),
                ('Ladrillo hueco Portante 18x18x33cm', 1400),
                ('Ladrillo hueco 18X18X33cm', 1160),
                ('Ladrillos comunes', 230),
                ('Ladrillos vista', 390),
                ('Llaves de Paso Agua', 25000),
                ('Llaves de Paso Gas', 25000),
                ('Loseta cemento 60x40cm', 5000),
                ('Martillo neumático', 25000),
                ('Metal desplegado', 3333.33),
                ('Mosaico calcáreo', 25000),
                ('Palito 1"x1"', 500),
                ('Pastina', 5500),
                ('Perlitas Telgopor', 106.67),
                ('Piedra Granítica', 93000),
                ('Pino encofrado 1"', 8860.76),
                ('Pino tabla machimbre', 11600),
                ('Pintura especial 1', 21250),
                ('Pintura especial 2', 22500),
                ('Pintura látex cielos', 7200),
                ('Pintura látex exterior', 7250),
                ('Pintura látex interior', 4750),
                ('Pintura satinol', 22500),
                ('Piso cerámico 1', 15000),
                ('Piso cerámico 2', 25000),
                ('Piso cerámico 3 (porcellanato)', 35000),
                ('Rvto.cerámico 1', 15000),
                ('Rvto.cerámico 2', 25000),
                ('Rvto.cerámico 3 (porcellanato)', 35000),
                ('Saligna   1"x2"', 1037.97),
                ('Saligna 1"x4"', 886.08),
                ('Saligna 3"x3"', 1766.67),
                ('Salpicrete', 80000),
                ('Super Iggam', 25000),
                ('Tarugo 6', 62),
                ('Tierra Colorada', 20000),
                ('Tirantes 2x6', 27166.67),
                ('Tornillo', 45),
                ('Tornillo c/arand goma', 118),
                ('Transporte material suelto', 19000),
                ('Viga Vipret 4m.', 3862.5),
                ('Zócalo cerámico 1', 1250),
                ('Zócalo cerámico 2', 2083.33),
                ('Zócalo cerámico 3 (Porcellanato)', 2916.67),
                ('Zócalo de madera', 3500),
            ]
            for sub, precio in PRECIOS_2Q:
                db.execute(
                    "UPDATE analisis_sub SET precio_ars=? WHERE sub_nombre=? AND es_material=1",
                    (precio, sub)
                )
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2q_done','2026-07-04')")
            db.commit()
            print(f"[migrate_db] 2q: {len(PRECIOS_2Q)} precios actualizados desde LISTA_MATERIALES_V3_formulafix.xlsx")

        # ── 2r. BUG: precio_mo_ars de las 4 instalaciones quedó desactualizado ──
        #   tras 2m. _actualizar_mo_analisis() escribe items_obra.precio_mo_ars UNA
        #   sola vez (guard: "WHERE precio_mo_ars IS NULL OR = 0") y nunca se vuelve
        #   a llamar. La migración 2m corrigió hof/hay de las 4 instalaciones pero NO
        #   recalculó precio_mo_ars — a diferencia de 2h, que sí lo hacía (ver arriba:
        #   "UPDATE items_obra SET precio_mo_ars = ROUND(hof*10000 + hay*5000, 2)").
        #   Resultado: el total de MO del presupuesto (paso2: precio_mo_ars × cantidad)
        #   usaba HOF/HAY viejos y "cruzados" para estas 4 instalaciones, aunque
        #   items_obra.hof/hay ya estuviera corregido desde 2m. Detectado 04/07/2026
        #   por Daniel comparando MO real (Excel COCHERA: $5.902.594) vs MO de la app
        #   ($8.736.994) para el presupuesto de Ezequiel Petrini.
        ya_2r = db.execute("SELECT valor FROM config WHERE clave='2r_done'").fetchone()
        if not ya_2r:
            db.execute("""
                UPDATE items_obra
                SET precio_mo_ars = ROUND(hof * 10000 + hay * 5000, 2)
                WHERE nombre IN ('Instalacion Desague','Instalacion Agua F/C',
                                 'Instalacion Gas','Instalacion Electrica')
            """)
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2r_done','2026-07-04')")
            db.commit()
            print("[migrate_db] 2r: precio_mo_ars recalculado para las 4 instalaciones (post 2m)")

        # ── 2s. "Accesorios" genérico → nombre específico (Desagues/TF/Gas) ─────
        #   Bug reportado por Daniel 05/07/2026: la app seguía mostrando
        #   "Accesorios" genérico en vez de "Accesorios Desagües" como indica la
        #   lista V3 (LISTA_MATERIALES_V3_formulafix.xlsx). Causa real: la
        #   migración 2n (resincronización contra PRESUPUESTO COCHERA.xlsx)
        #   reinsertó estos 3 materiales con el nombre genérico "Accesorios"
        #   para 'Instalacion Desague'/'Instalacion Agua F/C'/'Instalacion Gas',
        #   perdiendo la distinción que sí tenía analisis_data.py (fuente
        #   canónica: 'Accesorios Desagues', 'Accesorios TF', 'Accesorios Gas'
        #   respectivamente — Agua F/C usa "TF" de Termofusión). Efecto
        #   secundario: la migración 2q (precios reales de la lista V3) buscaba
        #   esos nombres específicos por sub_nombre exacto y nunca los
        #   encontraba, así que tampoco les actualizó el precio real.
        ya_2s = db.execute("SELECT valor FROM config WHERE clave='2s_done'").fetchone()
        if not ya_2s:
            RENOMBRES_2S = [
                # (item_nombre, nombre_correcto (lista V3), precio_real_v3)
                ('Instalacion Desague',  'Accesorios Desagues', 8500.0),
                ('Instalacion Agua F/C', 'Accesorios TF',       3000.0),
                ('Instalacion Gas',      'Accesorios Gas',      3500.0),
            ]
            for item_nombre, nombre_correcto, precio in RENOMBRES_2S:
                db.execute("""
                    UPDATE analisis_sub
                    SET sub_nombre = ?, precio_ars = ?
                    WHERE item_nombre = ? AND sub_nombre = 'Accesorios' AND es_material = 1
                """, (nombre_correcto, precio, item_nombre))
            db.commit()
            db.execute("INSERT OR REPLACE INTO config (clave,valor) VALUES ('2s_done','2026-07-05')")
            db.commit()
            print("[migrate_db] 2s: 'Accesorios' generico renombrado a nombre especifico (Desagues/TF/Gas) segun lista V3")

    except Exception as e:
        print(f"[migrate_db] {e}")
    finally:
        db.close()

def init_db():
    db = get_db()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        nombre TEXT,
        pais TEXT DEFAULT 'AR',
        active INTEGER DEFAULT 1,
        subscription_expires DATE,
        session_token TEXT,
        session_expires DATETIME,
        is_admin INTEGER DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS materiales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        unidad TEXT NOT NULL,
        precio_usd REAL NOT NULL,
        categoria TEXT DEFAULT 'General',
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS items_obra (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rubro_num TEXT NOT NULL,
        rubro_nombre TEXT NOT NULL,
        nombre TEXT NOT NULL,
        unidad TEXT NOT NULL,
        precio_ars REAL NOT NULL,
        hof REAL DEFAULT 0,
        hay REAL DEFAULT 0,
        precio_mo_ars REAL DEFAULT 0,
        orden INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS tipos_cambio (
        pais TEXT PRIMARY KEY,
        moneda TEXT NOT NULL,
        simbolo TEXT NOT NULL,
        tasa REAL NOT NULL,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS config (
        clave TEXT PRIMARY KEY,
        valor TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS analisis_sub (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        item_nombre TEXT NOT NULL,
        sub_nombre TEXT NOT NULL,
        cant_por_unit REAL NOT NULL,
        precio_ars REAL DEFAULT 0,
        es_material INTEGER DEFAULT 1
    );
    CREATE INDEX IF NOT EXISTS idx_analisis_item ON analisis_sub(item_nombre);

    CREATE TABLE IF NOT EXISTS empresa_perfil (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER UNIQUE REFERENCES users(id),
        nombre TEXT DEFAULT '',
        contacto TEXT DEFAULT '',
        telefono TEXT DEFAULT '',
        email TEXT DEFAULT '',
        slogan TEXT DEFAULT '',
        logo_data TEXT DEFAULT '',
        logo_filename TEXT DEFAULT '',
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS presupuestos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER REFERENCES users(id),
        nro TEXT,
        cliente_nombre TEXT,
        cliente_tel TEXT,
        cliente_email TEXT,
        obra_descripcion TEXT,
        obra_direccion TEXT,
        obra_tipo TEXT DEFAULT 'Vivienda nueva',
        fecha_presup TEXT,
        validez INTEGER DEFAULT 15,
        modo TEXT DEFAULT 'mo_mat',
        superficie REAL DEFAULT 0,
        rubros_json TEXT DEFAULT '[]',
        subcontratos_json TEXT DEFAULT '[]',
        indirectos_json TEXT DEFAULT '{}',
        materiales_json TEXT DEFAULT '[]',
        hh_total REAL DEFAULT 0,
        n_oficiales INTEGER DEFAULT 2,
        n_ayudantes INTEGER DEFAULT 1,
        jornal_oficial REAL DEFAULT 0,
        jornal_ayudante REAL DEFAULT 0,
        dias_obra INTEGER DEFAULT 0,
        pct_gg REAL DEFAULT 20,
        pct_impuestos REAL DEFAULT 7,
        pct_anticipo REAL DEFAULT 30,
        pct_final REAL DEFAULT 20,
        frecuencia_pago TEXT DEFAULT 'mensual',
        tipo_cambio REAL DEFAULT 1,
        costo_directo REAL DEFAULT 0,
        total_subcontratos REAL DEFAULT 0,
        total_indirectos REAL DEFAULT 0,
        total_mo REAL DEFAULT 0,
        total_materiales REAL DEFAULT 0,
        total_presupuesto REAL DEFAULT 0,
        descripcion_trabajos TEXT DEFAULT '',
        status TEXT DEFAULT 'completo',
        wizard_step INTEGER DEFAULT 8,
        session_json TEXT DEFAULT '{}',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS leads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT NOT NULL,
        apellido TEXT NOT NULL,
        telefono TEXT NOT NULL,
        email TEXT DEFAULT '',
        ciudad TEXT DEFAULT '',
        provincia TEXT DEFAULT '',
        estado TEXT DEFAULT 'nuevo',
        notas TEXT DEFAULT '',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS suscripciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER REFERENCES users(id),
        mp_preapproval_id TEXT UNIQUE,
        plan_nombre TEXT DEFAULT 'mensual',
        monto_ars REAL DEFAULT 0,
        estado TEXT DEFAULT 'pending',
        fecha_inicio DATE,
        fecha_fin DATE,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );

    -- Fix 05/07/2026: agregada también acá para instalaciones nuevas (ver
    -- misma tabla creada en migrate_db() para DBs ya existentes).
    CREATE TABLE IF NOT EXISTS password_reset_tokens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL REFERENCES users(id),
        token TEXT UNIQUE NOT NULL,
        expires_at DATETIME NOT NULL,
        used INTEGER DEFAULT 0,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    );
    """)

    # Admin user
    admin_exists = db.execute("SELECT 1 FROM users WHERE email='admin@presupuestopro.com'").fetchone()
    if not admin_exists:
        db.execute("""
            INSERT INTO users (email, password_hash, nombre, is_admin, active, pais)
            VALUES (?, ?, 'Administrador', 1, 1, 'AR')
        """, ('admin@presupuestopro.com', generate_password_hash('admin1234')))

    # Tipos de cambio iniciales
    db.executemany("""
        INSERT OR IGNORE INTO tipos_cambio (pais, moneda, simbolo, tasa)
        VALUES (?, ?, ?, ?)
    """, [
        ('AR', 'ARS', '$',   1200),
        ('CL', 'CLP', '$',    950),
        ('UY', 'UYU', '$U',    40),
        ('BR', 'BRL', 'R$',   5.2),
        ('PY', 'PYG', 'Gs',  7600),
    ])

    # Config por defecto
    db.executemany("""
        INSERT OR IGNORE INTO config (clave, valor) VALUES (?, ?)
    """, [
        ('pct_gg',         '20'),
        ('pct_impuestos',  '7'),
        ('hh_oficial_dia', '8'),
    ])

    # ── Items de obra (del Excel "Items") ──────────────────────────────────────
    # (rubro_num, rubro_nombre, nombre, unidad, precio_ars, hof, hay, orden)
    # orden = posición de display dentro del rubro (espacio de 10 para insertar)
    ITEMS = [
        # ─── 01 Preliminares ────────────────────────────────────────────────
        ('01','Preliminares','Cerco de obra',                    'ml',     40736,  0.3,   0.1,  30),
        ('01','Preliminares','Replanteo',                        'ml',     12409,  0.05,  0.02, 40),
        ('01','Preliminares','Picado de revoques',               'm2',      3270,  0.4,   0.2,  50),
        ('01','Preliminares','Rebajo de cordon',                 'ml',     30000,  0.5,   0.2,  60),
        ('01','Preliminares','Demolicion solado 20cm',           'm2',     10800,  0.3,   0.15, 70),
        ('01','Preliminares','Demolicion mamposteria',           'm3',     54000,  8.0,   4.0,  80),
        ('01','Preliminares','Demolicion hormigon armado',       'm3',    114000,  0,     4.0,  90),
        # ─── 02 Fundaciones ─────────────────────────────────────────────────
        ('02','Fundaciones','Excavacion zanjas a mano',          'm3',     23000,  4.0,   0,    10),
        ('02','Fundaciones','Excavacion bases a mano',           'm3',     30000,  5.0,   0,    20),
        ('02','Fundaciones','Excavacion de Pilotes',             'm3',     22500,  3.0,   0,    30),
        ('02','Fundaciones','Relleno y Compactacion',            'm3',     42117,  1.5,   0.5,  40),
        ('02','Fundaciones','Relleno de arena',                  'm3',     39000,  1.0,   0.5,  50),
        ('02','Fundaciones','Retiro de tierra',                  'm3',     27000,  0.5,   0.5,  60),
        ('02','Fundaciones','Zapata Ho Pobre',                   'm3',    280720,  8.0,   4.0,  80),
        ('02','Fundaciones','Mamp. ladrillo comun cimientos',    'm3',    190560, 18.0,   9.0,  90),
        ('02','Fundaciones','Mamp. ladrillo comun submuracion',  'm3',    263250, 20.0,  10.0, 100),
        # ─── 03 Hormigón Armado (Ho.Ado.) ───────────────────────────────────
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. base (50-3)',          'm3',    454314, 25.0, 12.0,  10),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. zapata (60-6)',        'm3',    568761, 30.0, 15.0,  30),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. viga fundacion (75-7)','m3',   653217, 40.0, 20.0,  40),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. columna (105-11)',     'm3',    871966, 50.0, 25.0,  50),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. viga (120-16)',        'm3',   1056095, 55.0, 27.0,  60),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. tabique 10cm (110-20)','m3',  1116076, 55.0, 27.0,  70),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. tabique 15cm (110-13)','m3',   949871, 50.0, 25.0,  80),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. tanque (90-13)',       'm3',    870487, 45.0, 22.0,  90),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. losa 10cm (80-10)',    'm3',    747818, 45.0, 22.0, 100),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. losa 15cm (75-6.66)', 'm3',    644745, 40.0, 20.0, 110),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. losa premoldeada',     'm2',     59996,  1.5,  0.7, 120),
        ('03','Hormigón Armado (Ho.Ado.)','Ho.Ado. pavimento 15cm',       'm2',     44563,  1.5,  0.7, 130),
        # ─── 04 Hormigón Elaborado (H.Elab.) ────────────────────────────────
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. base (50-3)',           'm3',    405714, 18.0,  9.0,  10),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. zapata (60-6)',         'm3',    512661, 22.0, 11.0,  30),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. viga fundacion (75-7)', 'm3',    585867, 28.0, 14.0,  40),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. columna (105-11)',      'm3',    782116, 35.0, 17.0,  50),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. viga (120-16)',         'm3',    954995, 40.0, 20.0,  60),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. tabique 10cm (110-20)', 'm3',   1022476, 38.0, 19.0,  70),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. tabique 15cm (110-13)', 'm3',    856271, 35.0, 17.0,  80),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. tanque (90-13)',        'm3',    791887, 32.0, 16.0,  90),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. losa 10cm (80-10)',     'm3',    676718, 30.0, 15.0, 100),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. losa 15cm (75-6.66)',   'm3',    577395, 28.0, 14.0, 110),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. losa premoldeada',      'm2',     52969,  1.2,  0.6, 120),
        ('04','Hormigón Elaborado (H.Elab.)','H.Elab. pavimento 15cm',        'm2',     41023,  1.2,  0.6, 130),
        # ─── 05 Cemento ─────────────────────────────────────────────────────
        ('05','Cemento','Cemento: capa aisladora s/muros',       'm2',     19052,  0.3,  0.1,  10),
        ('05','Cemento','Carpeta bajo piso',                      'm2',     10215,  0.3,  0.1,  20),
        ('05','Cemento','Cemento: carpeta azotea',                'm2',     11465, 0.25,  0.1,  30),
        ('05','Cemento','Cemento: toma de juntas',                'm2',     15505, 0.15, 0.05,  40),
        ('05','Cemento','Cemento: azotado',                       'm2',      4930, 0.15, 0.05,  50),
        ('05','Cemento','Cemento: revoque',                       'm2',     28164, 0.45,  0.2,  60),
        ('05','Cemento','Cemento: zocalo 10cm',                   'ml',      6732,  0.2, 0.05,  80),
        ('05','Cemento','Cemento: piso rodillado',                'm2',     13912,  0.3,  0.1,  90),
        ('05','Cemento','Cemento: piso alisado',                  'm2',     23642,  0.4, 0.15, 100),
        # ─── 06 Mampostería ─────────────────────────────────────────────────
        ('06','Mampostería','Mamp. ladrillo comun 15cm',          'm3',    234310, 18.0,  9.0,  10),
        ('06','Mampostería','Mamp. ladrillo comun 30cm',          'm3',    208060, 22.0, 11.0,  20),
        ('06','Mampostería','Mamp. ladrillo vista 15cm',          'm3',    356950, 22.0, 11.0,  30),
        ('06','Mampostería','Mamp. ladrillo vista 30cm',          'm3',    316950, 26.0, 13.0,  40),
        ('06','Mampostería','Tabique ladrillo comun 8cm',         'm2',     22600,  1.0,  0.5,  50),
        ('06','Mampostería','Tabique ladrillo hueco 8x18x33',     'm2',     22510,  1.0,  0.5,  60),
        ('06','Mampostería','Tabique ladrillo hueco 12x18x33',    'm2',     24950,  1.2,  0.6,  70),
        ('06','Mampostería','Tabique ladrillo hueco 18x18x33',    'm2',     31640,  1.4,  0.7,  80),
        ('06','Mampostería','Tabique hueco portante 12x19x33cm',  'm2',     32223,  0.8,  0.4,  90),
        ('06','Mampostería','Tabique hueco portante 18x19x33cm',  'm2',     38050,  0.9, 0.45, 100),
        ('06','Mampostería','Dinteles pared 15cm',                'ml',     13644,  0.5, 0.25, 110),
        ('06','Mampostería','Dinteles pared 30cm',                'ml',     21321,  0.7, 0.35, 120),
        # ─── 07 Contrapisos ─────────────────────────────────────────────────
        ('07','Contrapisos','Contrapiso cascotes 15cm',           'm2',     18655, 0.35, 0.15,  10),
        ('07','Contrapisos','Contrapiso cascotes 10cm',           'm2',     13482,  0.3, 0.12,  20),
        ('07','Contrapisos','Banquina Ho Pobre 10cm',             'm2',     13482,  0.3, 0.12,  30),
        ('07','Contrapisos','Contrapiso c/perlitas 10cm',         'm2',     18067, 0.35, 0.15,  40),
        ('07','Contrapisos','Contrapiso c/perlitas 5cm',          'm2',     11913, 0.25,  0.1,  50),
        ('07','Contrapisos','Ho de Pendiente c/perlitas 10cm',    'm2',     20249,  0.4, 0.15,  60),
        # ─── 08 Revoques ────────────────────────────────────────────────────
        ('08','Revoques','R. exterior cal (az+gr+f)',              'm2',     18050,  0.5, 0.25,  10),
        ('08','Revoques','R. interior cal (gr+f) fratas',         'm2',     14777, 0.45,  0.2,  20),
        ('08','Revoques','R. interior cal (gr+f) fieltro',        'm2',     16027,  0.5, 0.22,  30),
        ('08','Revoques','R. grueso interior b/rvto.',            'm2',      9231, 0.25,  0.1,  40),
        ('08','Revoques','Cielorraso s/losa (gr+f)',               'm2',     19943,  0.6,  0.3,  50),
        ('08','Revoques','Cielo s/metal despl. (az+gr+f)',         'm2',     22578,  0.7, 0.35,  60),
        ('08','Revoques','Estructura cielo suspendido (3-6m)',     'm2',     67291,  1.5, 0.75,  70),
        # ─── 09 Revestimientos ──────────────────────────────────────────────
        ('09','Revestimientos','Enlucido S.Iggam',                 'm2',     15000, 0.25,  0.1,  10),
        ('09','Revestimientos','Enlucido Salpicrete',              'm2',      5750,  0.2, 0.08,  20),
        ('09','Revestimientos','Revest. Texturado',                'm2',     25750,  0.3,  0.1,  30),
        # ─── 10 Techos ──────────────────────────────────────────────────────
        ('10','Techos','Techo de chapas',                          'm2',    112523,  0.5, 0.25,  10),
        # ─── 11 Instalaciones ───────────────────────────────────────────────
        ('11','Instalaciones','Colocacion de Aberturas',           'Un',     36653,  3.0,  1.5,  10),
        ('11','Instalaciones','Instalacion Desague',               'Un',    490215, 40.0, 20.0,  20),
        ('11','Instalaciones','Instalacion Agua F/C',              'Un',    554033, 30.0, 15.0,  30),
        ('11','Instalaciones','Instalacion Gas',                         'Un',    532303, 20.0, 10.0,  40),
        ('11','Instalaciones','Instalacion Electrica',             'Bca',    90347,  4.0,  2.0,  50),
        # ─── 12 Cerámicos y Porcellanatos ───────────────────────────────────
        ('12','Cerámicos y Porcellanatos','Piso ceramico azotea',              'm2',  33107,  0.7, 0.35,  10),
        ('12','Cerámicos y Porcellanatos','Piso ceramico 1',                   'm2',  21735,  0.7, 0.35,  20),
        ('12','Cerámicos y Porcellanatos','Piso ceramico 2',                   'm2',  30660, 0.75, 0.35,  30),
        ('12','Cerámicos y Porcellanatos','Piso ceramico 3 (porcellanato)',    'm2',  56660,  0.9, 0.45,  40),
        ('12','Cerámicos y Porcellanatos','Piso loseta cemento',               'm2',  35482, 0.65,  0.3,  50),
        ('12','Cerámicos y Porcellanatos','Piso calcareo vereda',              'm2',  43932,  0.6,  0.3,  60),
        ('12','Cerámicos y Porcellanatos','Zocalo ceramico 1',                 'ml',   4824,  0.3,  0.1,  70),
        ('12','Cerámicos y Porcellanatos','Zocalo ceramico 2',                 'ml',   5538,  0.3,  0.1,  80),
        ('12','Cerámicos y Porcellanatos','Zocalo ceramico 3',                 'ml',   6798, 0.35, 0.12,  90),
        ('12','Cerámicos y Porcellanatos','Zocalo de pino',                    'ml',  15529,  0.2, 0.05, 100),
        ('12','Cerámicos y Porcellanatos','Rvto. ceramico 1',                  'm2',  24080,  0.8,  0.4, 110),
        ('12','Cerámicos y Porcellanatos','Rvto. ceramico 2',                  'm2',  33430, 0.85,  0.4, 120),
        ('12','Cerámicos y Porcellanatos','Rvto. ceramico 3 porcellanato',     'm2',  59580,  1.0,  0.5, 130),
        ('12','Cerámicos y Porcellanatos','Rvto. marmol',                      'm2', 250290,  1.2,  0.6, 140),
        # ─── 13 Pintura ─────────────────────────────────────────────────────
        ('13','Pintura','Pintura cal',                             'm2',      5817, 0.12, 0.04,  10),
        ('13','Pintura','Pintura latex exterior',                  'm2',      6890, 0.15, 0.05,  20),
        ('13','Pintura','Pintura latex interior',                  'm2',      7340, 0.15, 0.05,  30),
        ('13','Pintura','Pintura latex cielos',                    'm2',      7160, 0.18, 0.06,  40),
        ('13','Pintura','Pintura especial 1',                      'm2',     10220,  0.2, 0.07,  50),
        ('13','Pintura','Pintura especial 2',                      'm2',     10850, 0.22, 0.08,  60),
        ('13','Pintura','Enduido s/paredes',                       'm2',      6875,  0.2, 0.05,  70),
        ('13','Pintura','Pintura satinol muros',                   'm2',     10850,  0.2, 0.07,  80),
        ('13','Pintura','Pintura s/carp. madera',                  'm2',     13825, 0.25, 0.05,  90),
        ('13','Pintura','Pintura s/carp. metalica',                'm2',     12825,  0.2, 0.05, 100),
        ('13','Pintura','Ayuda gremios y varios',                  'h/m2',    7500,  1.0,  0.5, 110),
    ]

    existing = db.execute("SELECT COUNT(*) as c FROM items_obra").fetchone()['c']
    if existing == 0:
        db.executemany("""
            INSERT INTO items_obra (rubro_num, rubro_nombre, nombre, unidad, precio_ars, hof, hay, orden)
            VALUES (?,?,?,?,?,?,?,?)
        """, ITEMS)

    # ── Materiales base ────────────────────────────────────────────────────────
    # Precios en USD = precio_ars / tasa_ars (tasa: 1495 — actualizado jun/2026)
    # Precios ARS de referencia confirmados:
    #   Cemento Portland 25kg $6.500 | Plasticor 25kg $8.300 | Arena gruesa m3 $32.000
    #   Piedra 6/20 m3 $70.000 | Ladrillo común millar $175.000
    #   Perlitas Telgopor 75Lts $6.500 | Ladrillo hueco 12x18x33 u $875
    materiales_base = [
        ('Cementos y aglomerantes', 'Cemento portland',            'bolsa 25kg',   4.35),
        ('Cementos y aglomerantes', 'Plasticor',                   'bolsa 25kg',   5.55),
        ('Cementos y aglomerantes', 'Cal aérea',                   'bolsa 30kg',  12.00),
        ('Cementos y aglomerantes', 'Cal hidráulica',              'bolsa 30kg',  12.00),
        ('Cementos y aglomerantes', 'Yeso',                        'bolsa 30kg',   8.00),
        ('Áridos y gravas',         'Arena gruesa',                'm3',          21.40),
        ('Áridos y gravas',         'Arena fina',                  'm3',          21.40),
        ('Áridos y gravas',         'Piedra partida 6/20',         'm3',          46.82),
        ('Áridos y gravas',         'Cascotes',                    'm3',           0.80),
        ('Áridos y gravas',         'Perlitas Telgopor (75 Lts)',  'bolsa',        4.35),
        ('Ladrillos y bloques',     'Ladrillo común 18x9x5cm',    'millar',      117.06),
        ('Ladrillos y bloques',     'Ladrillo hueco 8x18x33cm',   'u',            0.47),
        ('Ladrillos y bloques',     'Ladrillo hueco 12x18x33cm',  'u',            0.585),
        ('Ladrillos y bloques',     'Ladrillo hueco 18x18x33cm',  'u',            0.70),
        ('Ladrillos y bloques',     'Bloque portante 12x19x33cm', 'u',            3.50),
        ('Ladrillos y bloques',     'Bloque portante 18x19x33cm', 'u',            4.20),
        ('Acero y hierro',          'Hierro Ø 8mm (fy=420)',       'kg',           1.00),
        ('Acero y hierro',          'Hierro Ø 10mm (fy=420)',      'kg',           1.00),
        ('Acero y hierro',          'Hierro Ø 12mm (fy=420)',      'kg',           1.00),
        ('Acero y hierro',          'Hierro Ø 16mm (fy=420)',      'kg',           1.00),
        ('Acero y hierro',          'Hierro Ø 20mm (fy=420)',      'kg',           1.00),
        ('Acero y hierro',          'Malla electrosoldada Ø6',     'm2',           2.50),
        ('Madera encofrado',        'Tabla pino 1"x8"x3m',        'u',            3.00),
        ('Madera encofrado',        'Tirante 2"x3"x3m',           'u',            2.50),
        ('Madera encofrado',        'Fenólico 18mm',               'm2',           8.00),
        ('Impermeabilizantes',      'Aditivo Sika 1',              'lt',           0.40),
        ('Impermeabilizantes',      'Membrana asfáltica 4mm',      'm2',           1.20),
        ('Impermeabilizantes',      'Membrana líquida',            'lt',           0.30),
        ('Instalacion sanitaria',   'Caño PVC 110mm',              'm',            0.80),
        ('Instalacion sanitaria',   'Caño PVC 63mm',               'm',            0.60),
        ('Instalacion sanitaria',   'Caño PVC 40mm',               'm',            0.50),
        ('Instalacion sanitaria',   'Caño cobre 3/4"',             'm',            1.20),
        ('Instalacion sanitaria',   'Caño cobre 1/2"',             'm',            0.80),
        ('Instalacion electrica',   'Caño corrugado 3/4"',         'ml',           0.15),
        ('Instalacion electrica',   'Cajas metalicas',             'u',            0.30),
        ('Instalacion electrica',   'Cable 2.5 mm',                'ml',           0.48),
        ('Instalacion electrica',   'Cable 1.5 mm',                'ml',           0.37),
        ('Pisos y revestimientos',  'Pastina',                     'kg',           6.00),
        ('Pisos y revestimientos',  'Klaukol pegamento',           'kg',           0.54),
        ('Pisos y revestimientos',  'Revestimiento ceramico 1',    'm2',           6.50),
        ('Pisos y revestimientos',  'Revestimiento ceramico 2',    'm2',          15.00),
        ('Pisos y revestimientos',  'Porcelanato revestimiento',   'm2',          30.00),
        ('Pisos y revestimientos',  'Piso ceramico 1',             'm2',           6.50),
        ('Pisos y revestimientos',  'Piso ceramico 2',             'm2',          15.00),
        ('Pisos y revestimientos',  'Porcelanato piso',            'm2',          30.00),
        ('Pisos y revestimientos',  'Mosaico calcario 20x20cm',   'm2',          25.00),
        ('Pisos y revestimientos',  'Loseta cemento 40x60cm',     'm2',          22.00),
        ('Pisos y revestimientos',  'Baldosa ceramica',            'm2',          12.50),
        ('Pinturas',                'Pintura látex exterior',      'lt',           0.08),
        ('Pinturas',                'Pintura látex interior',      'lt',           0.07),
        ('Pinturas',                'Enduido plástico',            'kg',           0.30),
        ('Pinturas',                'Pintura cal hidráulica',      'kg',           0.40),
    ]

    mat_exist = db.execute("SELECT COUNT(*) as c FROM materiales").fetchone()['c']
    if mat_exist == 0:
        db.executemany("""
            INSERT INTO materiales (categoria, nombre, unidad, precio_usd)
            VALUES (?, ?, ?, ?)
        """, materiales_base)


    # ── Análisis sub-items (materiales + MO por item) ─────────────────────
    analisis_count = db.execute("SELECT COUNT(*) as c FROM analisis_sub").fetchone()['c']
    if analisis_count == 0:
        try:
            from utils.analisis_data import ANALISIS
            rows = []
            for item_nombre, subs in ANALISIS.items():
                for s in subs:
                    rows.append((item_nombre, s['n'], s['c'], s['p'], s['m']))
            db.executemany(
                "INSERT INTO analisis_sub (item_nombre, sub_nombre, cant_por_unit, precio_ars, es_material) VALUES (?,?,?,?,?)",
                rows
            )
            print(f"[init_db] analisis_sub: {len(rows)} filas insertadas")
        except Exception as e:
            print(f"[init_db] analisis_sub no cargado: {e}")

    # ── precio_mo_ars: fallback hof/hay si no hay Excel ────────────────
    _actualizar_mo_analisis(db)

    # ── Actualización de precios v1 (Jun-2026) ───────────────────────────
    _migrar_precios_v1(db)

    db.commit()
    db.close()
    print("Base de datos inicializada OK")


def _migrar_precios_v1(db):
    """Actualiza precio_ars en analisis_sub con precios reales del corralón (Jun-2026).
    Es idempotente: usa config 'precios_v1_ok' para no re-ejecutar si ya se aplicó."""
    ya_aplicado = db.execute(
        "SELECT valor FROM config WHERE clave='precios_v1_ok'"
    ).fetchone()
    if ya_aplicado:
        return

    PRECIOS = [
        ('Accesorios Desagues',                  8500),
        ('Accesorios Gas',                       3500),
        ('Accesorios TF',                        3000),
        ('Alambre negro',                        7000),
        ('Arena común',                         31000),
        ('Baldosa cerámica azotea',             33000),
        ('Cable 1,5 mm',                          510),
        ('Cable 2,5 mm',                          850),
        ('Cajas Metalicas',                       850),
        ('Cal Hidráulica',                        332),
        ('Cal aérea Milagro',                     372),
        ('Caño Awaduct 110',                     8125),
        ('Caño Awaduct 40',                      2900),
        ('Caño Awaduct 50',                      5300),
        ('Caño Awaduct 63',                      6250),
        ('Caño Corrugado 1"',                     376),
        ('Caño Corrugado 3/4"',                   260),
        ('Caño Epoxi 3/4',                    12890.63),
        ('Caño TF 20',                           2625),
        ('Caño TF 25',                           5000),
        ('Caño epoxi 1/2',                    6406.25),
        ('Cemento Albañilería',                   296),
        ('Cemento portland bolsas',               332),
        ('Chapas Cerco',                        14800),
        ('Chapas Techo',                        14800),
        ('Clavadores 2 x 2',                  1007.59),
        ('Clavos 2"',                            5600),
        ('Clavos 2" 1/2',                        5600),
        ('Clavos 3"',                            7200),
        ('Clavos 4"',                            7200),
        ('Color pintura cal',                   20000),
        ('Enduido sintético',                    7500),
        ('Escurridores 1/2 x 2',              459.46),
        ('Esmalte albalux',                     23175),
        ('Fondo Base',                           3800),
        ('Granza',                              36000),
        ('Hidrófugo',                            2900),
        ('Hierro redondo d=10mm',            2553.76),
        ('Hormigon elaborado colado',          190000),
        ('Issolant',                             4950),
        ('Klaukol',                               720),
        ('Ladrillo Telgopor 12*38*1m',           6200),
        ('Ladrillo hueco 12X18X33cm',             830),
        ('Ladrillo hueco 18X18X33cm',            1160),
        ('Ladrillo hueco 8x18x33cm',              720),
        ('Ladrillo hueco Portante 12x18x33cm',   1160),
        ('Ladrillo hueco Portante 18x18x33cm',   1400),
        ('Ladrillos comunes',                     230),
        ('Ladrillos vista',                       390),
        ('Llaves de Paso Agua',                 25000),
        ('Llaves de Paso Gas',                  25000),
        ('Loseta cemento 60x40cm',               5000),
        ('Martillo neumático',                  25000),
        ('Metal desplegado',                  3333.33),
        ('Mosaico calcáreo',                    25000),
        ('Palito 1"x1"',                          500),
        ('Pastina',                              5500),
        ('Perlitas Telgopor (75 Lts)',         106.67),
        ('Piedra Granítica',                    93000),
        ('Pino encofrado 1"',                8860.76),
        ('Pino tabla machimbre',                11600),
        ('Pintura especial 1',                  21250),
        ('Pintura especial 2',                  22500),
        ('Pintura látex cielos',                 7200),
        ('Pintura látex exterior',               7250),
        ('Pintura látex interior',               4750),
        ('Pintura satinol',                     22500),
        ('Piso cerámico 1',                     15000),
        ('Piso cerámico 2',                     25000),
        ('Piso cerámico 3 (porcellanato)',       35000),
        ('Rev Text.',                         4166.67),
        ('Rvto.cerámico 1',                     15000),
        ('Rvto.cerámico 2',                     25000),
        ('Rvto.cerámico 3 (porcellanato)',       35000),
        ('Saligna   1"x2"',                   1037.97),
        ('Saligna 1"x4"',                      886.08),
        ('Saligna 3"x3"',                    1766.67),
        ('Salpicrete',                        2666.67),
        ('Super Iggam',                        833.33),
        ('Tarugo 6',                               62),
        ('Tierra Colorada',                     20000),
        ('Tirantes 2x6',                     27166.67),
        ('Tornillo',                               45),
        ('Tornillo c/arand goma',                 118),
        ('Transporte material suelto',          19000),
        ('Viga Vipret 4m.',                    3862.5),
        ('Zócalo cerámico 1',                   1250),
        ('Zócalo cerámico 2',                2083.33),
        ('Zócalo cerámico 3 (Porcellanato)',  2916.67),
        ('Zócalo de madera',                    3500),
    ]

    for sub_nombre, precio in PRECIOS:
        db.execute(
            "UPDATE analisis_sub SET precio_ars=? WHERE sub_nombre=?",
            (precio, sub_nombre)
        )

    db.execute("INSERT INTO config (clave, valor) VALUES ('precios_v1_ok', '2026-06-28')")
    print(f"[migrar_precios_v1] {len(PRECIOS)} materiales actualizados")
